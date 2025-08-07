"""
Clean Rule-Based Data Extraction Tools
Fixed version with proper structure and all requested improvements
"""

import pandas as pd
import requests
import time
import yfinance as yf
from datetime import datetime, timedelta
import json
import os
import logging
from pathlib import Path
from crewai.tools import BaseTool
import numpy as np
from typing import Dict, List, Any, Optional

logger = logging.getLogger(__name__)

class SimpleRuleBasedExtractor(BaseTool):
    """
    Simple rule-based data extractor - NO LLM needed
    Returns consistent format and creates Excel files directly
    """
    
    name: str = "simple_rule_based_extractor"
    description: str = "Extract data using simple rules and create Excel files"
    def _run(self, validation_results: str) -> str:
        """Extract data and create Excel files directly"""
        try:
            print("SIMPLE RULE-BASED EXTRACTION")
            print("=" * 50)
            
            # Parse validation results
            sources = self._parse_validation_results(validation_results)
            if not sources:
                return json.dumps({"error": "No valid sources found"})
            
            print(f"Processing {len(sources)} sources...")
            
            # Create output directory
            output_dir = Path("extracted_data")
            output_dir.mkdir(exist_ok=True)
            print(f"Output directory: {output_dir.absolute()}")
            
            # Extract data by category
            macro_data = {}
            market_data = []
            files_created = []
            
            # Load FRED API key
            fred_api_key = os.getenv('FRED_API_KEY')
            
            for i, source in enumerate(sources, 1):
                print(f"\n[{i}/{len(sources)}] {source['name']}")
                
                # Determine category using simple rules
                if self._is_market_source(source['name']):
                    # Market data
                    result = self._extract_market_data(source)
                    if result:
                        market_data.append(result)
                        print(f"   ✅ Market: {result['records']} records")
                    else:
                        print(f"   ❌ Market extraction failed")
                else:
                    # Macro data
                    result = self._extract_macro_data(source, fred_api_key)
                    if result:
                        country = result['country']
                        data_type = result['data_type']
                        
                        if country not in macro_data:
                            macro_data[country] = {}
                        macro_data[country][data_type] = result['data']
                        
                        print(f"   ✅ Macro: {result['records']} records ({country}/{data_type})")
                    else:
                        print(f"   ❌ Macro extraction failed")
                
                time.sleep(0.5)  # Be respectful
            
            # Create Excel files
            if macro_data:
                for country, country_data in macro_data.items():
                    filename = self._create_macro_excel(country, country_data, output_dir)
                    if filename:
                        files_created.append(filename)
            
            if market_data:
                filename = self._create_market_excel(market_data, output_dir)
                if filename:
                    files_created.append(filename)
            
            # Return simple summary
            summary = {
                "status": "completed",
                "files_created": files_created,
                "macro_countries": len(macro_data),
                "market_indices": len(market_data),
                "total_files": len(files_created)
            }
            
            print(f"\n✅ EXTRACTION COMPLETED")
            print(f"   Files created: {len(files_created)}")
            for file in files_created:
                print(f"   • {Path(file).name}")
            
            return json.dumps(summary, indent=2)
            
        except Exception as e:
            error_msg = f"Extraction failed: {str(e)}"
            print(f"❌ {error_msg}")
            import traceback
            traceback.print_exc()
            return json.dumps({"error": error_msg})
    
    def _parse_validation_results(self, validation_results: str) -> List[Dict[str, str]]:
        """Parse validation results into list of sources"""
        sources = []
        
        try:
            lines = validation_results.strip().split('\n')
            current_source = {}
            
            for line in lines:
                line = line.strip()
                if line.startswith("Source:"):
                    if current_source:
                        sources.append(current_source)
                    current_source = {"name": line.replace("Source:", "").strip()}
                elif line.startswith("Method:"):
                    current_source["method"] = line.replace("Method:", "").strip()
                elif line.startswith("Status:"):
                    current_source["status"] = line.replace("Status:", "").strip()
                elif line.startswith("URL:"):
                    current_source["url"] = line.replace("URL:", "").strip()
            
            if current_source:
                sources.append(current_source)
            
            # Filter only working sources
            working_sources = [s for s in sources if s.get("status") == "working"]
            return working_sources
            
        except Exception as e:
            print(f"Error parsing validation: {e}")
            return []
    
    def _is_market_source(self, source_name: str) -> bool:
        """Check if source is in MARKET_INDICES_SOURCES - scalable approach"""
        try:
            from macro_sources import MARKET_INDICES_SOURCES
            
            # Check if source name matches any in MARKET_INDICES_SOURCES
            for market_source in MARKET_INDICES_SOURCES:
                if market_source.name == source_name:
                    return True
                # Also check partial matches for robustness
                if source_name.lower() in market_source.name.lower() or market_source.name.lower() in source_name.lower():
                    return True
            
            # If not found in MARKET_INDICES_SOURCES, it's macro data
            return False
            
        except ImportError:
            print(f"   Warning: Could not import MARKET_INDICES_SOURCES, using fallback")
            # Fallback to basic keywords if import fails
            market_keywords = ["s&p 500", "nasdaq", "nikkei", "hang seng", "sti", "vix"]
            name_lower = source_name.lower()
            return any(keyword in name_lower for keyword in market_keywords)
    
    def _extract_macro_data(self, source: Dict[str, str], fred_api_key: str) -> Optional[Dict[str, Any]]:
        """Extract macro data using universal methods + country-specific methods"""
        try:
            url = source.get("url", "")
            name = source.get("name", "")
            
            # Find the source object
            macro_source = self._find_macro_source(name)
            if not macro_source:
                print(f"   Warning: Could not find source object for {name}")
                return None
            
            # UNIVERSAL METHODS FIRST
            if "fred.stlouisfed.org" in url: # FRED API
                df = self._extract_fred_api(macro_source, fred_api_key)
            elif "data.imf.org" in url or "api.imf.org" in url: # IMF API
                df = self._extract_imf_api(macro_source)
            elif hasattr(macro_source, 'wb_series_id') and macro_source.wb_series_id:  # World Bank API
                df = self._extract_world_bank_data(macro_source)
            elif macro_source.source_type == 'excel_download': # Excel/CSV download
                df = self._extract_excel_download(macro_source)
            
            # COUNTRY-SPECIFIC METHODS
            elif "data.gov.sg" in url: # Singapore API
                df = self._extract_singapore_api(macro_source)
            elif "data.gov.my" in url: # Malaysia API
                df = self._extract_malaysia_api(macro_source)
            elif "financialmarkets.bnm.gov.my" in url: # Malaysia BNM
                df = self._extract_malaysia_bnm(macro_source)
            
            else:
                print(f"   ❌ Unknown source type or URL pattern")
                df = None
            
            if df is not None and len(df) > 0:
                return {
                    'country': macro_source.country,
                    'data_type': macro_source.data_type,
                    'data': df,
                    'records': len(df),
                    'source_name': name
                }
            
            return None
            
        except Exception as e:
            print(f"   Error extracting macro data: {e}")
            return None
    
    def _extract_singapore_api(self, source) -> Optional[pd.DataFrame]:
        """Extract from Singapore API with clean date formatting and enhanced metadata"""
        try:
            response = requests.get(source.api_url + "&limit=1000", timeout=30)
            response.raise_for_status()
            
            data = response.json()
            if 'result' not in data or 'records' not in data['result']:
                return None
            
            records = data['result']['records']
            if not records:
                return None
            
            df = pd.DataFrame(records)
            
            # Enhanced date filtering with clean formatting
            date_cols = [col for col in df.columns if 'date' in col.lower() or 'quarter' in col.lower() or 'period' in col.lower()]
            if date_cols:
                try:
                    import warnings
                    # Suppress the pandas date parsing warning
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore", UserWarning)
                        df[date_cols[0]] = pd.to_datetime(df[date_cols[0]], errors='coerce')
                    
                    # Format date as clean YYYY-MM-DD (remove 00:00:00 time component)
                    df[date_cols[0]] = df[date_cols[0]].dt.strftime('%Y-%m-%d')
                    
                    # Filter for data from 2000 onwards (now works with string dates)
                    df = df[df[date_cols[0]] >= '2000-01-01']
                except Exception as e:
                    print(f"   Date formatting warning: {e}")
            
            # Clean up quarterly data format if present
            quarter_cols = [col for col in df.columns if 'quarter' in col.lower()]
            for quarter_col in quarter_cols:
                if quarter_col in df.columns:
                    try:
                        # If it's datetime, format it cleanly
                        if pd.api.types.is_datetime64_any_dtype(df[quarter_col]):
                            df[quarter_col] = pd.to_datetime(df[quarter_col]).dt.strftime('%Y-%m-%d')
                    except:
                        pass
            
            # Add enhanced metadata
            df['source_name'] = source.name
            df['extraction_time'] = datetime.now().isoformat()
            df['data_type'] = source.data_type
            df['country'] = source.country
            
            # Add currency information for monetary data
            if source.data_type in ['GDP', 'Property_Price']:
                df['currency'] = 'SGD'
                df['unit'] = 'Chained 2015 Dollars' if 'GDP' in source.data_type else 'Index (Base Quarter 2009-Q1 = 100)'
            elif 'CPI' in source.data_type:
                df['unit'] = 'Index (2019=100)'
            elif 'Interest' in source.data_type:
                df['unit'] = 'Percent per annum'
            elif 'Population' in source.data_type:
                df['unit'] = 'Number of persons'
            
            return df
            
        except Exception as e:
            print(f"   Singapore API error: {e}")
            return None
    
    def _extract_fred_api(self, source, fred_api_key: str) -> Optional[pd.DataFrame]:
        """Extract from FRED API with enhanced data cleaning and detailed removal logging"""
        if not fred_api_key:
            return None
        
        try:
            # First, get series metadata to extract real units
            series_id = getattr(source, 'fred_series_id', 'Unknown')
            if series_id == 'Unknown':
                print(f"   ❌ No FRED series ID found")
                return None
            
            # Get series info for units
            series_info_url = f"https://api.stlouisfed.org/fred/series?series_id={series_id}&api_key={fred_api_key}&file_type=json"
            
            print(f"   Getting series info from FRED API...")
            series_response = requests.get(series_info_url, timeout=30)
            series_response.raise_for_status()
            series_data = series_response.json()
            
            # Extract real units from FRED
            real_units = "Unknown"
            if 'seriess' in series_data and len(series_data['seriess']) > 0:
                series_info = series_data['seriess'][0]
                real_units = series_info.get('units', 'Unknown')
                print(f"   FRED Units: {real_units}")
            
            # Now get the actual data
            api_url = source.get_fred_api_url()
            if not api_url:
                print(f"   ❌ Could not construct FRED API URL")
                return None
            
            print(f"   Getting observations from FRED API...")
            response = requests.get(api_url, timeout=30)
            response.raise_for_status()
            
            data = response.json()
            if 'observations' not in data:
                print(f"   ❌ No observations in FRED response")
                return None
            
            observations = data['observations']
            print(f"   Raw observations from FRED: {len(observations)}")
            
            df = pd.DataFrame(observations)
            print(f"   Initial DataFrame shape: {df.shape}")
            print(f"   Sample of first 5 observations:")
            print(f"   {df[['date', 'value']].head().to_dict('records')}")
            
            # Enhanced data cleaning with detailed logging
            original_count = len(df)
            removed_records = {
                'invalid_dates': [],
                'missing_values': [],
                'non_numeric': [],
                'negative_population': []
            }
            
            # Step 1: Convert dates and track invalid ones
            print(f"\n   📋 DATA CLEANING DETAILS:")
            df['date'] = pd.to_datetime(df['date'], errors='coerce')
            
            # Track records with invalid dates
            invalid_date_mask = df['date'].isna()
            if invalid_date_mask.any():
                invalid_date_records = df[invalid_date_mask][['date', 'value']].to_dict('records')
                removed_records['invalid_dates'] = invalid_date_records
                print(f"   ❌ Removing {len(invalid_date_records)} records with invalid dates:")
                for record in invalid_date_records[:3]:  # Show first 3
                    print(f"      • Date: {record['date']}, Value: {record['value']}")
                if len(invalid_date_records) > 3:
                    print(f"      • ... and {len(invalid_date_records) - 3} more")
            
            df = df.dropna(subset=['date'])
            print(f"   ✅ After date cleaning: {len(df)} records remaining")
            
            # Step 2: Handle FRED's missing value indicators
            print(f"   📊 Value column analysis:")
            unique_values = df['value'].unique()
            print(f"      Unique values sample: {list(unique_values[:10])}")
            
            # Track different types of missing values
            missing_patterns = ['.', '', 'NaN']
            for pattern in missing_patterns:
                pattern_mask = df['value'] == pattern
                if pattern_mask.any():
                    pattern_records = df[pattern_mask][['date', 'value']].to_dict('records')
                    removed_records['missing_values'].extend(pattern_records)
                    print(f"   ❌ Removing {len(pattern_records)} records with '{pattern}' values:")
                    for record in pattern_records[:2]:  # Show first 2
                        print(f"      • Date: {record['date']}, Value: '{record['value']}'")
                    if len(pattern_records) > 2:
                        print(f"      • ... and {len(pattern_records) - 2} more")
            
            # Remove NaN values
            nan_mask = df['value'].isna()
            if nan_mask.any():
                nan_records = df[nan_mask][['date', 'value']].to_dict('records')
                removed_records['missing_values'].extend(nan_records)
                print(f"   ❌ Removing {len(nan_records)} records with NaN values")
            
            # Apply all missing value filters
            df = df[df['value'].notna()].copy()
            df = df[df['value'] != '.'].copy()
            df = df[df['value'] != ''].copy()
            df = df[df['value'] != 'NaN'].copy()
            
            print(f"   ✅ After missing value removal: {len(df)} records remaining")
            
            # Step 3: Convert to numeric and track conversion failures
            before_numeric = len(df)
            df['value'] = pd.to_numeric(df['value'], errors='coerce')
            
            # Track records that couldn't be converted to numeric
            numeric_nan_mask = df['value'].isna()
            if numeric_nan_mask.any():
                non_numeric_records = df[numeric_nan_mask][['date', 'value']].to_dict('records')
                removed_records['non_numeric'] = non_numeric_records
                print(f"   ❌ Removing {len(non_numeric_records)} records that couldn't convert to numeric:")
                for record in non_numeric_records[:2]:
                    print(f"      • Date: {record['date']}, Value: {record['value']}")
            
            df = df.dropna(subset=['value'])
            print(f"   ✅ After numeric conversion: {len(df)} records remaining")
            
            # Step 4: Domain-specific validation (population only)
            if 'Population' in source.data_type:
                negative_mask = df['value'] <= 0
                if negative_mask.any():
                    negative_records = df[negative_mask][['date', 'value']].to_dict('records')
                    removed_records['negative_population'] = negative_records
                    print(f"   ❌ Removing {len(negative_records)} records with non-positive population:")
                    for record in negative_records[:2]:
                        print(f"      • Date: {record['date']}, Value: {record['value']}")
                    
                    df = df[df['value'] > 0]
                    print(f"   ✅ After population validation: {len(df)} records remaining")
            
            # Step 5: Sort by date
            df = df.sort_values('date')
            
            # Summary of cleaning
            total_removed = original_count - len(df)
            print(f"\n   📊 CLEANING SUMMARY:")
            print(f"      Original records: {original_count}")
            print(f"      Final records: {len(df)}")
            print(f"      Total removed: {total_removed} ({total_removed/original_count*100:.1f}%)")
            
            # Breakdown of removals
            if total_removed > 0:
                print(f"      Breakdown:")
                if removed_records['invalid_dates']:
                    print(f"        • Invalid dates: {len(removed_records['invalid_dates'])}")
                if removed_records['missing_values']:
                    print(f"        • Missing values: {len(removed_records['missing_values'])}")
                if removed_records['non_numeric']:
                    print(f"        • Non-numeric: {len(removed_records['non_numeric'])}")
                if removed_records['negative_population']:
                    print(f"        • Negative population: {len(removed_records['negative_population'])}")
            
            if len(df) == 0:
                print(f"   ❌ No valid data after cleaning!")
                return None
            
            # Format date as clean YYYY-MM-DD (remove time component)
            df['date'] = df['date'].dt.strftime('%Y-%m-%d')
            
            # Show final date range
            print(f"   ✅ Final data range: {df['date'].iloc[0]} to {df['date'].iloc[-1]}")
            print(f"   ✅ Final record count: {len(df)}")
            
            # Create standardized DataFrame with enhanced metadata
            standardized_df = pd.DataFrame()
            
            # Required columns in specific order
            standardized_df['source_name'] = [source.name] * len(df)
            standardized_df['series_id'] = [series_id] * len(df)
            standardized_df['country'] = [source.country] * len(df)
            standardized_df['data_type'] = [source.data_type] * len(df)
            standardized_df['date'] = df['date'].values
            standardized_df['value'] = df['value'].values
            
            # Use REAL units from FRED API
            standardized_df['unit'] = [real_units] * len(df)
            
            # Add currency for monetary data
            if source.data_type in ['GDP', 'Property_Price']:
                if source.country == 'US':
                    standardized_df['currency'] = ['USD'] * len(df)
                elif source.country == 'Euro Area':
                    standardized_df['currency'] = ['EUR'] * len(df)
                elif source.country == 'Japan':
                    standardized_df['currency'] = ['JPY'] * len(df)
            
            # Add cleaning metadata
            standardized_df['records_removed'] = [total_removed] * len(df)
            standardized_df['cleaning_applied'] = [f"Date validation, missing value removal, numeric conversion{', population validation' if 'Population' in source.data_type else ''}"] * len(df)
            
            # Final column: extraction time
            standardized_df['extraction_time'] = [datetime.now().isoformat()] * len(df)
            
            print(f"   ✅ Standardized {len(standardized_df)} records with real units: {real_units}")
            return standardized_df
        
        except Exception as e:
            print(f"   ❌ FRED API error: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _extract_malaysia_api(self, source) -> Optional[pd.DataFrame]:
        """Extract from Malaysia Government API"""
        try:
            from malaysia_data_extractor import MalaysiaDataExtractor
            
            # Use the Malaysia extractor for specific source
            extractor = MalaysiaDataExtractor()
            
            # Get source info
            source_info = {
                'name': source.name,
                'url': source.url,
                'api_url': source.api_url,
                'source_type': source.source_type,
                'authority': source.authority,
                'description': source.description
            }
            
            # Extract based on data type
            if source.data_type == 'GDP':
                df = extractor._extract_malaysia_gov_api(source_info, 'GDP')
            elif source.data_type == 'CPI':
                df = extractor._extract_malaysia_gov_api(source_info, 'CPI')
            elif source.data_type == 'Population':
                df = extractor._extract_malaysia_gov_api(source_info, 'Population')
            else:
                df = None
            
            return df
            
        except Exception as e:
            print(f"   Malaysia API error: {e}")
            return None
    
    def _extract_malaysia_bnm(self, source) -> Optional[pd.DataFrame]:
        """Extract from Bank Negara Malaysia website"""
        try:
            from malaysia_data_extractor import MalaysiaDataExtractor
            
            extractor = MalaysiaDataExtractor()
            
            source_info = {
                'name': source.name,
                'url': source.url,
                'api_url': source.url,  # Use URL as api_url for web scraping
                'source_type': 'web_scraping',
                'authority': source.authority,
                'description': source.description
            }
            
            df = extractor._extract_web_scraping(source_info, 'Interest_Rate')
            return df
            
        except Exception as e:
            print(f"   Malaysia BNM error: {e}")
            return None
    
    def _extract_market_data(self, source: Dict[str, str]) -> Optional[Dict[str, Any]]:
        """Extract market data using yfinance - updated for 2020+ with last trading day"""
        try:
            name = source.get("name", "")
            symbol = self._get_market_symbol(name)
            
            if not symbol:
                print(f"   No symbol found for {name}")
                return None
            
            ticker = yf.Ticker(symbol)
            
            # Get daily data since January 1, 2020
            start_date = "2020-01-01"
            end_date = datetime.now().strftime('%Y-%m-%d')
            
            print(f"   Fetching daily data from {start_date} to {end_date}")
            
            daily_data = ticker.history(
                start=start_date,
                end=end_date
            )
            
            if daily_data.empty:
                return None
            
            # # Convert to monthly (last trading day of month, not calendar day)
            # daily_data_reset = daily_data.reset_index()
            # daily_data_reset['YearMonth'] = daily_data_reset['Date'].dt.to_period('M')
            
            # # Get the last trading day for each month (actual last day with data)
            # monthly_data = daily_data_reset.groupby('YearMonth').tail(1)
            
            # Create clean DataFrame
            # data = monthly_data[['Date', 'Close']].copy()
            data = daily_data.reset_index()[['Date', 'Close', 'Volume']].copy()
            data.reset_index(drop=True, inplace=True)
            
            # Format Date as clean YYYY-MM-DD (remove time component)
            data['Date'] = data['Date'].dt.strftime('%Y-%m-%d')
            
            # Add metadata in specific order
            data.insert(0, 'Index_Name', name)
            data.insert(1, 'Symbol', symbol)
            data.insert(2, 'Country', self._get_country_from_name(name))
            
            # Final column order: Index_Name, Symbol, Country, Date, Close, Volume
            cols = ['Index_Name', 'Symbol', 'Country', 'Date', 'Close', 'Volume']
            data = data[cols]
            
            return {
                'name': name,
                'symbol': symbol,
                'country': self._get_country_from_name(name),
                'data': data,
                'records': len(data)
            }
            
        except Exception as e:
            print(f"   Market extraction error: {e}")
            return None
    
    def _get_market_symbol(self, name: str) -> Optional[str]:
        """Get yfinance symbol for market index"""
        symbol_map = {
            "s&p 500": "^GSPC",
            "nasdaq": "^IXIC", 
            "nikkei": "^N225",
            "hang seng": "^HSI",
            "euro stoxx": "^STOXX50E",
            "shanghai": "000001.SS",
            "sti": "^STI",
            "straits times": "^STI",
            "vix": "^VIX",
            "ftse 100": "^FTSE",
            "sensex": "^BSESN",
            "bse sensex": "^BSESN",
            "jakarta": "^JKSE",
            "jkse": "^JKSE",
            "klci": "^KLSE",
            "bursa": "^KLSE",
            "set": "^SET.BK",
            "set index": "^SET.BK",
            "vn-index": "VNINDEX.VN",
            "vnindex": "VNINDEX.VN",
        }
        
        name_lower = name.lower()
        for key, symbol in symbol_map.items():
            if key in name_lower:
                return symbol
        
        return None
    
    def _get_country_from_name(self, name: str) -> str:
        """Get country from index name"""
        name_lower = name.lower()
        
        if any(word in name_lower for word in ["singapore", "sti", "straits"]):
            return "Singapore"
        elif any(word in name_lower for word in ["s&p", "nasdaq", "vix"]):
            return "US"
        elif "euro" in name_lower:
            return "EU"
        elif "nikkei" in name_lower:
            return "Japan"
        elif "shanghai" in name_lower:
            return "China"
        elif "hang seng" in name_lower:
            return "Hong Kong"
        elif "ftse 100" in name_lower:
            return "UK"
        elif any(word in name_lower for word in ["sensex", "bse"]):
            return "India"
        elif any(word in name_lower for word in ["jakarta", "jkse"]):
            return "Indonesia"
        elif any(word in name_lower for word in ["klci", "bursa", "malaysia"]):
            return "Malaysia"
        elif any(word in name_lower for word in ["set", "thailand"]):
            return "Thailand"
        elif any(word in name_lower for word in ["vn-index", "vnindex", "vietnam"]):
            return "Vietnam"
        else:
            return "Global"
    
    def _find_macro_source(self, name: str):
        """Find macro source object"""
        try:
            from macro_sources import (
            SINGAPORE_SOURCES, CHINA_SOURCES, US_SOURCES, EU_SOURCES, JAPAN_SOURCES,
            UK_SOURCES, INDIA_SOURCES, INDONESIA_SOURCES, MALAYSIA_SOURCES, THAILAND_SOURCES, VIETNAM_SOURCES
        )
            
            all_sources = (SINGAPORE_SOURCES + CHINA_SOURCES + US_SOURCES + EU_SOURCES + 
                           JAPAN_SOURCES + UK_SOURCES + INDIA_SOURCES + 
                           INDONESIA_SOURCES + MALAYSIA_SOURCES + THAILAND_SOURCES + VIETNAM_SOURCES)
            
            for source in all_sources:
                if source.name == name:
                    return source
            
            return None
        except:
            return None
    
    def _get_source_url_by_name(self, source_name: str) -> str:
        """Get source URL by matching source name - FIXED for FRED sources"""
        try:
            # Import all sources and find matching URL
            from macro_sources import (
            SINGAPORE_SOURCES, CHINA_SOURCES, US_SOURCES, EU_SOURCES, JAPAN_SOURCES,
            UK_SOURCES, INDIA_SOURCES, INDONESIA_SOURCES, THAILAND_SOURCES, 
            MARKET_INDICES_SOURCES
            )
            all_sources = (SINGAPORE_SOURCES + CHINA_SOURCES + US_SOURCES + EU_SOURCES + 
                           JAPAN_SOURCES + UK_SOURCES + INDIA_SOURCES + 
                           INDONESIA_SOURCES + THAILAND_SOURCES + 
                           MARKET_INDICES_SOURCES)
            # First try exact match
            for source in all_sources:
                if source.name == source_name:
                    print(f"   Found exact URL match for: {source_name}")
                    return source.url
            
            # Then try partial match for FRED sources (in case of slight name differences)
            for source in all_sources:
                if (source_name.lower() in source.name.lower() or 
                    source.name.lower() in source_name.lower()):
                    print(f"   Found partial URL match: {source.name} <-> {source_name}")
                    return source.url
            
            print(f"   No URL match found for: {source_name}")
            return "URL not found"
        except Exception as e:
            print(f"   Error finding URL: {e}")
            return "N/A"
    
    def _extract_excel_download(self, source) -> Optional[pd.DataFrame]:
        """
        FIXED: Extract data from Excel/CSV download with proper format detection
        Uses self.session for HTTP requests
        """
        try:
            print(f"   🔄 File Download (Fixed): {source.api_url}")
            
            response = requests.get(source.api_url, timeout=60)
            if response.status_code != 200:
                print(f"   ❌ Download failed: {response.status_code}")
                return None
            
            # Check content type
            content_type = response.headers.get('content-type', '').lower()
            print(f"   📄 Content type: {content_type}")
            
            # FIXED: Determine if it's CSV or Excel based on content
            content_preview = response.content[:1000].decode('utf-8', errors='ignore')
            is_csv = (',' in content_preview and '\n' in content_preview and 
                     not content_preview.startswith('PK'))  # Excel files start with PK
            
            # Save to temporary file with appropriate extension
            import tempfile
            import os
            
            if is_csv:
                temp_suffix = '.csv'
                print(f"   📊 Detected CSV format")
            else:
                temp_suffix = '.xlsx'
                print(f"   📊 Detected Excel format")
            
            with tempfile.NamedTemporaryFile(delete=False, suffix=temp_suffix) as temp_file:
                temp_file.write(response.content)
                temp_file_path = temp_file.name
            
            try:
                # FIXED: Read based on actual format
                if is_csv:
                    # Try different encodings for CSV
                    for encoding in ['utf-8', 'iso-8859-1', 'cp1252']:
                        try:
                            df = pd.read_csv(temp_file_path, encoding=encoding)
                            print(f"   📊 Parsed as CSV ({encoding}): {len(df)} rows, {len(df.columns)} columns")
                            break
                        except UnicodeDecodeError:
                            continue
                        except Exception as e:
                            if encoding == 'cp1252':  # Last attempt
                                print(f"   ❌ CSV parsing failed: {e}")
                                return None
                            continue
                else:
                    # Excel format
                    try:
                        df = pd.read_excel(temp_file_path, engine='openpyxl')
                        print(f"   📊 Parsed as Excel: {len(df)} rows, {len(df.columns)} columns")
                    except Exception as e:
                        print(f"   ❌ Excel parsing failed: {e}")
                        return None
                
                if df is None or len(df) == 0:
                    print(f"   ❌ Could not parse downloaded file")
                    return None
                
                # Show column info for debugging
                print(f"   📋 Columns: {list(df.columns)}")
                
                # FIXED: Enhanced date and value column detection for Thailand THOR data
                date_col = None
                value_col = None
                
                for col in df.columns:
                    col_str = str(col).lower().strip()
                    
                    # Look for date columns
                    if col_str in ['as of', 'date', 'วันที่']:
                        date_col = col
                        print(f"   📅 Found date column: {col}")
                        break
                    elif any(keyword in col_str for keyword in ['date', 'day', 'time']):
                        if date_col is None:
                            date_col = col
                            print(f"   📅 Found potential date column: {col}")
                
                for col in df.columns:
                    col_str = str(col).lower().strip()
                    
                    # Look for rate/value columns
                    if col_str in ['rate', 'อัตรา']:
                        value_col = col
                        print(f"   💰 Found value column: {col}")
                        break
                    elif any(keyword in col_str for keyword in ['value', 'price']) and pd.api.types.is_numeric_dtype(df[col]):
                        if value_col is None:
                            value_col = col
                            print(f"   💰 Found potential value column: {col}")
                
                # Filter to THOR O/N rates if applicable
                if 'Code' in df.columns and 'Tenor' in df.columns:
                    thor_mask = (df['Code'] == 'THOR') & (df['Tenor'] == 'O/N')
                    df_filtered = df[thor_mask].copy()
                    print(f"   🔍 Filtered to THOR O/N rates: {len(df)} → {len(df_filtered)} records")
                    df = df_filtered
                
                if date_col is not None and value_col is not None and len(df) > 0:
                    # Clean and format data
                    clean_df = df[[date_col, value_col]].copy()
                    clean_df.columns = ['date', 'value']
                    
                    # Remove missing values
                    clean_df = clean_df.dropna()
                    clean_df['value'] = pd.to_numeric(clean_df['value'], errors='coerce')
                    clean_df = clean_df.dropna()
                    
                    # Date conversion
                    print(f"   📅 Sample date values: {clean_df['date'].head(3).tolist()}")
                    
                    date_converted = False
                    for date_format in ['%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d', '%Y/%m/%d']:
                        try:
                            clean_df['date'] = pd.to_datetime(clean_df['date'], format=date_format)
                            print(f"   ✅ Date format matched: {date_format}")
                            date_converted = True
                            break
                        except:
                            continue
                    
                    if not date_converted:
                        try:
                            clean_df['date'] = pd.to_datetime(clean_df['date'], errors='coerce')
                            date_converted = True
                            print(f"   ✅ Automatic date parsing succeeded")
                        except:
                            pass
                    
                    if not date_converted:
                        print(f"   ❌ Could not parse dates")
                        return None
                    
                    clean_df = clean_df.dropna()
                    
                    if len(clean_df) == 0:
                        print(f"   ❌ No valid records after date conversion")
                        return None
                    
                    # FIXED: Add proper metadata to match other extraction methods
                    clean_df['source_name'] = source.name
                    clean_df['country'] = source.country
                    clean_df['data_type'] = source.data_type
                    clean_df['unit'] = 'Percent per annum'
                    clean_df['source'] = source.authority
                    clean_df['extraction_time'] = datetime.now().isoformat()
                    
                    # Sort by date and format
                    clean_df = clean_df.sort_values('date')
                    clean_df['date'] = clean_df['date'].dt.strftime('%Y-%m-%d')
                    
                    print(f"   ✅ File extraction successful: {len(clean_df)} records")
                    print(f"   📅 Date range: {clean_df['date'].min()} to {clean_df['date'].max()}")
                    
                    return clean_df
                else:
                    print(f"   ❌ Could not identify date/value columns or no data")
                    return None
                    
            finally:
                # Clean up temp file
                if os.path.exists(temp_file_path):
                    os.unlink(temp_file_path)
                    
        except Exception as e:
            print(f"   ❌ File download error: {e}")
            import traceback
            traceback.print_exc()
            return None

    # Extraction method for IMF API
    def _extract_imf_api(self, source) -> Optional[pd.DataFrame]:
        """Extract data from IMF API (China GDP, Thailand GDP, CPI)"""
        try:
            import urllib.request
            import json
            import ssl
            
            print(f"   🔄 IMF API: {source.api_url}")
            
            # CREATE SSL CONTEXT THAT ALLOWS UNVERIFIED CERTIFICATES
            ssl_context = ssl.create_default_context()
            ssl_context.check_hostname = False
            ssl_context.verify_mode = ssl.CERT_NONE

            headers = {
                'Cache-Control': 'no-cache',
                'User-Agent': 'Economic-Data-Pipeline/1.0'
            }
            
            req = urllib.request.Request(source.api_url, headers=headers)
            response = urllib.request.urlopen(req, timeout=30, context=ssl_context)
            
            if response.getcode() != 200:
                print(f"   ❌ IMF API request failed: {response.getcode()}")
                return None
            
            data = json.loads(response.read().decode('utf-8'))
            
            # Navigate to the data structure
            if 'data' not in data or 'dataSets' not in data['data']:
                print(f"   ❌ Invalid IMF response structure")
                return None
                
            dataset = data['data']['dataSets'][0]
            series_data = dataset.get('series', {})
            
            # Extract time periods from structure
            time_values = self._extract_imf_time_periods(data)
            
            if not time_values:
                print(f"   ❌ Could not extract time periods from IMF response")
                return None
            
            # Process series data
            all_records = []
            
            for series_key, series_info in series_data.items():
                observations = series_info.get('observations', {})
                
                for obs_key, obs_data in observations.items():
                    try:
                        time_index = int(obs_key)
                        
                        if isinstance(obs_data, list) and len(obs_data) > 0:
                            value = obs_data[0]
                            
                            if value is not None:
                                numeric_value = float(value)
                                
                                if time_index < len(time_values):
                                    time_period = time_values[time_index]
                                    
                                    # Convert time format based on data type
                                    if source.data_type == 'GDP':
                                        date_str = self._convert_quarterly_to_date(time_period)
                                    else:  # CPI
                                        date_str = self._convert_imf_monthly_to_date(time_period)
                                    
                                    if date_str:
                                        all_records.append({
                                            'date': date_str,
                                            'value': numeric_value,
                                            'time_period': time_period,
                                            'country': source.country,
                                            'data_type': source.data_type,
                                            'source_name': source.name, 
                                            'source': source.authority,
                                            'extraction_time': pd.Timestamp.now().isoformat()
                                        })
                    
                    except (ValueError, IndexError, TypeError):
                        continue
            
            if all_records:
                df = pd.DataFrame(all_records)
                df['date'] = pd.to_datetime(df['date'])
                df = df.sort_values('date')
                df = df.drop_duplicates(subset=['date'], keep='last')
                df['date'] = df['date'].dt.strftime('%Y-%m-%d')
                
                if 'source_name' not in df.columns and 'source' in df.columns:
                    df['source_name'] = df['source']

                print(f"   ✅ IMF extraction successful: {len(df)} records")
                return df
            else:
                print(f"   ❌ No valid records extracted from IMF API")
                return None
                
        except Exception as e:
            print(f"   ❌ IMF API error: {e}")
            return None

    # Helper methods for IMF data processing
    def _extract_imf_time_periods(self, data: dict) -> list:
        """Extract time periods from IMF API response structure"""
        try:
            if 'data' in data and 'structures' in data['data'] and len(data['data']['structures']) > 0:
                structure = data['data']['structures'][0]
                dimensions = structure.get('dimensions', {})
                for dim in dimensions.get('observation', []):
                    if dim.get('id') == 'TIME_PERIOD':
                        values = dim.get('values', [])
                        time_periods = [v['value'] for v in values if isinstance(v, dict) and 'value' in v]
                        if time_periods:
                            return time_periods
            return []
        except Exception:
            return []

    def _convert_quarterly_to_date(self, time_period: str) -> Optional[str]:
        """Convert quarterly time period to date (2024-Q1 -> 2024-01-01)"""
        try:
            if not time_period or '-Q' not in time_period:
                return None
            
            parts = time_period.split('-Q')
            if len(parts) != 2:
                return None
            
            year = parts[0]
            quarter = parts[1]
            
            quarter_months = {'1': '01', '2': '04', '3': '07', '4': '10'}
            month = quarter_months.get(quarter)
            
            if month and len(year) == 4 and year.isdigit():
                return f"{year}-{month}-01"
            
            return None
        except Exception:
            return None

    def _convert_imf_monthly_to_date(self, time_period: str) -> Optional[str]:
        """Convert IMF monthly time period to date (2010-M01 -> 2010-01-01)"""
        try:
            if not time_period or '-M' not in time_period:
                return None
            
            parts = time_period.split('-M')
            if len(parts) != 2:
                return None
            
            year = parts[0]
            month = parts[1].zfill(2)
            
            if len(year) == 4 and year.isdigit() and len(month) == 2 and month.isdigit():
                month_int = int(month)
                if 1 <= month_int <= 12:
                    return f"{year}-{month}-01"
            
            return None
        except Exception:
            return None

    # Extraction method for World Bank API:
    def _extract_world_bank_data(self, source) -> Optional[pd.DataFrame]:
        """Extract data from World Bank API using wbgapi - Universal method"""
        try:
            # Import wbgapi here to avoid dependency issues
            import wbgapi as wb
            
            series_id = source.wb_series_id
            country_code = source.wb_country_code
            
            if not series_id or not country_code:
                print(f"   ❌ Missing World Bank identifiers")
                return None
            
            print(f"   🔄 World Bank API: {series_id} for {country_code}")
            
            # Get data using wbgapi
            data = wb.data.DataFrame(series_id, country_code)
            
            if data.empty:
                print(f"   ❌ No data returned from World Bank API")
                return None
            
            # Process the data (reuse logic from vietnam_data_extractor)
            records = []
            data_reset = data.reset_index()
            
            # Check if data has year columns or time index
            for col in data_reset.columns:
                if col not in ['Country', 'economy', 'time']:
                    try:
                        year = int(str(col).replace('YR', ''))
                        if 1960 <= year <= datetime.now().year:
                            for _, row in data_reset.iterrows():
                                value = row[col]
                                if pd.notna(value) and value != 0:
                                    date_str = f"{year}-12-31"
                                    records.append({
                                        'date': date_str,
                                        'value': float(value)
                                    })
                    except (ValueError, TypeError):
                        continue
            
            if not records:
                print(f"   ❌ No valid records found after processing")
                return None
            
            df = pd.DataFrame(records)
            df['date'] = pd.to_datetime(df['date'])
            df = df.sort_values('date').reset_index(drop=True)
            df['date'] = df['date'].dt.strftime('%Y-%m-%d')
            
            # Add metadata
            df['source_name'] = source.name
            df['country'] = source.country
            df['data_type'] = source.data_type
            df['extraction_time'] = datetime.now().isoformat()
            
            print(f"   ✅ Extracted {len(df)} records from World Bank")
            return df
            
        except Exception as e:
            print(f"   ❌ World Bank API error: {e}")
            return None

    def _create_macro_excel(self, country: str, country_data: Dict[str, pd.DataFrame], output_dir: Path) -> Optional[str]:
        """Create Excel file for macro data with enhanced contents sheet matching original format"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"macro_data_{country.lower()}_{timestamp}.xlsx"
            filepath = output_dir / filename
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Create enhanced contents sheet
                contents_df = self._create_enhanced_contents_sheet(country, country_data)
                contents_df.to_excel(writer, sheet_name='Contents', index=False)
                
                # Data sheets with improved column widths
                for data_type, df in country_data.items():
                    sheet_name = data_type.replace(' ', '_')[:31]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Auto-adjust column widths for data sheets
                    try:
                        worksheet = writer.sheets[sheet_name]
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            
                            # Set minimum width for date columns to prevent ######
                            if 'date' in worksheet[f'{column_letter}1'].value.lower() if worksheet[f'{column_letter}1'].value else False:
                                adjusted_width = max(max_length + 2, 20)  # Minimum 20 for dates
                            else:
                                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                            
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                    except Exception as width_error:
                        print(f"   Warning: Could not adjust column widths: {width_error}")
                
                # Apply styling to match original format
                self._style_enhanced_contents_sheet(writer, country, len(contents_df.columns))
            
            print(f"   Created: {filename}")
            return str(filepath)
            
        except Exception as e:
            print(f"   Error creating macro Excel: {e}")
            return None
    
    def _create_enhanced_contents_sheet(self, country: str, country_data: Dict[str, pd.DataFrame]) -> pd.DataFrame:
        """Create enhanced contents sheet with source URLs and improved date detection"""
        contents_data = []
        
        row_num = 1
        for data_type, df in country_data.items():
            total_records = len(df) if df is not None else 0
            
            # Enhanced date range detection - handles both date formats and quarterly data
            date_range = "No date data"
            if df is not None and len(df) > 0:
                date_cols = [col for col in df.columns if any(word in col.lower() 
                           for word in ['date', 'time', 'period', 'quarter', 'year'])]
                
                if date_cols:
                    try:
                        date_col = date_cols[0]
                        print(f"   Analyzing date column '{date_col}' for {data_type}")
                        
                        # Since dates are now formatted as strings, we can sort them directly
                        if df[date_col].dtype == 'object':
                            # String dates - sort and get range
                            sorted_dates = df[date_col].dropna().sort_values()
                            if len(sorted_dates) > 0:
                                date_range = f"{sorted_dates.iloc[0]} to {sorted_dates.iloc[-1]}"
                                print(f"   Date range: {date_range}")
                        else:
                            # Still datetime format - convert
                            if pd.api.types.is_datetime64_any_dtype(df[date_col]):
                                valid_dates = df[date_col].dropna()
                                if len(valid_dates) > 0:
                                    min_date = valid_dates.min()
                                    max_date = valid_dates.max()
                                    date_range = f"{min_date.strftime('%Y-%m-%d')} to {max_date.strftime('%Y-%m-%d')}"
                                    print(f"   Date range: {date_range}")
                                    
                    except Exception as e:
                        print(f"   Date analysis error for {data_type}: {e}")
                        date_range = f"Analysis error: {str(e)[:50]}"
            
            # Get source URL from the data
            source_url = "N/A"
            if df is not None and len(df) > 0:
                # Try to find the original source URL by matching source name
                source_name = df['source_name'].iloc[0] if 'source_name' in df.columns else None
                if source_name:
                    source_url = self._get_source_url_by_name(source_name)
            
            # Create sheet name for hyperlink
            sheet_name = data_type.replace(' ', '_').replace('/', '_')
            sheet_name = sheet_name[:31]  # Excel limit
            
            contents_data.append({
                'No.': row_num,
                'Data Type': data_type,
                'Status': '✅ Success' if total_records > 0 else '❌ Failed',
                'Records': total_records,
                'Date Range': date_range,
                'Last Extraction Time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Sheet Link': f'=HYPERLINK("#{sheet_name}!A1","Go to {data_type}")' if total_records > 0 else 'No data available',
                'Source URL': source_url
            })
            row_num += 1
        
        return pd.DataFrame(contents_data)
    
    def _style_enhanced_contents_sheet(self, writer, country: str, num_cols: int):
        """Apply enhanced styling matching original format with updated column widths"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            workbook = writer.book
            contents_sheet = workbook['Contents']
            
            # Insert title rows
            contents_sheet.insert_rows(1, 2)
            
            # Set title
            title_cell = contents_sheet['A1']
            title_cell.value = f"{country.replace('_', ' ')} Macroeconomic Data"
            
            # Style title - matching original blue theme
            title_font = Font(size=18, bold=True, color="FFFFFF")
            title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            title_alignment = Alignment(horizontal="center", vertical="center")
            
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = title_alignment
            
            # Merge title cells
            contents_sheet.merge_cells(f'A1:{get_column_letter(num_cols)}2')
            
            # Style header row (now row 3) - matching original darker blue
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in contents_sheet[3]:
                if cell.value:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths with new Source URL column
            column_widths = {
                'A': 8,   # No.
                'B': 20,  # Data Type
                'C': 12,  # Status
                'D': 10,  # Records
                'E': 35,  # Date Range (wider for full date ranges)
                'F': 20,  # Last Extraction Time
                'G': 25,  # Sheet Link
                'H': 40   # Source URL (new column)
            }
            
            for col_letter, width in column_widths.items():
                if ord(col_letter) - ord('A') < num_cols:
                    contents_sheet.column_dimensions[col_letter].width = width
            
            # Enhanced hyperlink styling - make sure all hyperlinks are blue with underline
            link_font = Font(color="0563C1", underline="single")
            
            # Apply link styling to Sheet Link column (second to last)
            if num_cols >= 2:
                sheet_link_col = get_column_letter(num_cols - 1)  # Sheet Link is second to last
                
                for row_num in range(4, contents_sheet.max_row + 1):
                    link_cell = contents_sheet[f'{sheet_link_col}{row_num}']
                    if link_cell.value and ('HYPERLINK' in str(link_cell.value) or 'Go to' in str(link_cell.value)):
                        link_cell.font = link_font
                        print(f"   Applied hyperlink styling to {sheet_link_col}{row_num}")
            
        except Exception as e:
            print(f"Warning: Could not style contents sheet: {e}")
    
    def _create_market_excel(self, market_data: List[Dict], output_dir: Path) -> Optional[str]:
        """Create Excel file for market data with enhanced contents sheet"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"market_indices_data_{timestamp}.xlsx"
            filepath = output_dir / filename
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Create enhanced contents sheet for market data
                contents_df = self._create_market_contents_sheet(market_data)
                contents_df.to_excel(writer, sheet_name='Contents', index=False)
                
                # Data sheets
                for item in market_data:
                    sheet_name = item['name'].replace(' ', '_').replace('&', 'and')[:31]
                    item['data'].to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Apply market-specific styling
                self._style_market_contents_sheet(writer, len(contents_df.columns))
            
            print(f"   Created: {filename}")
            return str(filepath)
            
        except Exception as e:
            print(f"   Error creating market Excel: {e}")
            return None
    
    def _create_market_contents_sheet(self, market_data: List[Dict]) -> pd.DataFrame:
        """Create enhanced contents sheet for market data with source URLs"""
        contents_data = []
        
        for i, item in enumerate(market_data, 1):
            data = item['data']
            name = item['name']
            symbol = item['symbol']
            country = item['country']
            
            # Get latest data
            latest_date = data['Date'].iloc[-1] if not data.empty else "N/A"
            latest_close = f"{data['Close'].iloc[-1]:.2f}" if not data.empty else "N/A"
            
            # Get additional info from MARKET_INDICES_SOURCES including URL
            authority = "Yahoo Finance via yfinance"
            coverage_start = "2020"  # Updated to reflect new start date
            description = f"{name} stock market index, monthly closing data since 2020"
            source_url = "N/A"
            
            try:
                from macro_sources import MARKET_INDICES_SOURCES
                for source in MARKET_INDICES_SOURCES:
                    if source.name == name or name.lower() in source.name.lower():
                        authority = source.authority
                        coverage_start = "2020"  # Override with our actual start date
                        description = f"{source.description} (monthly data since 2020)"
                        source_url = source.url  # Get the actual source URL
                        break
            except:
                pass
            
            # Create sheet name for hyperlink
            sheet_name = name.replace(' ', '_').replace('&', 'and').replace('(', '').replace(')', '')
            sheet_name = sheet_name[:31]
            
            contents_data.append({
                'No.': i,
                'Index Name': name,
                'Symbol': symbol,
                'Country': country,
                'Authority': authority,
                'Status': "✅ Success",
                'Records': len(data),
                'Latest Date': latest_date,
                'Latest Close': latest_close,
                'Coverage Start': coverage_start,
                'Description': description,
                'Sheet Link': f'=HYPERLINK("#{sheet_name}!A1","Go to {name}")',
                'Source URL': source_url
            })
        
        return pd.DataFrame(contents_data)
    
    def _style_market_contents_sheet(self, writer, num_cols: int):
        """Apply styling for market data contents sheet with Source URL column"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            workbook = writer.book
            contents_sheet = workbook['Contents']
            
            # Insert title rows
            contents_sheet.insert_rows(1, 2)
            
            # Set title
            title_cell = contents_sheet['A1']
            title_cell.value = "Financial Market Data"
            
            # Style title - same blue theme
            title_font = Font(size=18, bold=True, color="FFFFFF")
            title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            title_alignment = Alignment(horizontal="center", vertical="center")
            
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = title_alignment
            
            # Merge title cells
            contents_sheet.merge_cells(f'A1:{get_column_letter(num_cols)}2')
            
            # Style header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in contents_sheet[3]:
                if cell.value:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths for market data with Source URL
            column_widths = {
                'A': 6,   # No.
                'B': 25,  # Index Name
                'C': 12,  # Symbol
                'D': 12,  # Country
                'E': 25,  # Authority
                'F': 12,  # Status
                'G': 10,  # Records
                'H': 15,  # Latest Date
                'I': 15,  # Latest Close
                'J': 15,  # Coverage Start
                'K': 35,  # Description
                'L': 20,  # Sheet Link
                'M': 40   # Source URL (new column)
            }
            
            for col_letter, width in column_widths.items():
                if ord(col_letter) - ord('A') < num_cols:
                    contents_sheet.column_dimensions[col_letter].width = width
            
            # Style hyperlinks in Sheet Link column (second to last)
            if num_cols >= 2:
                link_font = Font(color="0563C1", underline="single")
                sheet_link_col = get_column_letter(num_cols - 1)  # Sheet Link is second to last
                
                for row_num in range(4, contents_sheet.max_row + 1):
                    link_cell = contents_sheet[f'{sheet_link_col}{row_num}']
                    if link_cell.value and 'HYPERLINK' in str(link_cell.value):
                        link_cell.font = link_font
            
        except Exception as e:
            print(f"Warning: Could not style market contents sheet: {e}")

# Create tool instance
simple_rule_based_extractor = SimpleRuleBasedExtractor()
