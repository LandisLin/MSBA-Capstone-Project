"""
Malaysia Complete Data Extractor
Extracts Malaysia macroeconomic data from multiple sources into unified Excel file
Sources: Malaysia Gov API (GDP, CPI, Population), BNM website (Interest Rate), FRED (Property Price)
"""

import requests
import pandas as pd
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional
from table_data_scraper import UniversalTableScraper

class MalaysiaDataExtractor:
    """Complete Malaysia macroeconomic data extractor"""
    
    def __init__(self, output_dir: str = "./extracted_data"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.session = self._create_session()
        self.sources = self._load_malaysia_sources()
    
    def _create_session(self) -> requests.Session:
        """Create HTTP session with proper headers"""
        session = requests.Session()
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'application/json, text/html, */*',
            'Accept-Language': 'en-US,en;q=0.9'
        })
        return session
    
    def _load_malaysia_sources(self) -> Dict:
        """Load Malaysia sources from macro_sources.py"""
        try:
            from macro_sources import MALAYSIA_SOURCES
            
            sources_dict = {}
            for source in MALAYSIA_SOURCES:
                sources_dict[source.data_type] = {
                    'name': source.name,
                    'url': source.url,
                    'api_url': source.api_url if source.api_url else source.url,
                    'source_type': source.source_type,
                    'authority': source.authority,
                    'description': source.description,
                    'fred_series_id': getattr(source, 'fred_series_id', None)
                }
            
            print(f"✅ Loaded {len(sources_dict)} Malaysia sources from macro_sources.py")
            return sources_dict
            
        except ImportError:
            print("❌ Could not import MALAYSIA_SOURCES from macro_sources.py")
            return {}
        except Exception as e:
            print(f"❌ Error loading Malaysia sources: {e}")
            return {}
    
    def extract_all_data(self) -> Optional[str]:
        """Extract all Malaysia data sources and combine into single Excel file"""
        
        print("🇲🇾 MALAYSIA COMPLETE DATA EXTRACTION")
        print("=" * 50)
        
        if not self.sources:
            print("❌ No sources loaded. Please add MALAYSIA_SOURCES to macro_sources.py")
            return None
        
        extracted_data = {}
        
        # Extract each data source
        for indicator, source_info in self.sources.items():
            print(f"\n📊 Extracting {indicator}: {source_info['name']}")
            
            try:
                if source_info['source_type'] == 'api':
                    if 'data.gov.my' in source_info['api_url']:
                        data = self._extract_malaysia_gov_api(source_info, indicator)
                    elif source_info.get('fred_series_id'):
                        data = self._extract_fred_api(source_info, indicator)
                    else:
                        print(f"   ❌ Unknown API type")
                        data = None
                elif source_info['source_type'] == 'web_scraping':
                    data = self._extract_web_scraping(source_info, indicator)
                else:
                    print(f"   ❌ Unknown source type: {source_info['source_type']}")
                    data = None
                
                if data is not None and not data.empty:
                    extracted_data[indicator] = data
                    print(f"   ✅ Success: {len(data)} records")
                    print(f"   📅 Range: {data['date'].min()} to {data['date'].max()}")
                else:
                    print(f"   ❌ Failed to extract data")
                    
            except Exception as e:
                print(f"   ❌ Error: {e}")
                continue
        
        if extracted_data:
            excel_file = self._save_combined_excel(extracted_data)
            return excel_file
        else:
            print("\n❌ No data extracted")
            return None
    
    def _extract_malaysia_gov_api(self, source_info: Dict, indicator: str) -> Optional[pd.DataFrame]:
        """Extract data from Malaysia Government API"""
        try:
            api_url = source_info['api_url']
            print(f"   🔄 Malaysia Gov API: {api_url}")
            
            response = self.session.get(api_url, timeout=30)
            if response.status_code != 200:
                print(f"   ❌ API request failed: {response.status_code}")
                return None
            
            data_json = response.json()
            
            # Get data array based on API response structure
            if isinstance(data_json, dict) and 'data_catalogue' in data_json:
                data_array = data_json['data_catalogue']
            elif isinstance(data_json, list):
                data_array = data_json
            else:
                print(f"   ⚠️ Unexpected API response structure")
                return None
            
            print(f"   📊 Found {len(data_array)} total records")
            
            # Process the data with filtering
            records = []
            for item in data_array:
                if isinstance(item, dict) and self._should_include_record(item, indicator):
                    record = self._parse_malaysia_gov_record(item, indicator)
                    if record:
                        records.append(record)
            
            if records:
                df = pd.DataFrame(records)
                df = df.sort_values('date')
                print(f"   ✅ Processed: {len(df)} valid records")
                return df
            else:
                print(f"   ❌ No valid records found")
                return None
                
        except Exception as e:
            print(f"   ❌ Malaysia Gov API error: {e}")
            return None
    
    def _should_include_record(self, item: Dict, indicator: str) -> bool:
        """Filter records based on indicator-specific criteria"""
        try:
            if indicator == 'GDP':
                return item.get('series', '').lower() == 'abs'
            elif indicator == 'CPI':
                return item.get('division', '').lower() == 'overall'
            elif indicator == 'Population':
                return (item.get('age', '').lower() == 'overall' and 
                       item.get('sex', '').lower() == 'both' and
                       item.get('ethnicity', '').lower() == 'overall')
            else:
                return True
        except Exception:
            return False
    
    def _parse_malaysia_gov_record(self, item: Dict, indicator: str) -> Optional[Dict]:
        """Parse individual record from Malaysia Gov API"""
        date_fields = ['date', 'period', 'time', 'quarter', 'year', 'x']
        value_fields = ['value', 'amount', 'index', 'rate', 'population', 'gdp', 'cpi', 'y']
        
        date_val = None
        value_val = None
        
        # Find date and value fields
        for field in date_fields:
            if field in item and item[field] is not None:
                date_val = str(item[field])
                break
        
        for field in value_fields:
            if field in item and item[field] is not None:
                try:
                    value_val = float(item[field])
                    break
                except (ValueError, TypeError):
                    continue
        
        if date_val and value_val is not None:
            try:
                # Standardize date format
                if len(date_val) == 4:  # Year only
                    date_obj = pd.to_datetime(f"{date_val}-01-01")
                elif 'Q' in date_val.upper():  # Quarterly
                    if '-Q' in date_val:
                        year, quarter = date_val.split('-Q')
                    else:
                        year = date_val[:-2]
                        quarter = date_val[-1]
                    
                    quarter_month = {'1': '01', '2': '04', '3': '07', '4': '10'}
                    month = quarter_month.get(quarter, '01')
                    date_obj = pd.to_datetime(f"{year}-{month}-01")
                else:
                    date_obj = pd.to_datetime(date_val)
                
                return {
                    'date': date_obj.strftime('%Y-%m-%d'),
                    'value': value_val,
                    'source_name': f"Malaysia {indicator}",
                    'data_type': indicator,
                    'country': 'Malaysia',
                    'unit': self._get_unit_for_indicator(indicator),
                    'extraction_time': datetime.now().isoformat()
                }
            except Exception:
                pass
        
        return None
    
    def _get_unit_for_indicator(self, indicator: str) -> str:
        """Get appropriate unit for each indicator"""
        unit_mapping = {
            'GDP': 'Million RM (2015 prices)',
            'CPI': 'Index (2010=100)',
            'Interest_Rate': 'Percent per annum',
            'Population': 'Number of persons',
            'Property_Price': 'Index (2010=100)'
        }
        return unit_mapping.get(indicator, 'Unknown units')
    
    def _extract_fred_api(self, source_info: Dict, indicator: str) -> Optional[pd.DataFrame]:
        """Extract data from FRED API"""
        from dotenv import load_dotenv
        
        load_dotenv()
        fred_api_key = os.getenv('FRED_API_KEY')
        
        if not fred_api_key:
            print(f"   ❌ FRED API key not found")
            return None
        
        try:
            fred_series_id = source_info.get('fred_series_id')
            if not fred_series_id:
                print(f"   ❌ No FRED series ID found")
                return None
            
            api_url = f"https://api.stlouisfed.org/fred/series/observations?series_id={fred_series_id}&api_key={fred_api_key}&file_type=json"
            print(f"   🔄 FRED API: series {fred_series_id}")
            
            response = self.session.get(api_url, timeout=15)
            if response.status_code != 200:
                print(f"   ❌ FRED API failed: {response.status_code}")
                return None
            
            data = response.json()
            if 'observations' not in data:
                print(f"   ❌ No observations in FRED response")
                return None
            
            observations = data['observations']
            records = []
            
            for obs in observations:
                try:
                    date_str = obs['date']
                    value_str = obs['value']
                    
                    # Skip missing values
                    if value_str != '.' and value_str != '':
                        date_obj = pd.to_datetime(date_str)
                        value_num = float(value_str)
                        
                        records.append({
                            'date': date_obj.strftime('%Y-%m-%d'),
                            'value': value_num,
                            'source_name': source_info['name'],
                            'data_type': indicator,
                            'country': 'Malaysia',
                            'unit': 'Index (Base 2010=100)',
                            'extraction_time': datetime.now().isoformat()
                        })
                except Exception:
                    continue
            
            if records:
                df = pd.DataFrame(records)
                df = df.sort_values('date')
                print(f"   ✅ FRED: {len(df)} records")
                return df
            else:
                print(f"   ❌ No valid FRED records found")
                return None
                
        except Exception as e:
            print(f"   ❌ FRED API error: {e}")
            return None
    
    def _extract_web_scraping(self, source_info: Dict, indicator: str) -> Optional[pd.DataFrame]:
        """Extract data using web scraping"""
        try:
            scraper = UniversalTableScraper()
            df = scraper.extract_table_data(
                url=source_info['api_url'],
                data_name=source_info['name'],
                date_format="%d/%m/%Y",
                value_range=(0.25, 15.0),
                country="Malaysia",
                data_type=indicator,
                unit="Percent per annum",
                verbose=False
            )
            return df
        except Exception as e:
            print(f"   ❌ Web scraping error: {e}")
            return None
    
    def _save_combined_excel(self, extracted_data: Dict[str, pd.DataFrame]) -> str:
        """Save all extracted data to combined Excel file"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"macro_data_malaysia_{timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        print(f"\n💾 Creating Excel file: {filename}")
        
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Create Contents sheet
                self._create_contents_sheet(writer, extracted_data)
                
                # Create individual data sheets
                for indicator, df in extracted_data.items():
                    df.to_excel(writer, sheet_name=indicator, index=False)
                    print(f"   📄 Created sheet: {indicator} ({len(df)} records)")
            
            print(f"✅ Excel file created: {filename}")
            return str(filepath)
            
        except Exception as e:
            print(f"❌ Error creating Excel file: {e}")
            raise
    
    def _create_contents_sheet(self, writer, extracted_data: Dict[str, pd.DataFrame]):
        """Create Contents sheet matching other countries' format"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            workbook = writer.book
            contents_sheet = workbook.create_sheet('Contents', 0)
            
            # Title
            contents_sheet['A1'] = "Malaysia Macroeconomic Data"
            contents_sheet.merge_cells('A1:H2')
            
            # Style title
            title_font = Font(size=18, bold=True, color="FFFFFF")
            title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            title_alignment = Alignment(horizontal="center", vertical="center")
            contents_sheet['A1'].font = title_font
            contents_sheet['A1'].fill = title_fill
            contents_sheet['A1'].alignment = title_alignment
            
            # Headers
            headers = ['No.', 'Data Type', 'Status', 'Records', 'Date Range', 'Last Extraction Time', 'Sheet Link', 'Source URL']
            for col, header in enumerate(headers, 1):
                cell = contents_sheet.cell(row=3, column=col)
                cell.value = header
                
                # Style header
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Data rows
            for i, (indicator, df) in enumerate(extracted_data.items(), 1):
                row = 3 + i
                records = len(df)
                date_range = f"{df['date'].min()} to {df['date'].max()}" if records > 0 else "No data"
                source_url = self.sources[indicator]['url']
                
                contents_sheet.cell(row=row, column=1).value = i
                contents_sheet.cell(row=row, column=2).value = indicator
                contents_sheet.cell(row=row, column=3).value = "✅ Success" if records > 0 else "❌ Failed"
                contents_sheet.cell(row=row, column=4).value = records
                contents_sheet.cell(row=row, column=5).value = date_range
                contents_sheet.cell(row=row, column=6).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # Sheet link with hyperlink
                if records > 0:
                    hyperlink_formula = f'=HYPERLINK("#{indicator}!A1","Go to {indicator}")'
                    link_cell = contents_sheet.cell(row=row, column=7)
                    link_cell.value = hyperlink_formula
                    link_font = Font(color="0563C1", underline="single")
                    link_cell.font = link_font
                else:
                    contents_sheet.cell(row=row, column=7).value = "No data available"
                
                contents_sheet.cell(row=row, column=8).value = source_url
            
            # Set column widths
            column_widths = {'A': 8, 'B': 20, 'C': 12, 'D': 10, 'E': 35, 'F': 20, 'G': 20, 'H': 60}
            for col_letter, width in column_widths.items():
                contents_sheet.column_dimensions[col_letter].width = width
            
        except Exception as e:
            print(f"   ⚠️ Error creating styled Contents sheet: {e}")
            # Create basic contents sheet without styling
            contents_data = []
            for i, (indicator, df) in enumerate(extracted_data.items(), 1):
                contents_data.append({
                    'No.': i,
                    'Data Type': indicator,
                    'Status': '✅ Success' if len(df) > 0 else '❌ Failed',
                    'Records': len(df),
                    'Date Range': f"{df['date'].min()} to {df['date'].max()}" if len(df) > 0 else "No data",
                    'Source URL': self.sources[indicator]['url']
                })
            
            contents_df = pd.DataFrame(contents_data)
            contents_df.to_excel(writer, sheet_name='Contents', index=False)


def clean_malaysia_data(input_file_path: str = None) -> Optional[str]:
    """
    Clean Malaysia data: Filter to 2000+ and create cleaned file with updated Contents sheet
    
    Args:
        input_file_path: Path to Malaysia data file to clean (auto-detects if None)
        
    Returns:
        Path to cleaned file if successful, None if failed
    """
    
    # Auto-find latest Malaysia file if not specified
    if input_file_path is None:
        data_dir = Path("./extracted_data")
        if not data_dir.exists():
            print("❌ Data directory not found")
            return None
        
        # Find latest Malaysia file
        malaysia_files = list(data_dir.glob("macro_data_malaysia_*.xlsx"))
        if not malaysia_files:
            print("❌ No Malaysia files found")
            return None
        
        input_file_path = max(malaysia_files, key=lambda x: x.stat().st_mtime)
        print(f"📄 Found Malaysia file: {input_file_path.name}")
    
    input_path = Path(input_file_path)
    output_file_path = input_path.parent / f"cleaned_{input_path.name}"
    
    print(f"\n🧹 CLEANING MALAYSIA DATA")
    print("=" * 40)
    print(f"Input:  {input_path.name}")
    print(f"Output: {output_file_path.name}")
    print(f"Filter: 2000-01-01 onwards")
    
    try:
        # Read all sheets except Contents
        excel_data = pd.read_excel(input_file_path, sheet_name=None)
        cleaned_data = {}
        
        # First, filter all data sheets (excluding Contents)
        for sheet_name, df in excel_data.items():
            if sheet_name != 'Contents':
                if 'date' in df.columns:
                    original_count = len(df)
                    df_filtered = df[df['date'] >= '2000-01-01'].copy()
                    cleaned_data[sheet_name] = df_filtered
                    
                    print(f"   📊 {sheet_name}: {original_count} → {len(df_filtered)} records")
                    if len(df_filtered) > 0:
                        print(f"       Range: {df_filtered['date'].min()} to {df_filtered['date'].max()}")
                else:
                    cleaned_data[sheet_name] = df
                    print(f"   ⚠️ {sheet_name}: No date column, kept as-is")
        
        # Write cleaned data with updated Contents sheet
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            # Create updated Contents sheet with actual cleaned data stats
            _create_updated_contents_sheet(writer, cleaned_data, input_file_path)
            
            # Write data sheets
            for sheet_name, df in cleaned_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"\n✅ Cleaning completed!")
        print(f"📁 Output: {output_file_path.name}")
        return str(output_file_path)
        
    except Exception as e:
        print(f"❌ Cleaning error: {e}")
        return None


def _create_updated_contents_sheet(writer, cleaned_data: Dict[str, pd.DataFrame], input_file_path: str):
    """Create updated Contents sheet with cleaned data statistics"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        
        workbook = writer.book
        contents_sheet = workbook.create_sheet('Contents', 0)
        
        # Title (rows 1-2)
        contents_sheet['A1'] = "Malaysia Macroeconomic Data"
        contents_sheet.merge_cells('A1:H2')
        
        # Style title
        title_font = Font(size=18, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        contents_sheet['A1'].font = title_font
        contents_sheet['A1'].fill = title_fill
        contents_sheet['A1'].alignment = title_alignment
        
        # Headers (row 3)
        headers = ['No.', 'Data Type', 'Status', 'Records', 'Date Range', 'Last Extraction Time', 'Sheet Link', 'Source URL']
        for col, header in enumerate(headers, 1):
            cell = contents_sheet.cell(row=3, column=col)
            cell.value = header
            
            # Style header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Get source URLs from original file
        source_urls = _get_source_urls_from_original(input_file_path)
        
        # Data rows with updated statistics
        for i, (indicator, df) in enumerate(cleaned_data.items(), 1):
            row = 3 + i
            
            # Calculate actual cleaned data stats
            records = len(df)
            date_range = f"{df['date'].min()} to {df['date'].max()}" if records > 0 and 'date' in df.columns else "No data"
            source_url = source_urls.get(indicator, "N/A")
            
            # Fill row data
            contents_sheet.cell(row=row, column=1).value = i
            contents_sheet.cell(row=row, column=2).value = indicator
            contents_sheet.cell(row=row, column=3).value = "✅ Success" if records > 0 else "❌ Failed"
            contents_sheet.cell(row=row, column=4).value = records
            contents_sheet.cell(row=row, column=5).value = date_range
            contents_sheet.cell(row=row, column=6).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Sheet link with hyperlink
            if records > 0:
                hyperlink_formula = f'=HYPERLINK("#{indicator}!A1","Go to {indicator}")'
                link_cell = contents_sheet.cell(row=row, column=7)
                link_cell.value = hyperlink_formula
                link_font = Font(color="0563C1", underline="single")
                link_cell.font = link_font
            else:
                contents_sheet.cell(row=row, column=7).value = "No data available"
            
            contents_sheet.cell(row=row, column=8).value = source_url
        
        # Set column widths
        column_widths = {'A': 8, 'B': 20, 'C': 12, 'D': 10, 'E': 35, 'F': 20, 'G': 20, 'H': 60}
        for col_letter, width in column_widths.items():
            contents_sheet.column_dimensions[col_letter].width = width
        
        print(f"   📋 Contents sheet updated with cleaned data statistics")
        
    except Exception as e:
        print(f"   ⚠️ Error creating updated Contents sheet: {e}")


def _get_source_urls_from_original(input_file_path: str) -> Dict[str, str]:
    """Extract source URLs from MALAYSIA_SOURCES directly"""
    source_urls = {}
    
    try:
        # Get URLs directly from MALAYSIA_SOURCES (same as original extraction)
        from macro_sources import MALAYSIA_SOURCES
        
        for source in MALAYSIA_SOURCES:
            source_urls[source.data_type] = source.url
        
        print(f"   📎 Loaded Malaysia URLs from macro_sources.py: {len(source_urls)} sources")
        return source_urls
        
    except ImportError:
        print(f"   ⚠️ Could not import MALAYSIA_SOURCES from macro_sources.py")
        return {}
    except Exception as e:
        print(f"   ⚠️ Error loading Malaysia source URLs: {e}")
        return {}


def main():
    """Main execution function"""
    print("🚀 MALAYSIA DATA EXTRACTOR")
    print("Sources: Malaysia Gov API, BNM Website, FRED API")
    print("=" * 50)
    
    try:
        # Extract data
        extractor = MalaysiaDataExtractor()
        excel_file = extractor.extract_all_data()
        
        if excel_file:
            print(f"\n🎉 EXTRACTION COMPLETE!")
            print(f"📁 File: {Path(excel_file).name}")
            
            # Clean the data
            cleaned_file = clean_malaysia_data(excel_file)
            if cleaned_file:
                print(f"\n🧹 CLEANING COMPLETE!")
                print(f"📁 Cleaned file: {Path(cleaned_file).name}")
            
        return excel_file
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return None


# Convenience functions
def extract_malaysia_data() -> Optional[str]:
    """Convenience function to extract Malaysia data"""
    extractor = MalaysiaDataExtractor()
    return extractor.extract_all_data()


if __name__ == "__main__":
    main()