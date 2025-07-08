"""
Thailand Complete Data Extractor - FIXED VERSION
Fixes: IMF API endpoints, Excel download format, Contents sheet error
"""

import requests
import pandas as pd
import os
import urllib.request
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional
import tempfile

class ThailandDataExtractor:
    """Complete Thailand macroeconomic data extractor - FIXED"""
    
    def __init__(self, output_dir: str = "./extracted_data"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.session = self._create_session()
        self.sources = self._load_thailand_sources()
        self.fred_api_key = os.getenv('FRED_API_KEY')
    
    def _create_session(self) -> requests.Session:
        """Create HTTP session with proper headers"""
        session = requests.Session()
        session.headers.update({
            'User-Agent': 'Economic-Data-Pipeline/1.0',
            'Accept': 'application/json, text/csv, application/vnd.ms-excel, */*',
            'Cache-Control': 'no-cache'
        })
        return session
    
    def _load_thailand_sources(self) -> Dict:
        """Load Thailand sources from macro_sources.py"""
        try:
            from macro_sources import THAILAND_SOURCES
            
            sources_dict = {}
            for source in THAILAND_SOURCES:
                sources_dict[source.data_type] = {
                    'name': source.name,
                    'url': source.url,
                    'api_url': source.api_url if source.api_url else source.url,
                    'source_type': source.source_type,
                    'authority': source.authority,
                    'description': source.description,
                    'fred_series_id': getattr(source, 'fred_series_id', None)
                }
            
            print(f"✅ Loaded {len(sources_dict)} Thailand sources from macro_sources.py")
            return sources_dict
            
        except ImportError:
            print("❌ Could not import THAILAND_SOURCES from macro_sources.py")
            return {}
        except Exception as e:
            print(f"❌ Error loading Thailand sources: {e}")
            return {}
    
    def extract_all_data(self) -> Optional[str]:
        """Extract all Thailand data sources and combine into single Excel file"""
        
        print("🇹🇭 THAILAND COMPLETE DATA EXTRACTION")
        print("=" * 50)
        
        if not self.sources:
            print("❌ No sources loaded. Please add THAILAND_SOURCES to macro_sources.py")
            return None
        
        extracted_data = {}
        
        # Extract each data source
        for indicator, source_info in self.sources.items():
            print(f"\n📊 Extracting {indicator}: {source_info['name']}")
            
            try:
                if source_info['source_type'] == 'api':
                    if ('api.imf.org' in source_info['api_url'] or 
                        'data.imf.org' in source_info['api_url'] or
                        indicator in ['GDP', 'CPI']):
                        data = self._extract_imf_api_fixed(source_info, indicator)
                    elif source_info.get('fred_series_id'):
                        data = self._extract_fred_api(source_info, indicator)
                    else:
                        print(f"   ❌ Unknown API type")
                        data = None
                elif source_info['source_type'] == 'excel_download':
                    data = self._extract_excel_download_fixed(source_info, indicator)
                else:
                    print(f"   ❌ Unknown source type: {source_info['source_type']}")
                    data = None
                
                if data is not None and not data.empty:
                    extracted_data[indicator] = data
                    print(f"   ✅ Success: {len(data)} records")
                    print(f"   📅 Range: {data['date'].min()} to {data['date'].max()}")
                else:
                    print(f"   ❌ Failed to extract data")
                    
            except Exception as e:
                print(f"   ❌ Error: {e}")
                continue
        
        if extracted_data:
            excel_file = self._save_combined_excel(extracted_data)
            return excel_file
        else:
            print("\n❌ No data extracted")
            return None
    
    def _extract_imf_api_fixed(self, source_info: Dict, indicator: str) -> Optional[pd.DataFrame]:
        """
        FIXED: Extract data from IMF API using proper API endpoints
        The original URLs were for web interface, not API
        """
        try:
            # Use the configured API URL from macro_sources.py
            api_url = source_info['api_url']
            
            print(f"   🔄 IMF API (Fixed): {api_url}")
            
            headers = {
                'Cache-Control': 'no-cache',
                'User-Agent': 'Economic-Data-Pipeline/1.0',
                'Accept': 'application/json'
            }
            
            req = urllib.request.Request(api_url, headers=headers)
            response = urllib.request.urlopen(req, timeout=30)
            
            if response.getcode() != 200:
                print(f"   ❌ IMF API request failed: {response.getcode()}")
                return None
            
            # Read and parse response
            response_text = response.read().decode('utf-8')
            
            # Check if response is actually JSON
            if not response_text.strip().startswith('{'):
                print(f"   ❌ IMF API returned non-JSON response")
                print(f"   📝 Response preview: {response_text[:200]}...")
                return None
            
            data = json.loads(response_text)
            
            # Navigate to the data structure
            if 'data' not in data or 'dataSets' not in data['data']:
                print(f"   ❌ Invalid IMF response structure")
                return None
                
            dataset = data['data']['dataSets'][0]
            series_data = dataset.get('series', {})
            
            # Extract time periods from structure
            time_values = self._extract_imf_time_periods(data)
            
            if not time_values:
                print(f"   ❌ Could not extract time periods from IMF response")
                return None
            
            # Process series data
            all_records = []
            
            for series_key, series_info in series_data.items():
                observations = series_info.get('observations', {})
                
                for obs_key, obs_data in observations.items():
                    try:
                        time_index = int(obs_key)
                        
                        if isinstance(obs_data, list) and len(obs_data) > 0:
                            value = obs_data[0]
                            
                            if value is not None:
                                numeric_value = float(value)
                                
                                if time_index < len(time_values):
                                    time_period = time_values[time_index]
                                    
                                    # Convert time format based on data type
                                    if indicator == 'GDP':
                                        date_str = self._convert_quarterly_to_date(time_period)
                                    else:  # CPI
                                        date_str = self._convert_imf_monthly_to_date(time_period)
                                    
                                    if date_str:
                                        all_records.append({
                                            'date': date_str,
                                            'value': numeric_value,
                                            'time_period': time_period,
                                            'country': 'Thailand',
                                            'indicator': indicator,
                                            'source': 'IMF',
                                            'extraction_time': datetime.now().isoformat()
                                        })
                    
                    except (ValueError, IndexError, TypeError):
                        continue
            
            if all_records:
                df = pd.DataFrame(all_records)
                df['date'] = pd.to_datetime(df['date'])
                df = df.sort_values('date')
                df = df.drop_duplicates(subset=['date'], keep='last')
                df['date'] = df['date'].dt.strftime('%Y-%m-%d')
                
                print(f"   ✅ IMF extraction successful: {len(df)} records")
                return df
            else:
                print(f"   ❌ No valid records extracted from IMF API")
                return None
                
        except Exception as e:
            print(f"   ❌ IMF API error: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _extract_excel_download_fixed(self, source_info: Dict, indicator: str) -> Optional[pd.DataFrame]:
        """
        CORRECTED: Extract data from Excel/CSV download with proper date column detection
        """
        try:
            download_url = source_info['api_url']
            print(f"   🔄 File Download (Corrected): {download_url}")
            
            # Download file
            response = self.session.get(download_url, timeout=60)
            if response.status_code != 200:
                print(f"   ❌ Download failed: {response.status_code}")
                return None
            
            # Check content type
            content_type = response.headers.get('content-type', '').lower()
            print(f"   📄 Content type: {content_type}")
            
            # Save to temporary file
            with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                temp_file.write(response.content)
                temp_file_path = temp_file.name
            
            try:
                # Try different parsing methods
                df = None
                
                # Method 1: Try as CSV first
                try:
                    df = pd.read_csv(temp_file_path, encoding='utf-8')
                    print(f"   📊 Parsed as CSV: {len(df)} rows, {len(df.columns)} columns")
                except:
                    try:
                        df = pd.read_csv(temp_file_path, encoding='iso-8859-1')
                        print(f"   📊 Parsed as CSV (ISO): {len(df)} rows, {len(df.columns)} columns")
                    except:
                        pass
                
                if df is None or len(df) == 0:
                    print(f"   ❌ Could not parse downloaded file")
                    return None
                
                # Show column info for debugging
                print(f"   📋 Columns: {list(df.columns)}")
                print(f"   📋 First few rows:")
                print(df.head(3).to_string())
                
                # CORRECTED: Proper date and value column detection for Thailand THOR data
                date_col = None
                value_col = None
                
                # For Thailand THOR data, we know the structure from your image:
                # Columns: ['As of', 'Code', 'Tenor', 'Rate', 'Days Count']
                # We want: 'As of' for date, 'Rate' for value
                
                for col in df.columns:
                    col_str = str(col).lower().strip()
                    
                    # Look for date columns - prioritize "As of" for THOR data
                    if col_str in ['as of', 'date', 'วันที่']:
                        date_col = col
                        print(f"   📅 Found date column: {col}")
                        break
                    elif any(keyword in col_str for keyword in ['date', 'day', 'time']):
                        if date_col is None:  # Only set if we haven't found a better match
                            date_col = col
                            print(f"   📅 Found potential date column: {col}")
                
                for col in df.columns:
                    col_str = str(col).lower().strip()
                    
                    # Look for rate/value columns - prioritize "Rate" for THOR data
                    if col_str in ['rate', 'อัตรา']:
                        value_col = col
                        print(f"   💰 Found value column: {col}")
                        break
                    elif any(keyword in col_str for keyword in ['value', 'price']) and df[col].dtype in ['float64', 'int64']:
                        if value_col is None:
                            value_col = col
                            print(f"   💰 Found potential value column: {col}")
                
                # For THOR data specifically, filter to overnight rates only (O/N tenor)
                if 'Code' in df.columns and 'Tenor' in df.columns:
                    # Filter for THOR (overnight) rates only
                    thor_mask = (df['Code'] == 'THOR') & (df['Tenor'] == 'O/N')
                    df_filtered = df[thor_mask].copy()
                    print(f"   🔍 Filtered to THOR O/N rates: {len(df)} → {len(df_filtered)} records")
                    df = df_filtered
                
                if date_col is not None and value_col is not None and len(df) > 0:
                    # Clean and format data
                    clean_df = df[[date_col, value_col]].copy()
                    clean_df.columns = ['date', 'value']
                    
                    # Remove missing values
                    clean_df = clean_df.dropna()
                    
                    # Convert value to numeric
                    clean_df['value'] = pd.to_numeric(clean_df['value'], errors='coerce')
                    clean_df = clean_df.dropna()
                    
                    # CORRECTED: Proper date conversion for Thailand format
                    # Thailand date format is likely M/D/YYYY (e.g., 1/5/2015)
                    print(f"   📅 Sample date values: {clean_df['date'].head(3).tolist()}")
                    
                    # Try different date formats
                    date_converted = False
                    for date_format in ['%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d', '%Y/%m/%d']:
                        try:
                            clean_df['date'] = pd.to_datetime(clean_df['date'], format=date_format)
                            print(f"   ✅ Date format matched: {date_format}")
                            date_converted = True
                            break
                        except:
                            continue
                    
                    if not date_converted:
                        # Try automatic date parsing as fallback
                        try:
                            clean_df['date'] = pd.to_datetime(clean_df['date'], errors='coerce')
                            date_converted = True
                            print(f"   ✅ Automatic date parsing succeeded")
                        except:
                            pass
                    
                    if not date_converted:
                        print(f"   ❌ Could not parse dates")
                        return None
                    
                    # Remove rows with invalid dates
                    clean_df = clean_df.dropna()
                    
                    if len(clean_df) == 0:
                        print(f"   ❌ No valid records after date conversion")
                        return None
                    
                    # Add metadata
                    clean_df['country'] = 'Thailand'
                    clean_df['data_type'] = indicator
                    clean_df['source'] = source_info['authority']
                    clean_df['extraction_time'] = datetime.now().isoformat()
                    
                    # Sort by date and format
                    clean_df = clean_df.sort_values('date')
                    clean_df['date'] = clean_df['date'].dt.strftime('%Y-%m-%d')
                    
                    print(f"   ✅ File extraction successful: {len(clean_df)} records")
                    print(f"   📅 Actual date range: {clean_df['date'].min()} to {clean_df['date'].max()}")
                    
                    return clean_df
                else:
                    print(f"   ❌ Could not identify proper date/value columns or no data after filtering")
                    print(f"   📋 Date column: {date_col}, Value column: {value_col}, Records: {len(df)}")
                    return None
                    
            finally:
                # Clean up temp file
                if os.path.exists(temp_file_path):
                    os.unlink(temp_file_path)
                    
        except Exception as e:
            print(f"   ❌ File download error: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _extract_fred_api(self, source_info: Dict, indicator: str) -> Optional[pd.DataFrame]:
        """Extract data from FRED API"""
        try:
            if not self.fred_api_key:
                print(f"   ❌ FRED API key not found in environment")
                return None
            
            fred_series_id = source_info['fred_series_id']
            api_url = f"https://api.stlouisfed.org/fred/series/observations?series_id={fred_series_id}&api_key={self.fred_api_key}&file_type=json"
            
            print(f"   🔄 FRED API: {fred_series_id}")
            
            response = self.session.get(api_url, timeout=30)
            if response.status_code != 200:
                print(f"   ❌ FRED API failed: {response.status_code}")
                return None
            
            data = response.json()
            observations = data.get('observations', [])
            
            records = []
            for obs in observations:
                date_str = obs.get('date')
                value_str = obs.get('value')
                
                if date_str and value_str and value_str != '.' and value_str != '':
                    try:
                        date_obj = pd.to_datetime(date_str)
                        value_num = float(value_str)
                        
                        records.append({
                            'date': date_obj.strftime('%Y-%m-%d'),
                            'value': value_num,
                            'source_name': source_info['name'],
                            'data_type': indicator,
                            'country': 'Thailand',
                            'extraction_time': datetime.now().isoformat()
                        })
                    except Exception:
                        continue
            
            if records:
                df = pd.DataFrame(records)
                df = df.sort_values('date')
                return df
            else:
                return None
                
        except Exception as e:
            print(f"   ❌ FRED API error: {e}")
            return None
    
    def _extract_imf_time_periods(self, data: dict) -> list:
        """Extract time periods from IMF API response structure"""
        try:
            if 'data' in data and 'structures' in data['data'] and len(data['data']['structures']) > 0:
                structure = data['data']['structures'][0]
                dimensions = structure.get('dimensions', {})
                for dim in dimensions.get('observation', []):
                    if dim.get('id') == 'TIME_PERIOD':
                        values = dim.get('values', [])
                        time_periods = [v['value'] for v in values if isinstance(v, dict) and 'value' in v]
                        if time_periods:
                            return time_periods
            return []
        except Exception:
            return []

    def _convert_quarterly_to_date(self, time_period: str) -> Optional[str]:
        """Convert quarterly time period to date (2024-Q1 -> 2024-01-01)"""
        try:
            if not time_period or '-Q' not in time_period:
                return None
            
            parts = time_period.split('-Q')
            if len(parts) != 2:
                return None
            
            year = parts[0]
            quarter = parts[1]
            
            quarter_months = {'1': '01', '2': '04', '3': '07', '4': '10'}
            month = quarter_months.get(quarter)
            
            if month and len(year) == 4 and year.isdigit():
                return f"{year}-{month}-01"
            
            return None
        except Exception:
            return None

    def _convert_imf_monthly_to_date(self, time_period: str) -> Optional[str]:
        """Convert IMF monthly time period to date (2010-M01 -> 2010-01-01)"""
        try:
            if not time_period or '-M' not in time_period:
                return None
            
            parts = time_period.split('-M')
            if len(parts) != 2:
                return None
            
            year = parts[0]
            month = parts[1].zfill(2)
            
            if len(year) == 4 and year.isdigit() and len(month) == 2 and month.isdigit():
                month_int = int(month)
                if 1 <= month_int <= 12:
                    return f"{year}-{month}-01"
            
            return None
        except Exception:
            return None
    
    def _save_combined_excel(self, extracted_data: Dict[str, pd.DataFrame]) -> str:
        """Save all extracted data to combined Excel file"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"macro_data_thailand_{timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        print(f"\n💾 Creating Excel file: {filename}")
        
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Create Contents sheet (FIXED - avoid merged cell issues)
                self._create_contents_sheet(writer, extracted_data)

                # Create individual data sheets
                for indicator, df in extracted_data.items():
                    df.to_excel(writer, sheet_name=indicator, index=False)
                    print(f"   📄 Created sheet: {indicator} ({len(df)} records)")
            
            print(f"✅ Excel file created: {filename}")
            return str(filepath)
            
        except Exception as e:
            print(f"❌ Error creating Excel file: {e}")
            raise
    
    def _create_contents_sheet(self, writer, extracted_data: Dict[str, pd.DataFrame]):
        """Create Contents sheet matching other countries' format with proper styling"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            workbook = writer.book
            contents_sheet = workbook.create_sheet('Contents', 0)
            
            # Title (rows 1-2)
            contents_sheet['A1'] = "Thailand Macroeconomic Data"
            contents_sheet.merge_cells('A1:H2')
            
            # Style title
            title_font = Font(size=18, bold=True, color="FFFFFF")
            title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            title_alignment = Alignment(horizontal="center", vertical="center")
            contents_sheet['A1'].font = title_font
            contents_sheet['A1'].fill = title_fill
            contents_sheet['A1'].alignment = title_alignment
            
            # Headers (row 3)
            headers = ['No.', 'Data Type', 'Status', 'Records', 'Date Range', 'Last Extraction Time', 'Sheet Link', 'Source URL']
            for i, header in enumerate(headers):
                cell = contents_sheet.cell(row=3, column=i+1, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows (starting from row 4)
            for i, (indicator, df) in enumerate(extracted_data.items(), 1):
                row_num = i + 3
                
                # Data
                contents_sheet[f'A{row_num}'] = i
                contents_sheet[f'B{row_num}'] = indicator
                contents_sheet[f'C{row_num}'] = '✅ Success' if len(df) > 0 else '❌ Failed'
                contents_sheet[f'D{row_num}'] = len(df)
                contents_sheet[f'E{row_num}'] = f"{df['date'].min()} to {df['date'].max()}" if len(df) > 0 else "No data"
                contents_sheet[f'F{row_num}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                contents_sheet[f'G{row_num}'] = f"Go to {indicator}"
                contents_sheet[f'H{row_num}'] = self.sources[indicator]['url']
                
                # Make "Go to" link blue
                link_cell = contents_sheet[f'G{row_num}']
                link_cell.font = Font(color="0563C1", underline="single")
            
            # Set column widths to match other countries
            column_widths = {
                'A': 8,   # No.
                'B': 20,  # Data Type
                'C': 12,  # Status
                'D': 10,  # Records
                'E': 35,  # Date Range
                'F': 20,  # Last Extraction Time
                'G': 25,  # Sheet Link
                'H': 40   # Source URL
            }
            
            for col_letter, width in column_widths.items():
                contents_sheet.column_dimensions[col_letter].width = width
            
            print(f"   📋 Created Contents sheet with proper styling")
            
        except Exception as e:
            print(f"   ⚠️ Error creating Contents sheet: {e}")


# Include the same cleaning functions as before
def clean_thailand_data(input_file_path: str = None) -> Optional[str]:
    """Clean Thailand data: Filter to 2000+ and create cleaned file"""
    
    if input_file_path is None:
        data_dir = Path("./extracted_data")
        if not data_dir.exists():
            print("❌ Data directory not found")
            return None
        
        thailand_files = list(data_dir.glob("macro_data_thailand_*.xlsx"))
        if not thailand_files:
            print("❌ No Thailand files found")
            return None
        
        input_file_path = max(thailand_files, key=lambda x: x.stat().st_mtime)
        print(f"📄 Found Thailand file: {input_file_path.name}")
    
    input_path = Path(input_file_path)
    output_file_path = input_path.parent / f"cleaned_{input_path.name}"
    
    print(f"\n🧹 CLEANING THAILAND DATA")
    print("=" * 40)
    print(f"Input:  {input_path.name}")
    print(f"Output: {output_file_path.name}")
    print(f"Filter: 2000-01-01 onwards")
    
    try:
        excel_data = pd.read_excel(input_file_path, sheet_name=None)
        cleaned_data = {}
        
        for sheet_name, df in excel_data.items():
            if sheet_name != 'Contents':
                if 'date' in df.columns:
                    original_count = len(df)
                    df_filtered = df[df['date'] >= '2000-01-01'].copy()
                    cleaned_data[sheet_name] = df_filtered
                    
                    print(f"   📊 {sheet_name}: {original_count} → {len(df_filtered)} records")
                    if len(df_filtered) > 0:
                        print(f"       Range: {df_filtered['date'].min()} to {df_filtered['date'].max()}")
                else:
                    cleaned_data[sheet_name] = df
                    print(f"   ⚠️ {sheet_name}: No date column, kept as-is")
        
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            _create_updated_contents_sheet(writer, cleaned_data, input_file_path)
            
            for sheet_name, df in cleaned_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"\n✅ Cleaning completed!")
        print(f"📁 Output: {output_file_path.name}")
        return str(output_file_path)
        
    except Exception as e:
        print(f"❌ Cleaning error: {e}")
        return None


def _create_updated_contents_sheet(writer, cleaned_data: Dict[str, pd.DataFrame], input_file_path: str):
    """Create updated Contents sheet matching other countries' format"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        workbook = writer.book
        contents_sheet = workbook.create_sheet('Contents', 0)
        
        # Title (rows 1-2)
        contents_sheet['A1'] = "Thailand Macroeconomic Data (Cleaned - 2000+ Only)"
        contents_sheet.merge_cells('A1:H2')
        
        # Style title
        title_font = Font(size=16, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        contents_sheet['A1'].font = title_font
        contents_sheet['A1'].fill = title_fill
        contents_sheet['A1'].alignment = title_alignment
        
        # Headers (row 3) - NO CLEANING NOTE COLUMN
        headers = ['No.', 'Data Type', 'Status', 'Records', 'Date Range', 'Last Extraction Time', 'Sheet Link', 'Source URL']
        for i, header in enumerate(headers):
            cell = contents_sheet.cell(row=3, column=i+1, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Get source URLs from original Thailand sources
        source_urls = _get_source_urls_from_original_thailand(input_file_path)
        
        # Data rows (starting from row 4)
        for i, (sheet_name, df) in enumerate(cleaned_data.items(), 1):
            row_num = i + 3
            
            status = "✅ Success (Cleaned)" if len(df) > 0 else "⚠️ No data after 2000"
            date_range = f"{df['date'].min()} to {df['date'].max()}" if len(df) > 0 and 'date' in df.columns else "No date data"
            
            # Data
            contents_sheet[f'A{row_num}'] = i
            contents_sheet[f'B{row_num}'] = sheet_name
            contents_sheet[f'C{row_num}'] = status
            contents_sheet[f'D{row_num}'] = len(df)
            contents_sheet[f'E{row_num}'] = date_range
            contents_sheet[f'F{row_num}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            contents_sheet[f'G{row_num}'] = f"Go to {sheet_name}"
            contents_sheet[f'H{row_num}'] = source_urls.get(sheet_name, 'Unknown Source')
            
            # Make "Go to" link blue
            link_cell = contents_sheet[f'G{row_num}']
            link_cell.font = Font(color="0563C1", underline="single")
        
        # Set column widths to match other countries
        column_widths = {
            'A': 8,   # No.
            'B': 20,  # Data Type  
            'C': 15,  # Status
            'D': 10,  # Records
            'E': 35,  # Date Range
            'F': 20,  # Last Extraction Time
            'G': 25,  # Sheet Link
            'H': 40   # Source URL
        }
        
        for col_letter, width in column_widths.items():
            contents_sheet.column_dimensions[col_letter].width = width
        
        print("   📋 Created cleaned Contents sheet with proper styling and source URLs")
        
    except Exception as e:
        print(f"   ⚠️ Error creating cleaned Contents sheet: {e}")


def _get_source_urls_from_original_thailand(input_file_path: str) -> Dict[str, str]:
    """Extract source URLs from THAILAND_SOURCES directly"""
    source_urls = {}
    
    try:
        # Get URLs directly from THAILAND_SOURCES
        from macro_sources import THAILAND_SOURCES
        
        for source in THAILAND_SOURCES:
            source_urls[source.data_type] = source.url
        
        print(f"   📎 Loaded Thailand URLs from macro_sources.py: {len(source_urls)} sources")
        return source_urls
        
    except ImportError:
        print(f"   ⚠️ Could not import THAILAND_SOURCES from macro_sources.py")
        return {}
    except Exception as e:
        print(f"   ⚠️ Error loading Thailand source URLs: {e}")
        return {}

def main():
    """Main execution function"""
    print("🚀 THAILAND DATA EXTRACTOR (FIXED VERSION)")
    print("Sources: IMF API, Bank of Thailand, FRED API")
    print("=" * 50)
    
    try:
        extractor = ThailandDataExtractor()
        excel_file = extractor.extract_all_data()
        
        if excel_file:
            print(f"\n🎉 EXTRACTION COMPLETE!")
            print(f"📁 File: {Path(excel_file).name}")
            
            cleaned_file = clean_thailand_data(excel_file)
            if cleaned_file:
                print(f"\n🧹 CLEANING COMPLETE!")
                print(f"📁 Cleaned file: {Path(cleaned_file).name}")
            
        return excel_file
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return None


def extract_thailand_data() -> Optional[str]:
    """Convenience function to extract Thailand data"""
    extractor = ThailandDataExtractor()
    return extractor.extract_all_data()


if __name__ == "__main__":
    main()