import asyncio
import json
import os
from datetime import datetime
from typing import Dict, List, Any, Tuple
from openai import AsyncOpenAI
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Configuration
load_dotenv()
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY')
NEWS_DATA_FOLDER = "./news_data"
OUTPUT_FOLDER = "./news_analysis_output"

# Global URL cache for performance optimization
_url_cache = {}
_cache_file_path = None

# Economic indicators mapping for topic analysis
ECONOMIC_INDICATORS = {
    "GDP": ["GDP", "gross domestic product", "economic growth", "economic output"],
    "Inflation": ["inflation", "CPI", "consumer price index", "price level", "deflation"],
    "Interest Rates": ["interest rate", "central bank", "monetary policy", "fed", "BOE", "ECB", "BOJ"],
    "Employment": ["employment", "unemployment", "job", "labor market", "payroll"],
    "Trade": ["trade", "export", "import", "tariff", "trade deficit", "trade surplus"],
    "Currency": ["currency", "exchange rate", "USD", "EUR", "GBP", "JPY", "forex"],
    "Stock Market": ["stock", "equity", "share price", "market index", "S&P", "FTSE", "Nikkei"],
    "Housing": ["housing", "real estate", "property", "mortgage", "home price"],
    "Banking": ["bank", "financial institution", "credit", "lending", "deposit"],
    "Government Finance": ["budget deficit", "government debt", "fiscal policy", "public spending"],
    "Business Activity": ["PMI", "industrial production", "business confidence", "manufacturing"],
    "Consumer Spending": ["consumer confidence", "retail sales", "personal income", "consumption"],
    "Commodities": ["oil price", "commodity", "energy costs", "raw materials"],
    "International Finance": ["FDI", "current account", "balance of payments", "capital flows"],
}

# News sources configuration (expandable)
NEWS_SOURCES = {
    "bbc": {
        "name": "BBC Business",
        "pattern": "bbc_news_*.json",
        "website": "https://www.bbc.com/business"
    },
    "yahoo_economic": {
        "name": "Yahoo Finance-Economic",
        "pattern": "yahoo_economic_news_*.json",
        "website": "https://finance.yahoo.com/topic/economic-news/"
    },
    "yahoo_market": {
        "name": "Yahoo Finance-Stock Market", 
        "pattern": "yahoo_market_news_*.json",
        "website": "https://finance.yahoo.com/topic/stock-market-news/"
    },
    "business_times_sg": {
        "name": "Business Times Singapore",
        "pattern": "business_times_sg_news_*.json",
        "website": "https://www.businesstimes.com.sg/singapore"
    },
    
    # Future sources can be added here:
    # Examples:
    # "reuters": {"name": "Reuters", "pattern": "reuters_news_*.json", "website": "https://www.reuters.com"},
    # "bloomberg": {"name": "Bloomberg", "pattern": "bloomberg_news_*.json", "website": "https://www.bloomberg.com"}
}

openai_client = AsyncOpenAI(api_key=OPENAI_API_KEY)

class AnalysisConfig:
    """Configuration for analysis parameters"""
    SENTIMENT_ENSEMBLE_CALLS = int(os.getenv('SENTIMENT_ENSEMBLE_CALLS', '10'))  # Default 10, configurable via .env
    
    @classmethod
    def set_sentiment_calls(cls, num_calls: int):
        """Set number of sentiment calls programmatically"""
        cls.SENTIMENT_ENSEMBLE_CALLS = num_calls

def get_latest_news_file(source: str = "bbc") -> str:
    """Get the most recent news file for a given source"""
    if source not in NEWS_SOURCES:
        raise ValueError(f"Unknown news source: {source}")
    
    if not os.path.exists(NEWS_DATA_FOLDER):
        raise FileNotFoundError(f"Directory {NEWS_DATA_FOLDER} not found")
    
    pattern = NEWS_SOURCES[source]["pattern"]
    prefix = pattern.split("*")[0]
    
    matching_files = [f for f in os.listdir(NEWS_DATA_FOLDER) 
                     if f.startswith(prefix) and f.endswith(".json")]
    
    if not matching_files:
        raise FileNotFoundError(f"No news files found for source: {source}")
    
    latest_file = sorted(matching_files, reverse=True)[0]
    return os.path.join(NEWS_DATA_FOLDER, latest_file)

async def analyze_country_region(title: str, content: List[str]) -> str:
    """Identify the relevant countries/regions the news article impacts"""
    full_text = " ".join(content)
    
    # Countries from your project
    target_countries = [
        "Singapore", "United States", "Euro Area", "United Kingdom", "China", 
        "Japan", "India", "Indonesia", "Malaysia", "Thailand", "Vietnam"
    ]
    
    countries_text = ", ".join(target_countries)
    
    prompt = f"""
    You are an expert geographic economic analyst. Identify which countries/regions this news article significantly relates to or impacts.
    
    TARGET REGIONS: {countries_text}
    
    ANALYSIS CRITERIA:
    - Direct mention or focus on specific countries/economies
    - Economic policies that significantly impact certain regions
    - Trade relationships and bilateral/multilateral agreements
    - Regional economic integration and spillover effects
    - Market-specific impacts (local stock markets, currencies, etc.)
    - Supply chain or business operations affecting specific regions
    
    SELECTION RULES:
    - Choose 1-4 most relevant regions (maximum 4 to maintain focus)
    - Order by relevance: most relevant first
    - Include regions with both direct and significant indirect impacts
    - Use "Euro Area" for EU-wide policies/impacts
    - Use "United States" for US-wide policies/impacts
    - If no listed regions are significantly relevant, respond with "Other"
    - Avoid including regions with only minor or speculative connections
    
    REGIONAL IMPACT EXAMPLES:
    - Article about US-China trade war → "United States, China"
    - Article about ASEAN economic cooperation → "Singapore, Malaysia, Thailand, Indonesia"
    - Article about EU monetary policy → "Euro Area"
    - Article about global oil prices → "United States, China, Euro Area" (major economies)
    - Article about Singapore's fintech regulations → "Singapore"
    - Article about Brexit impact on Asia → "United Kingdom, Singapore" (if significant Asian focus)
    
    Article Title: {title}
    Article Content: {full_text}
    
    MANDATORY OUTPUT FORMAT:
    Respond with comma-separated region names ONLY, ordered by relevance.
    Examples: "Singapore" or "United States, China" or "Euro Area, United Kingdom"
    
    Do not include explanations, just the region names.
    """
    
    try:
        response = await openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=50,
            temperature=0
        )
        result = response.choices[0].message.content.strip()
        
        # Validation: ensure regions are from the target list
        regions = [r.strip() for r in result.split(',')]
        valid_regions = [r for r in regions if r in target_countries or r == "Other"]
        
        # Limit to maximum 4 regions
        if len(valid_regions) > 4:
            valid_regions = valid_regions[:4]
        
        return ', '.join(valid_regions) if valid_regions else "Other"
        
    except Exception as e:
        print(f"Error in country analysis: {e}")
        return "Unknown"

async def summarize_article(title: str, content: List[str]) -> str:
    """Generate a concise summary of the news article"""
    full_text = " ".join(content)
    
    prompt = f"""
    You are an expert economic news summarizer. Create a comprehensive yet concise summary of this news article focusing on economic and financial content.
    
    SUMMARY REQUIREMENTS:
    - Length: 3-4 sentences maximum
    - Focus ONLY on factual information explicitly stated in the article
    - Prioritize economic and financial information over general news
    - Include key economic data, figures, or metrics mentioned
    - Cover main economic impacts or implications stated in the article
    - Maintain objective, factual tone without interpretation or inference
    
    CONTENT PRIORITIES (include if mentioned in article):
    1. Economic data and financial metrics (GDP, inflation, employment, etc.)
    2. Policy decisions and regulatory changes
    3. Market movements and financial performance
    4. Corporate earnings, business developments, or strategic decisions
    5. Trade relationships and international economic developments
    6. Economic forecasts or projections explicitly stated
    7. Geographic regions and countries affected
    8. Time frames and effective dates mentioned
    
    STRICT GUIDELINES:
    - Use ONLY information explicitly stated in the article
    - Do NOT add background context, explanations, or implications not mentioned
    - Do NOT include your own analysis or interpretation
    - Do NOT speculate about future impacts beyond what the article states
    - Include specific numbers, percentages, or monetary amounts when mentioned
    - Mention key stakeholders (companies, governments, institutions) when relevant
    
    Article Title: {title}
    Article Content: {full_text}
    
    MANDATORY OUTPUT:
    Provide a factual summary in 3-4 complete sentences, focusing on economic and financial aspects explicitly mentioned in the article.
    """
    
    try:
        response = await openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=250,  # Increased for more comprehensive summaries
            temperature=0.1
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"Error in summarization: {e}")
        return "Summary generation failed"

async def analyze_news_impact(title: str, content: List[str]) -> str:
    """Analyze potential economic/financial impacts on various countries and markets"""
    full_text = " ".join(content)
    
    # Countries from your project
    target_countries = [
        "Singapore", "United States", "Euro Area", "United Kingdom", "China", 
        "Japan", "India", "Indonesia", "Malaysia", "Thailand", "Vietnam"
    ]
    
    # Financial markets from your project
    target_markets = [
        "Singapore STI", "S&P 500", "NASDAQ", "EURO STOXX 50", "Shanghai Composite", 
        "Hang Seng", "Nikkei 225", "VIX", "BSE Sensex", "FTSE 100", 
        "FTSE Bursa Malaysia KLCI", "Jakarta Composite", "SET Index"
    ]
    
    countries_text = ", ".join(target_countries)
    markets_text = ", ".join(target_markets)
    
    prompt = f"""
    You are an expert economic impact analyst. Analyze the potential economic and financial market impacts of this news article.
    
    TARGET REGIONS: {countries_text}
    TARGET MARKETS: {markets_text}
    
    Article Title: {title}
    Article Content: {full_text}...
    
    ANALYSIS REQUIREMENTS:
    1. Identify significantly impacted regions/markets only (skip minor or irrelevant impacts)
    2. For each impact, provide: impact direction with level + brief reasoning
    3. Impact directions with levels must be exactly one of these:
       - "Very Positive" - Major beneficial impact, significant economic gains
       - "Positive" - Clear beneficial impact, moderate economic gains
       - "Slightly Positive" - Minor beneficial impact, small economic gains
       - "Neutral" - No significant impact or balanced positive/negative effects
       - "Slightly Negative" - Minor adverse impact, small economic concerns
       - "Negative" - Clear adverse impact, moderate economic damage
       - "Very Negative" - Major adverse impact, significant economic damage
    4. Keep reasoning concise but comprehensive
    5. Ensure economic logic consistency with sentiment analysis
    
    MANDATORY OUTPUT FORMAT:
    Each impact must follow this EXACT format with double asterisks:
    **Country/Market: Impact Level** - Brief reasoning explaining the impact
    
    FORMATTING RULES:
    - Use ONLY double asterisks (**) for emphasis, NEVER use ### or # symbols
    - Each impact on a separate line
    - No bullet points, no numbering, no other formatting
    - Impact Level must be exactly one of the 7 levels listed above
    - Keep each reasoning to one sentence maximum
    
    CORRECT FORMAT EXAMPLES:
    **Singapore: Very Negative** - Trade-dependent economy faces severe disruption from global trade war escalation
    **United States: Positive** - Domestic manufacturing benefits from reduced foreign competition
    **S&P 500: Slightly Negative** - Mixed corporate earnings create minor downward pressure on markets
    **Euro Area: Neutral** - Balanced exposure to both positive and negative economic effects
    **China: Very Positive** - Major infrastructure investment significantly boosts economic growth prospects
    
    WRONG FORMATS TO AVOID:
    - ### Singapore: Negative ###
    - Singapore: Negative Impact
    - **Singapore: Moderately Negative**
    - • Singapore: Negative - reasoning
    - 1. Singapore: Negative
    
    Analyze only the most relevant impacts (typically 3-10 regions/markets). Focus on regions where the news has clear, logical economic consequences. Match the impact level to the severity and scope of expected economic effects.
    """
    
    try:
        response = await openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=500, 
            temperature=0.2
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"Error in news impact analysis: {e}")
        return "Impact analysis failed"

async def analyze_topics(title: str, content: List[str]) -> str:
    """Analyze topics and identify related economic indicators"""
    full_text = " ".join(content)
    indicators_text = "\n".join([f"- {k}: {', '.join(v[:3])}" for k, v in ECONOMIC_INDICATORS.items()])
    
    prompt = f"""
    You are an expert economic news analyst. Analyze this news article and provide a structured classification.

    TASK 1 - TOPIC CLASSIFICATION:
    Classify this article into the most relevant broad economic topics (maximum 5) from these categories:
    
    • Monetary Policy: Central bank actions, interest rate decisions, inflation control, quantitative easing
    • Trade & Commerce: International trade, tariffs, supply chains, trade agreements, import/export policies
    • Financial Markets: Stock markets, bond markets, currency markets, market volatility, investor sentiment
    • Economic Growth: GDP growth, recession/expansion, productivity, economic indicators, business cycles
    • Technology & Innovation: Fintech, digital economy, AI economic impact, tech sector developments
    • Energy & Commodities: Oil prices, renewable energy policies, commodity markets, energy security
    • Geopolitics & Economy: Trade wars, sanctions, international economic relations, political stability
    • Corporate & Business: Company earnings, mergers & acquisitions, business strategy, corporate governance
    • Real Estate & Housing: Property markets, housing policies, construction sector, mortgage markets
    • Government & Fiscal Policy: Government spending, taxation, budget policies, public debt, regulations
    • Labor & Employment: Job markets, unemployment, wage policies, labor disputes, workforce development
    • Consumer & Retail: Consumer spending, retail trends, consumer confidence, household economics

    Requirements:
    - Select ONLY from the categories above
    - Maximum 5 topics, minimum 1 topic
    - Order from MOST relevant to LEAST relevant
    - Choose the closest match if none perfectly fit

    TASK 2 - ECONOMIC INDICATORS:
    Identify the top economic indicators (maximum 5) most likely to be directly or indirectly impacted by this news:

    Available indicators:
    {indicators_text}

    Requirements:
    - Maximum 5 indicators from the list above
    - Order from MOST impacted to LEAST impacted
    - Consider both direct and indirect economic effects
    - If fewer than 5 relevant indicators exist in the list, you may create new relevant indicators based on article content
    - Focus on measurable economic metrics that could show changes due to this news

    Article Title: {title}
    Article Content: {full_text}

    CRITICAL OUTPUT FORMAT REQUIREMENTS:
    You MUST respond with EXACTLY this format with NO numbering, NO bullet points, NO other formatting:
    
    Topics: topic1, topic2, topic3, topic4, topic5
    Indicators: indicator1, indicator2, indicator3, indicator4, indicator5
    
    CORRECT EXAMPLES:
    Topics: Trade & Commerce, Geopolitics & Economy, Government & Fiscal Policy, Economic Growth, Corporate & Business
    Indicators: GDP, Employment, Trade, Stock Market, Currency
    
    WRONG FORMATS TO AVOID:
    - Topics: 1. Trade & Commerce 2. Geopolitics & Economy
    - Topics: • Trade & Commerce • Geopolitics & Economy
    - Topics: **Trade & Commerce**, **Geopolitics & Economy**
    - 1. Trade & Commerce 2. Geopolitics & Economy
    
    MANDATORY: Use ONLY comma-separated lists without any numbers, bullets, or special formatting.
    """
    
    try:
        response = await openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=150,
            temperature=0
        )
        
        result = response.choices[0].message.content.strip()
        
        # Post-processing to clean up any formatting issues
        lines = result.split('\n')
        cleaned_lines = []
        
        for line in lines:
            line = line.strip()
            if line.startswith('Topics:') or line.startswith('Indicators:'):
                # Remove any numbering or bullet points from the content
                if ':' in line:
                    prefix, content = line.split(':', 1)
                    content = content.strip()
                    
                    # Clean up numbered lists (1., 2., etc.)
                    import re
                    content = re.sub(r'\d+\.\s*', '', content)
                    
                    # Clean up bullet points
                    content = re.sub(r'[•\-\*]\s*', '', content)
                    
                    # Clean up extra formatting
                    content = content.replace('**', '').replace('*', '')
                    
                    # Reconstruct the line
                    cleaned_line = f"{prefix}: {content}"
                    cleaned_lines.append(cleaned_line)
        
        return '\n'.join(cleaned_lines) if cleaned_lines else result
        
    except Exception as e:
        print(f"Error in topic analysis: {e}")
        return "Topic analysis failed"

async def analyze_sentiment(title: str, content: List[str]) -> float:
    """Analyze sentiment and return only the score (for ensemble calls)"""
    full_text = " ".join(content)
    
    prompt = f"""
    You are an expert financial sentiment analyst. Analyze the sentiment of this news article specifically regarding its impact on the economy and financial markets.
    
    ANALYSIS FOCUS:
    - Economic growth prospects and business environment
    - Financial market stability and investor confidence
    - Trade, employment, and monetary policy implications
    - Regional economic impacts (Singapore, US, Euro Area, UK, China, Japan, India, Indonesia, Malaysia, Thailand, Vietnam)
    
    SENTIMENT SCORING:
    Provide a sentiment score from -1.0 to 1.0 based on economic/financial impact:
    • 0.7 to 1.0: Major economic gains, strong market rallies, significant positive policy changes
    • 0.4 to 0.6: Clear economic improvements, market optimism, beneficial developments
    • 0.2 to 0.3: Minor positive signals, cautious optimism, small improvements
    • -0.1 to 0.1: Mixed signals, uncertainty, or no significant economic impact
    • -0.3 to -0.2: Minor concerns, cautious pessimism, small negative developments
    • -0.6 to -0.4: Clear economic concerns, market pessimism, adverse developments
    • -1.0 to -0.7: Major economic threats, market crashes, severe policy failures
    
    IMPORTANT CONSIDERATIONS:
    - Focus on economic/financial impact, not just emotional tone
    - Consider both immediate and longer-term economic implications
    - Weight impacts on multiple regions and markets
    - Use increments of 0.1 only (e.g., -0.7, -0.6, -0.5, etc.)
    
    Article Title: {title}
    Article Content: {full_text}
    
    MANDATORY OUTPUT FORMAT:
    Respond with ONLY the numerical score (e.g., "0.3" or "-0.7").
    Do NOT include any labels or explanations.
    """
    
    try:
        response = await openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=10,  # Very short since we only need the score
            temperature=0.3
        )
        
        score_text = response.choices[0].message.content.strip()
        score = float(score_text)
        
        # Validation: ensure score is within bounds and round to 0.1
        score = round(score, 1)
        score = max(-1.0, min(1.0, score))
        
        return score
        
    except Exception as e:
        print(f"Error in sentiment score analysis: {e}")
        return 0.0

def standardize_sentiment_label(score: float) -> str:
    """Ensure consistent sentiment labels based on score"""
    if score <= -0.7:
        return "Very Negative"
    elif score <= -0.4:
        return "Negative"
    elif score < -0.1:
        return "Slightly Negative"
    elif score <= 0.1:
        return "Neutral"
    elif score < 0.4:
        return "Slightly Positive"
    elif score < 0.7:
        return "Positive"
    else:
        return "Very Positive"

async def analyze_sentiment_ensemble(title: str, content: List[str], num_calls: int = None) -> Tuple[float, str]:
    """Analyze sentiment with configurable ensemble calls and rate limiting"""
    if num_calls is None:
        num_calls = AnalysisConfig.SENTIMENT_ENSEMBLE_CALLS
    
    print(f"🔄 Running sentiment analysis with {num_calls} calls (with delays)...")
    
    # Make calls sequentially with delays instead of concurrently
    scores = []
    
    for i in range(num_calls):
        try:
            score = await analyze_sentiment(title, content)
            if isinstance(score, (int, float)) and -1.0 <= score <= 1.0:
                scores.append(float(score))
                print(f"   Call {i+1}/{num_calls}: {score}")
            
            # Add delay between calls (except for the last one)
            if i < num_calls - 1:
                await asyncio.sleep(0.5)  # 500ms delay between calls
                
        except Exception as e:
            print(f"   Call {i+1}/{num_calls} failed: {e}")
            continue
    
    if not scores:
        print("⚠️ All sentiment calls failed, using neutral fallback")
        return 0.0, "Neutral"
    
    # Get median score
    scores.sort()
    n = len(scores)
    if n % 2 == 0:
        median_score = (scores[n//2 - 1] + scores[n//2]) / 2
    else:
        median_score = scores[n//2]
    
    # Round to nearest 0.1
    median_score = round(median_score, 1)
    
    # Get corresponding label using existing function
    final_label = standardize_sentiment_label(median_score)
    
    print(f"📊 Sentiment ensemble results: {len(scores)}/{num_calls} successful calls, median: {median_score}, label: {final_label}")
    
    return median_score, final_label

def parse_published_date(date_str: str) -> datetime:
    """Parse published date string to datetime object"""
    try:
        if not date_str:
            return datetime.min
        # Handle the format "2025-07-14 05:46:00"
        return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
    except:
        try:
            # Fallback for other formats
            return datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        except:
            return datetime.min

def _get_existing_urls_optimized(file_path: str) -> set:
    """
    Get existing URLs with caching for maximum performance
    Only re-reads file if it has been modified
    """
    global _url_cache, _cache_file_path
    
    # Check if we need to refresh cache
    if _cache_file_path != file_path or file_path not in _url_cache:
        print(f"📊 Building URL cache from {file_path}")
        
        try:
            existing_urls = set()
            excel_file = pd.ExcelFile(file_path)
            
            for sheet_name in excel_file.sheet_names:
                if sheet_name == 'Contents':
                    continue
                    
                # Read only the URL column for maximum speed
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=2, usecols=['URL'])
                    urls = df['URL'].dropna().str.strip().str.lower()
                    existing_urls.update(urls)
                except Exception as e:
                    print(f"⚠️ Could not read URLs from {sheet_name}: {e}")
                    continue
            
            _url_cache[file_path] = existing_urls
            _cache_file_path = file_path
            print(f"✅ Cached {len(existing_urls)} URLs")
            
        except Exception as e:
            print(f"❌ Error building URL cache: {e}")
            return set()
    
    return _url_cache[file_path]

def _remove_duplicates_optimized(new_articles: List[Dict], file_path: str) -> List[Dict]:
    """
    Super-optimized duplicate removal using cached URL lookup
    """
    if not file_path or not os.path.exists(file_path):
        return new_articles
    
    # Get existing URLs with caching
    existing_urls = _get_existing_urls_optimized(file_path)
    
    if not existing_urls:
        return new_articles
    
    unique_articles = []
    duplicates_found = 0
    
    for article in new_articles:
        article_url = article.get('URL', '').strip().lower()
        
        if not article_url:
            print(f"⚠️ Skipping article without URL: {article.get('Title', 'Unknown')}")
            continue
        
        if article_url in existing_urls:
            duplicates_found += 1
            continue
            
        unique_articles.append(article)
        # Add to cache to prevent duplicates within same batch
        existing_urls.add(article_url)
    
    if duplicates_found > 0:
        print(f"🔄 Removed {duplicates_found} duplicate articles")
    
    print(f"✅ {len(unique_articles)} new unique articles to add")
    return unique_articles

def _load_existing_sheet_data(file_path: str, sheet_name: str) -> List[Dict]:
    """Load existing articles from a specific sheet"""
    try:
        df_existing = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=2)
        if len(df_existing) > 0:
            return df_existing.to_dict('records')
        return []
    except Exception as e:
        print(f"⚠️ Could not read existing data from {sheet_name}: {e}")
        return []

def _create_content_sheet(wb: Workbook, sheet_info: List[Dict]):
    """Create or update the content sheet with navigation - handles both new and existing files"""
    
    # Check if Contents sheet exists, create if not
    if 'Contents' not in wb.sheetnames:
        print("📝 Creating new Contents sheet")
        content_ws = wb.create_sheet(title='Contents', index=0)  # Insert as first sheet
        
        # Create the header structure for new sheet
        # Title row
        content_ws['A1'] = 'News Analysis Master File'
        content_ws.merge_cells('A1:G2')
        title_cell = content_ws['A1']
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers in row 3
        headers = ['No.', 'News Source', 'Articles', 'Date Range', 'Last Update', 'Sheet Link', 'Source URL']
        for col, header in enumerate(headers, 1):
            cell = content_ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
    else:
        print("📝 Updating existing Contents sheet")
        content_ws = wb['Contents']
    
    # Clear only data rows (keep title and headers)
    max_row = content_ws.max_row
    if max_row > 3:  # Only clear if there are data rows
        for row in range(4, max_row + 1):
            for col in range(1, 8):  # Assuming 7 columns
                content_ws.cell(row=row, column=col).value = None
    
    # Update data starting from row 4 (preserve existing formatting)
    for idx, info in enumerate(sheet_info, 1):
        row = idx + 3
        content_ws.cell(row=row, column=1, value=idx)
        content_ws.cell(row=row, column=2, value=info['source'])
        content_ws.cell(row=row, column=3, value=info['articles'])
        content_ws.cell(row=row, column=4, value=info['date_range'])
        content_ws.cell(row=row, column=5, value=info['last_extraction'])
        
        # Sheet link (preserve existing hyperlink formatting)
        sheet_link_cell = content_ws.cell(row=row, column=6, value=f"Go to {info['source']}")
        sheet_link_cell.hyperlink = f"#{info['sheet_link']}!A1"
        sheet_link_cell.font = Font(color='0000FF', underline='single')
        
        # Website URL (preserve existing hyperlink formatting)
        website_cell = content_ws.cell(row=row, column=7, value=info['website'])
        website_cell.hyperlink = info['website']
        website_cell.font = Font(color='0000FF', underline='single')

def _auto_adjust_content_sheet_widths(content_ws, sheet_info: List[Dict]):
    """Auto-adjust column widths for content sheet"""
    column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    
    for col_idx, column_letter in enumerate(column_letters, 1):
        max_length = 0
        
        # Check header length
        content_headers = ['No.', 'News Source', 'Articles', 'Date Range', 'Last Update', 'Sheet Link', 'Source URL']
        if col_idx <= len(content_headers):
            max_length = max(max_length, len(str(content_headers[col_idx - 1])))
        
        # Check data lengths
        for row_idx in range(4, len(sheet_info) + 4):
            try:
                cell = content_ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                continue
        
        # Set the column width
        adjusted_width = min(max_length + 2, 50)
        content_ws.column_dimensions[column_letter].width = adjusted_width

# Auto-adjust column widths for all columns
def _auto_adjust_all_column_widths(ws):
    """Auto-adjust column widths for all columns in the worksheet"""
    
    # Column letters mapping
    column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    
    for col_idx, column_letter in enumerate(column_letters, 1):
        max_length = 0
        
        # Check all rows in this column
        for row in range(1, ws.max_row + 1):
            try:
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None:
                    # Convert to string and get length
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            except:
                continue
        
        # Set column width with reasonable limits
        if max_length > 0:
            # Add padding and set reasonable min/max limits
            adjusted_width = max_length + 2
            
            # Set column-specific limits
            if col_idx == 1:  # Index column
                adjusted_width = min(max(adjusted_width, 6), 10)
            elif col_idx == 2:  # Title column  
                adjusted_width = min(max(adjusted_width, 20), 60)
            elif col_idx == 3:  # Description column
                adjusted_width = min(max(adjusted_width, 30), 80) 
            elif col_idx == 6:  # URL column
                adjusted_width = min(max(adjusted_width, 15), 40)
            else:
                adjusted_width = min(max(adjusted_width, 10), 50)
            
            ws.column_dimensions[column_letter].width = adjusted_width

def find_exact_sheet_name(workbook, source_display_name: str) -> str:
    """
    Find exact matching sheet name or use the source display name as-is
    """
    existing_sheets = workbook.sheetnames
    
    # Try exact match first
    if source_display_name in existing_sheets:
        print(f"✅ DEBUG: Found exact match sheet: '{source_display_name}'")
        return source_display_name
    
    # If no exact match, use the source display name as-is (no modifications)
    print(f"🆕 DEBUG: No existing sheet found, will create: '{source_display_name}'")
    return source_display_name

def append_articles_preserving_format(ws, existing_articles, new_articles, source_display_name):
    """Append new articles while preserving existing formatting"""
    
    # If this is a new sheet, create the structure
    if len(existing_articles) == 0:
        # Add title row
        ws['A1'] = source_display_name
        ws.merge_cells('A1:K2')  # Adjust range based on your column count
        title_cell = ws['A1']
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers in row 3
        headers = ['Index', 'Title', 'Description', 'Region', 'Published', 'URL', 
                  'Summary', 'Impact_Analysis', 'Topic_Analysis', 'Sentiment_Score', 'Sentiment_Label']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        start_row = 4
        next_index = 1
    else:
        # Find the last row with data
        start_row = ws.max_row + 1
        # Get the last index number
        last_index_cell = ws.cell(row=ws.max_row, column=1)
        try:
            next_index = int(last_index_cell.value) + 1
        except:
            next_index = len(existing_articles) + 1
    
    # Append only new articles
    for i, article in enumerate(new_articles):
        row = start_row + i
        
        # Add index and article data
        ws.cell(row=row, column=1, value=next_index + i)
        
        # Map article fields to columns
        field_mapping = {
            2: 'Title', 3: 'Description', 4: 'Region', 5: 'Published', 6: 'URL',
            7: 'Summary', 8: 'Impact_Analysis', 9: 'Topic_Analysis', 
            10: 'Sentiment_Score', 11: 'Sentiment_Label'
        }
        
        for col, field in field_mapping.items():
            value = article.get(field, '')
            cell = ws.cell(row=row, column=col, value=value)
            
            # FIXED: Add hyperlink formatting for URL column
            if field == 'URL' and value:
                cell.hyperlink = value
                cell.font = Font(color='0000FF', underline='single')
    
    # if len(new_articles) > 0 or len(existing_articles) == 0:
    #     _auto_adjust_all_column_widths(ws)

    if len(new_articles) > 0:
        # Calculate max width needed for Index column (column A)
        max_index_length = 0
        for row_num in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_num, column=1).value
            if cell_value is not None:
                max_index_length = max(max_index_length, len(str(cell_value)))
        
        # Set Index column width to fit content (minimum 6, maximum 10)
        index_width = min(max(max_index_length + 2, 6), 10)
        ws.column_dimensions['A'].width = index_width

def reorder_sheets_by_source_order(wb: Workbook):
    """Reorder sheets to match NEWS_SOURCES order"""
    try:
        # Desired order: Contents first, then sources in NEWS_SOURCES order
        desired_order = []
        
        # Contents sheet always first
        if 'Contents' in wb.sheetnames:
            desired_order.append('Contents')
        
        # Add source sheets in NEWS_SOURCES order
        for source_config in NEWS_SOURCES.values():
            source_name = source_config["name"]
            if source_name in wb.sheetnames:
                desired_order.append(source_name)
        
        # Add any other sheets that might exist
        for sheet_name in wb.sheetnames:
            if sheet_name not in desired_order:
                desired_order.append(sheet_name)
        
        # Reorder sheets
        for i, sheet_name in enumerate(desired_order):
            if sheet_name in wb.sheetnames:
                current_pos = wb.sheetnames.index(sheet_name)
                if current_pos != i:
                    wb.move_sheet(sheet_name, i)
                    print(f"📋 Moved {sheet_name} to position {i}")
        
        print(f"✅ Final sheet order: {wb.sheetnames}")
        
    except Exception as e:
        print(f"⚠️ Warning: Could not reorder sheets: {e}")

def create_excel_with_sheets(analyzed_data: Dict[str, List[Dict]], output_path: str = None, base_filename: str = "master_news_analysis.xlsx"):
    """
    Create or append to Excel file with content sheet and individual news source sheets
    Handles duplicate checking and maintains historical data with optimized performance
    
    Args:
        analyzed_data: Dictionary mapping source names to lists of analyzed articles
        output_path: Optional specific output path (if None, uses base_filename in default folder)
        base_filename: Default filename to use for master file
    """
    # Use a fixed filename instead of timestamp-based
    if output_path is None:
        output_folder = Path("./news_analysis_output")
        output_folder.mkdir(exist_ok=True)
        output_path = output_folder / base_filename
    
    # Check if file exists to determine if we're appending or creating new
    file_exists = os.path.exists(output_path)
    
    if file_exists:
        print(f"📁 Appending to existing file: {output_path}")
        wb = load_workbook(output_path)
    else:
        print(f"📁 Creating new file: {output_path}")
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
    
    # Store sheet information for content sheet
    sheet_info = []
    
    # Process each source
    for source_config in NEWS_SOURCES.values():
        source_display_name = source_config["name"]
        source_website = source_config["website"]
        print(f"\n📰 Processing {source_display_name}")

        articles = analyzed_data.get(source_display_name, [])
        
        # Define sheet_name
        if file_exists:
            sheet_name = find_exact_sheet_name(wb, source_display_name)
        else:
            sheet_name = source_display_name
        
        print(f"🔍 DEBUG: Using sheet name: '{sheet_name}' for source: '{source_display_name}'")
        
        # Load existing data from sheet if it exists
        existing_articles = []
        if file_exists and sheet_name in wb.sheetnames:
            existing_articles = _load_existing_sheet_data(str(output_path), sheet_name)
            print(f"📊 Found {len(existing_articles)} existing articles in {sheet_name}")

        # Handle the case of no new articles
        if not articles:
            print(f"⏭️ No articles for {source_display_name}")
            if existing_articles:
                df_existing = pd.DataFrame(existing_articles)
                date_range = f"{df_existing['Published'].min()} to {df_existing['Published'].max()}"
                article_count = len(existing_articles)
            else:
                date_range = "No articles"
                article_count = 0
            
            sheet_info.append({
                'source': source_display_name,
                'articles': article_count,
                'date_range': date_range,
                'last_extraction': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'sheet_link': sheet_name,
                'website': source_website
            })
            continue
                
        print(f"✅ Adding {len(articles)} new articles to {source_display_name}")
        
        # Use existing sheet or create new one
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"📝 Using existing sheet: {sheet_name}")
        else:
            ws = wb.create_sheet(title=sheet_name)
            print(f"📝 Creating new sheet: {sheet_name}")
        
        # Use append-only approach for sheet content
        append_articles_preserving_format(ws, existing_articles, articles, source_display_name)

        # ADDED: Auto-adjust column widths for existing sheets too
        # if len(existing_articles) > 0 and len(articles) == 0:
        #     _auto_adjust_all_column_widths(ws)
        
        # Sort the entire sheet by Published date while preserving format
        if ws.max_row > 3:
            # Get all data rows (starting from row 4)
            data_rows = []
            for row in range(4, ws.max_row + 1):
                row_data = []
                for col in range(1, 12):  # Columns A through K
                    row_data.append(ws.cell(row=row, column=col).value)
                data_rows.append(row_data)
            
            # Sort by Published date (Column E, index 4)
            try:
                data_rows.sort(key=lambda x: pd.to_datetime(x[4]) if x[4] else pd.to_datetime('1900-01-01'))
            except:
                print("⚠️ Warning: Could not sort by date")
            
            # Clear existing data rows and rewrite sorted data
            for row in range(4, ws.max_row + 1):
                for col in range(1, 12):
                    ws.cell(row=row, column=col).value = None
            
            # Write sorted data back with proper indexing
            for i, row_data in enumerate(data_rows, 1):
                row = i + 3  # Start from row 4
                ws.cell(row=row, column=1, value=i)  # Reindex sequentially
                for col, value in enumerate(row_data[1:], 2):
                    cell = ws.cell(row=row, column=col, value=value)
                    # Restore URL hyperlink formatting
                    if col == 6 and value:  # Column F (URL)
                        cell.hyperlink = value
                        cell.font = Font(color='0000FF', underline='single')
        
        # Get final article count for sheet_info
        final_article_count = ws.max_row - 3 if ws.max_row > 3 else 0
        
        # Calculate date range from the sorted data
        if final_article_count > 0:
            first_date = ws.cell(row=4, column=5).value
            last_date = ws.cell(row=ws.max_row, column=5).value
            date_range = f"{first_date} to {last_date}"
        else:
            date_range = "No articles"
            
        sheet_info.append({
            'source': source_display_name,
            'articles': final_article_count,
            'date_range': date_range,
            'last_extraction': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'sheet_link': sheet_name,
            'website': source_website
        })
    
    # ✅ REORDER ALL SHEETS AT THE END
    reorder_sheets_by_source_order(wb)
    
    # ✅ Update sheet_link references to match actual sheet names after reordering
    print("🔗 Updating sheet references...")
    for info in sheet_info:
        # The sheet name should still be the same, but let's make sure it exists
        actual_sheet_name = info['sheet_link']
        if actual_sheet_name in wb.sheetnames:
            # Sheet exists, hyperlink should work
            info['sheet_link'] = actual_sheet_name
            print(f"   ✅ {info['source']} -> {actual_sheet_name}")
        else:
            # Fallback: try to find the sheet by source name
            found_sheet = None
            for sheet_name in wb.sheetnames:
                if info['source'].lower() in sheet_name.lower():
                    found_sheet = sheet_name
                    break
            
            if found_sheet:
                info['sheet_link'] = found_sheet
                print(f"   🔧 {info['source']} -> {found_sheet} (corrected)")
            else:
                print(f"   ❌ Could not find sheet for {info['source']}")

    # Create or update content sheet
    _create_content_sheet(wb, sheet_info)
    
    # Auto-adjust content sheet column widths
    _auto_adjust_content_sheet_widths(wb['Contents'], sheet_info)
    
    # Save the workbook
    wb.save(output_path)
    print(f"\n✅ Excel file saved successfully: {output_path}")
    print(f"📊 Total sources: {len(sheet_info)}")
    total_articles = sum(info['articles'] for info in sheet_info)
    print(f"📰 Total articles: {total_articles}")
    
    # Clear cache after successful save to free memory
    global _url_cache, _cache_file_path
    _url_cache.clear()
    _cache_file_path = None
