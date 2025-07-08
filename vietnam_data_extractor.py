"""
Vietnam Data Extractor - FIXED VERSION
Sources: World Bank API (GDP, CPI) & FRED API (Population)
Output: Raw Excel file + Cleaned Excel file (filtered to 2000-01-01 onward)
"""

import pandas as pd
import numpy as np
import requests
import os
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Dict, Any, List
import warnings
import wbgapi as wb

# Suppress pandas warnings
warnings.filterwarnings('ignore', category=pd.errors.PerformanceWarning)

from dotenv import load_dotenv
load_dotenv()


class VietnamDataExtractor:
    """Extract macroeconomic data for Vietnam from World Bank and FRED APIs"""
    
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        
        # Vietnam data sources configuration
        self.vietnam_sources = {
            'GDP': {
                'wb_series_id': 'NY.GDP.MKTP.KN',
                'wb_country_code': 'VNM',
                'name': 'Vietnam Real GDP',
                'unit': 'Constant 2015 prices (VND)',
                'currency': 'VND',
                'source_url': 'https://databank.worldbank.org/reports.aspx?source=2&series=NY.GDP.MKTP.KN&country=VNM'
            },
            'CPI': {
                'wb_series_id': 'FP.CPI.TOTL',
                'wb_country_code': 'VNM', 
                'name': 'Vietnam Consumer Price Index',
                'unit': 'Index (2010 = 100)',
                'currency': None,
                'source_url': 'https://databank.worldbank.org/reports.aspx?source=2&series=FP.CPI.TOTL&country=VNM'
            },
            'Population': {
                'fred_series_id': 'POPTOTVNA647NWDB',
                'name': 'Vietnam Population',
                'unit': 'Number of persons',
                'currency': None,
                'source_url': 'https://fred.stlouisfed.org/series/POPTOTVNA647NWDB'
            }
        }
    
    def extract_all_data(self) -> Optional[str]:
        """Extract all Vietnam macroeconomic data and save to Excel (RAW, no filtering)"""
        print("🇻🇳 VIETNAM DATA EXTRACTION")
        print("Sources: World Bank API (GDP, CPI), FRED API (Population)")
        print("=" * 60)
        
        extracted_data = {}
        
        # Extract each indicator
        for indicator, config in self.vietnam_sources.items():
            print(f"\n📊 Extracting {indicator}...")
            
            if 'wb_series_id' in config:
                df = self._extract_world_bank_data(config, indicator)
            elif 'fred_series_id' in config:
                df = self._extract_fred_data(config, indicator)
            else:
                print(f"   ❌ No valid extraction method for {indicator}")
                continue
            
            if df is not None and len(df) > 0:
                # NO FILTERING - Keep raw data
                extracted_data[indicator] = df
                print(f"   ✅ Successfully extracted {len(df)} records (RAW)")
            else:
                print(f"   ❌ Failed to extract {indicator}")
        
        if not extracted_data:
            print("\n❌ No data extracted successfully")
            return None
        
        # Save RAW data to Excel
        return self._save_to_excel(extracted_data, is_cleaned=False)
    
    def _extract_world_bank_data(self, config: Dict, indicator: str) -> Optional[pd.DataFrame]:
        """Extract data from World Bank API using wbgapi with better error handling"""
        try:
            series_id = config['wb_series_id']
            country_code = config['wb_country_code']
            
            print(f"   🔄 World Bank API: {series_id} for {country_code}")
            
            try:
                print("   🔄 Getting all historical data...")
                data = wb.data.DataFrame(series_id, country_code)
                    
            except Exception as e:
                print(f"   ⚠️ World Bank API error: {e}")
                return None
            
            if data.empty:
                print(f"   ❌ No data returned from World Bank API")
                return None
            
            # Process the data
            records = []
            
            # Handle different data structures
            if hasattr(data, 'reset_index'):
                data_reset = data.reset_index()
                
                # Check if data has year columns or time index
                for col in data_reset.columns:
                    if col not in ['Country', 'economy', 'time']:
                        try:
                            # Try to parse as year
                            year = int(str(col).replace('YR', ''))
                            if 1960 <= year <= datetime.now().year:
                                for _, row in data_reset.iterrows():
                                    value = row[col]
                                    if pd.notna(value) and value != 0:
                                        date_str = f"{year}-12-31"
                                        records.append({
                                            'date': date_str,
                                            'value': float(value)
                                        })
                        except (ValueError, TypeError):
                            continue
                
                # Alternative: if time column exists
                if 'time' in data_reset.columns:
                    value_col = None
                    for col in data_reset.columns:
                        if col not in ['Country', 'economy', 'time']:
                            value_col = col
                            break
                    
                    if value_col:
                        for _, row in data_reset.iterrows():
                            try:
                                year = int(row['time'])
                                value = row[value_col]
                                if pd.notna(value) and value != 0:
                                    date_str = f"{year}-12-31"
                                    records.append({
                                        'date': date_str,
                                        'value': float(value)
                                    })
                            except (ValueError, TypeError):
                                continue
            
            if not records:
                print(f"   ❌ No valid records found after processing")
                return None
            
            df = pd.DataFrame(records)
            df['date'] = pd.to_datetime(df['date'])
            df = df.sort_values('date').reset_index(drop=True)
            df['date'] = df['date'].dt.strftime('%Y-%m-%d')
            
            print(f"   ✅ Extracted {len(df)} records from World Bank")
            return df
            
        except Exception as e:
            print(f"   ❌ World Bank API error: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _extract_fred_data(self, config: Dict, indicator: str) -> Optional[pd.DataFrame]:
        """Extract data from FRED API"""
        fred_api_key = os.getenv('FRED_API_KEY')
        if not fred_api_key:
            print(f"   ❌ FRED API key not found in environment")
            return None
        
        try:
            series_id = config['fred_series_id']
            api_url = f"https://api.stlouisfed.org/fred/series/observations?series_id={series_id}&api_key={fred_api_key}&file_type=json"
            
            print(f"   🔄 FRED API: {series_id}")
            
            response = self.session.get(api_url, timeout=30)
            response.raise_for_status()
            
            data = response.json()
            
            if 'observations' not in data:
                print(f"   ❌ No observations in FRED response")
                return None
            
            observations = data['observations']
            records = []
            
            for obs in observations:
                try:
                    date_str = obs['date']
                    value_str = obs['value']
                    
                    # Skip missing values
                    if value_str == '.' or value_str is None:
                        continue
                    
                    value = float(value_str)
                    records.append({
                        'date': date_str,
                        'value': value
                    })
                    
                except (ValueError, TypeError):
                    continue
            
            if not records:
                print(f"   ❌ No valid records found")
                return None
            
            df = pd.DataFrame(records)
            print(f"   ✅ Extracted {len(df)} records from FRED")
            return df
            
        except Exception as e:
            print(f"   ❌ FRED API error: {e}")
            return None
    
    def _create_standardized_dataframe(self, df: pd.DataFrame, indicator: str) -> pd.DataFrame:
        """Create standardized DataFrame with metadata"""
        config = self.vietnam_sources[indicator]
        
        # Create standardized DataFrame
        standardized_df = pd.DataFrame()
        
        # Required columns in specific order
        standardized_df['source_name'] = [config['name']] * len(df)
        standardized_df['series_id'] = [config.get('wb_series_id') or config.get('fred_series_id', 'Unknown')] * len(df)
        standardized_df['country'] = ['Vietnam'] * len(df)
        standardized_df['data_type'] = [indicator] * len(df)
        standardized_df['date'] = df['date'].values
        standardized_df['value'] = df['value'].values
        standardized_df['unit'] = [config['unit']] * len(df)
        
        # Add currency for monetary data
        if config.get('currency'):
            standardized_df['currency'] = [config['currency']] * len(df)
        
        # Add extraction metadata
        standardized_df['extraction_time'] = [datetime.now().isoformat()] * len(df)
        
        return standardized_df
    
    def _save_to_excel(self, extracted_data: Dict[str, pd.DataFrame], is_cleaned: bool = False) -> Optional[str]:
        """Save extracted data to Excel file with proper Contents sheet"""
        try:
            # Create timestamp for filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Create output directory if it doesn't exist
            output_dir = Path("./extracted_data")
            output_dir.mkdir(exist_ok=True)
            
            # Create filename
            prefix = "cleaned_macro_data" if is_cleaned else "macro_data"
            filename = f"{prefix}_vietnam_{timestamp}.xlsx"
            filepath = output_dir / filename
            
            # Prepare contents data and sheets
            contents_data = []
            sheet_data = {}
            
            for i, indicator in enumerate(['GDP', 'CPI', 'Population'], 1):  # Fixed order with numbering
                if indicator not in extracted_data:
                    continue
                    
                df = extracted_data[indicator]
                config = self.vietnam_sources[indicator]
                
                # Create standardized DataFrame
                std_df = self._create_standardized_dataframe(df, indicator)
                
                # Store sheet data
                sheet_data[indicator] = std_df
                
                # Add to contents - MATCHING US FORMAT EXACTLY
                contents_data.append({
                    'No.': i,
                    'Data Type': indicator,
                    'Status': '☑ Success',  # Changed to checkbox style
                    'Records': len(std_df),
                    'Date Range': f"{std_df['date'].iloc[0]} to {std_df['date'].iloc[-1]}",
                    'Last Extraction Time': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'Sheet Link': f"Go to {indicator}",  # Will be converted to hyperlink
                    'Source URL': config['source_url']
                })
            
            # Create Excel file with proper formatting
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                
                # 1. Write Contents sheet with US-style format directly
                self._write_formatted_contents_sheet(writer, contents_data)
                
                # 2. Write data sheets
                for sheet_name, std_df in sheet_data.items():
                    std_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"   📄 Written sheet: {sheet_name} ({len(std_df)} records)")
            
            file_type = "cleaned" if is_cleaned else "raw"
            print(f"\n📁 {file_type.title()} data saved to: {filename}")
            print(f"📊 Total indicators: {len(extracted_data)}")
            total_records = sum(len(df) for df in extracted_data.values())
            print(f"📈 Total records: {total_records}")
            
            return str(filepath)
            
        except Exception as e:
            print(f"❌ Error saving to Excel: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _write_formatted_contents_sheet(self, writer, contents_data):
        """Write Contents sheet with proper US-style formatting from the start"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # Create worksheet
            workbook = writer.book
            contents_sheet = workbook.create_sheet('Contents', 0)  # Insert at position 0
            
            # ROW 1-2: Title
            contents_sheet['A1'] = "Vietnam Macroeconomic Data"
            contents_sheet.merge_cells('A1:H2')
            
            # Style title
            title_font = Font(size=18, bold=True, color="FFFFFF")
            title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            title_alignment = Alignment(horizontal="center", vertical="center")
            contents_sheet['A1'].font = title_font
            contents_sheet['A1'].fill = title_fill
            contents_sheet['A1'].alignment = title_alignment
            
            # ROW 3: Headers
            headers = ['No.', 'Data Type', 'Status', 'Records', 'Date Range', 'Last Extraction Time', 'Sheet Link', 'Source URL']
            for col, header in enumerate(headers, 1):
                cell = contents_sheet.cell(row=3, column=col)
                cell.value = header
                
                # Style header
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # ROW 4+: Data with hyperlinks
            from openpyxl.styles import Font as OpenpyxlFont
            
            for i, row_data in enumerate(contents_data, 4):
                for col, (key, value) in enumerate(row_data.items(), 1):
                    cell = contents_sheet.cell(row=i, column=col, value=value)
                    
                    # Special handling for Sheet Link column (column 7)
                    if key == 'Sheet Link' and col == 7:
                        # Extract indicator name from "Go to {indicator}"
                        indicator_name = value.replace("Go to ", "")
                        
                        # Create hyperlink formula
                        hyperlink_formula = f'=HYPERLINK("#{indicator_name}!A1","{value}")'
                        cell.value = hyperlink_formula
                        
                        # Style as blue underlined link
                        link_font = OpenpyxlFont(color="0563C1", underline="single")
                        cell.font = link_font
            
            # Set column widths
            column_widths = {
                'A': 8,   # No.
                'B': 15,  # Data Type
                'C': 12,  # Status
                'D': 10,  # Records
                'E': 25,  # Date Range
                'F': 20,  # Last Extraction Time
                'G': 15,  # Sheet Link
                'H': 60   # Source URL
            }
            
            for col_letter, width in column_widths.items():
                contents_sheet.column_dimensions[col_letter].width = width
            
            print("   ✅ Contents sheet formatted successfully")
            
        except Exception as e:
            print(f"   ❌ Error formatting Contents sheet: {e}")
            import traceback
            traceback.print_exc()


def clean_vietnam_data(input_file: str) -> Optional[str]:
    """Clean Vietnam data - filter to 2000-01-01 onward and create cleaned file"""
    print(f"\n🧹 VIETNAM DATA CLEANING")
    print(f"Input: {Path(input_file).name}")
    print("=" * 50)
    
    try:
        # Read the Excel file - get all sheet names first
        all_sheets = pd.ExcelFile(input_file).sheet_names
        print(f"   📋 Found sheets: {all_sheets}")
        
        # Read all sheets
        data_sheets = pd.read_excel(input_file, sheet_name=None)
        
        # Filter date for cleaning
        filter_date = datetime(2000, 1, 1)
        cleaned_data = {}
        
        # Process each data sheet
        for sheet_name, df in data_sheets.items():
            if sheet_name == 'Contents':
                print(f"   ⏭️ Skipping Contents sheet")
                continue
            
            print(f"   🔍 Processing sheet: {sheet_name} ({len(df)} records)")
            
            # Check if this is a data sheet with date column
            if 'date' not in df.columns:
                print(f"   ⚠️ No date column in {sheet_name}, skipping")
                continue
            
            indicator = sheet_name  # GDP, CPI, Population
            
            print(f"   🧹 Cleaning {indicator}...")
            
            # Convert date and filter
            df_copy = df.copy()
            df_copy['date'] = pd.to_datetime(df_copy['date'])
            
            # Filter to 2000+
            df_filtered = df_copy[df_copy['date'] >= filter_date].copy()
            
            # Convert date back to string
            df_filtered['date'] = df_filtered['date'].dt.strftime('%Y-%m-%d')
            
            # Sort and reset index
            df_filtered = df_filtered.sort_values('date').reset_index(drop=True)
            
            # Add cleaning metadata
            if 'cleaning_applied' not in df_filtered.columns:
                df_filtered['cleaning_applied'] = 'Date filtering (2000+), missing value removal'
            
            cleaned_data[indicator] = df_filtered
            
            removed_count = len(df) - len(df_filtered)
            print(f"   ✅ {indicator}: {len(df_filtered)} records (removed {removed_count})")
        
        if not cleaned_data:
            print("   ❌ No data to clean - no valid data sheets found")
            return None
        
        print(f"   📊 Total cleaned datasets: {len(cleaned_data)}")
        
        # Save cleaned data using the same extractor method
        extractor = VietnamDataExtractor()
        cleaned_file = extractor._save_to_excel(cleaned_data, is_cleaned=True)
        
        return cleaned_file
        
    except Exception as e:
        print(f"   ❌ Error cleaning data: {e}")
        import traceback
        traceback.print_exc()
        return None


def main():
    """Main execution function"""
    print("🚀 VIETNAM DATA EXTRACTOR")
    print("Sources: World Bank API (GDP, CPI), FRED API (Population)")
    print("=" * 60)
    
    try:       
        # Extract RAW data
        extractor = VietnamDataExtractor()
        excel_file = extractor.extract_all_data()
        
        if excel_file:
            print(f"\n🎉 RAW EXTRACTION COMPLETE!")
            print(f"📁 File: {Path(excel_file).name}")
            
            # Clean the data
            cleaned_file = clean_vietnam_data(excel_file)
            if cleaned_file:
                print(f"\n🧹 CLEANING COMPLETE!")
                print(f"📁 Cleaned file: {Path(cleaned_file).name}")
            else:
                print(f"\n❌ CLEANING FAILED!")
            
        return excel_file
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return None


# Convenience functions
def extract_vietnam_data() -> Optional[str]:
    """Convenience function to extract Vietnam data"""
    extractor = VietnamDataExtractor()
    return extractor.extract_all_data()


if __name__ == "__main__":
    main()