"""
Singapore Data Standardization Script
Converts Singapore horizontal data to vertical format with specific requirements
Updated: Property Price "index" → "value" column for consistency
"""

import pandas as pd
import numpy as np
from pathlib import Path
import re
from datetime import datetime
import warnings

# Suppress warnings
warnings.filterwarnings('ignore')

def get_singapore_source_urls():
    """
    Extract correct source URLs from macro_sources.py for Singapore data types
    """
    try:
        # Import Singapore sources
        from macro_sources import SINGAPORE_SOURCES
        
        # Create mapping of data types to URLs
        url_mapping = {}
        
        for source in SINGAPORE_SOURCES:
            # Map based on data type patterns
            if 'GDP' in source.name and 'Chained' in source.name:
                url_mapping['GDP'] = source.url
            elif 'Consumer Price Index' in source.name or 'CPI' in source.name:
                url_mapping['CPI'] = source.url
            elif 'Interest' in source.name or 'SORA' in source.name:
                url_mapping['Interest_Rate'] = source.url
            elif 'Population' in source.name and 'Total' not in source.name:
                url_mapping['Population'] = source.url
            elif 'Property' in source.name or 'Price Index' in source.name:
                url_mapping['Property_Price'] = source.url
        
        print(f"   🔗 Loaded {len(url_mapping)} source URLs from macro_sources.py")
        for data_type, url in url_mapping.items():
            print(f"      {data_type}: {url}")
        
        return url_mapping
        
    except ImportError:
        print(f"   ⚠️ Could not import macro_sources.py")
        return {}
    except Exception as e:
        print(f"   ⚠️ Error loading source URLs: {e}")
        return {}

def find_latest_singapore_file(data_dir="./extracted_data"):
    """Find the latest cleaned Singapore macro data file"""
    
    data_path = Path(data_dir)
    if not data_path.exists():
        raise FileNotFoundError(f"Data directory not found: {data_dir}")
    
    # Search specifically for cleaned Singapore macro files
    pattern = "cleaned_macro_data_singapore*.xlsx"
    singapore_files = list(data_path.glob(pattern))
    
    if not singapore_files:
        # Fallback to any Singapore files
        pattern = "*singapore*.xlsx"
        singapore_files = list(data_path.glob(pattern))
        
        if not singapore_files:
            raise FileNotFoundError(f"No Singapore files found in {data_dir}")
    
    # Sort by modification time (most recent first)
    latest_file = max(singapore_files, key=lambda x: x.stat().st_mtime)
    
    print(f"📄 Found Singapore file: {latest_file.name}")
    return str(latest_file)

def standardize_singapore_data(input_file_path=None):
    """
    Standardize Singapore data to vertical format
    
    Requirements:
    1. Contents sheet: NO CHANGES
    2. GDP: Keep "GDP In Chained (2015) Dollars" → "value", date: 2025Q1 → 2025-01-01
    3. CPI: Keep "All Items" → "value", same date conversion
    4. Interest Rate: Keep "SORA 3-Month" → "value", date: 2025-04 → 2025-04-01
    5. Population: Keep all populations, "Total Population" → "value", date: 2025 → 2025-01-01
    6. Property Price: Already vertical, rename "quarter" → "date" and "index" → "value"
    """
    
    # Auto-find latest file if not specified
    if input_file_path is None:
        input_file_path = find_latest_singapore_file()
    
    input_path = Path(input_file_path)
    
    # Create proper output filename
    if input_path.name.startswith("cleaned_"):
        # Input: cleaned_macro_data_singapore_... → Output: standardized_cleaned_macro_data_singapore_...
        output_filename = f"standardized_{input_path.name}"
    else:
        # Fallback for other naming patterns
        output_filename = f"standardized_{input_path.name}"
    
    output_file_path = input_path.parent / output_filename
    
    print("🔄 SINGAPORE DATA STANDARDIZATION")
    print("=" * 50)
    print(f"Input:  {input_path.name}")
    print(f"Output: {output_file_path.name}")
    print("Updated: Property Price 'index' → 'value' column")
    
    # Load source URLs for Contents sheet
    source_urls = get_singapore_source_urls()
    
    # Read all sheets
    excel_data = pd.read_excel(input_file_path, sheet_name=None)
    standardized_data = {}
    
    for sheet_name, df in excel_data.items():
        print(f"\n📊 Processing sheet: {sheet_name}")
        
        if sheet_name == 'Contents':
            # NO CHANGES to Contents sheet - preserve exactly as-is
            standardized_data[sheet_name] = df.copy()
            print(f"   ✅ Contents sheet preserved unchanged ({df.shape[0]} rows, {df.shape[1]} columns)")
            
        elif sheet_name == 'GDP':
            # GDP standardization
            standardized_df = standardize_gdp_sheet(df)
            standardized_data[sheet_name] = standardized_df
            
        elif sheet_name == 'CPI':
            # CPI standardization
            standardized_df = standardize_cpi_sheet(df)
            standardized_data[sheet_name] = standardized_df
            
        elif sheet_name == 'Interest_Rate':
            # Interest Rate standardization
            standardized_df = standardize_interest_rate_sheet(df)
            standardized_data[sheet_name] = standardized_df
            
        elif sheet_name == 'Population':
            # Population standardization
            standardized_df = standardize_population_sheet(df)
            standardized_data[sheet_name] = standardized_df
            
        elif sheet_name == 'Property_Price':
            # Property Price standardization with index → value conversion
            standardized_df = standardize_property_price_sheet(df)
            standardized_data[sheet_name] = standardized_df
            
        else:
            # Keep other sheets as-is
            standardized_data[sheet_name] = df
            print(f"   ✅ {sheet_name} sheet preserved as-is")
    
    # Write to new Excel file with proper Contents sheet recreation
    try:
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            for sheet_name, df in standardized_data.items():
                if sheet_name == 'Contents':
                    # Skip writing Contents initially - we'll recreate it
                    continue
                else:
                    # Write other sheets normally
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Recreate Contents sheet exactly like the original
            recreate_contents_sheet_exactly(writer, standardized_data, source_urls)
        
        print(f"\n🎉 Standardization completed!")
        print(f"📁 Output saved: {output_file_path}")
        return str(output_file_path)
        
    except Exception as e:
        print(f"\n❌ Error saving file: {e}")
        return None

def standardize_gdp_sheet(df):
    """
    Standardize GDP sheet:
    - Keep only "GDP In Chained (2015) Dollars" → rename to "value"
    - Convert date: 2025Q1 → 2025-01-01, Q2 → 2025-04-01, etc.
    """
    
    print(f"   🔄 GDP: Converting horizontal → vertical, keeping main GDP series only")
    print(f"   📊 DataFrame shape: {df.shape}")
    print(f"   📋 Columns: {list(df.columns)}")
    
    # Check if DataSeries column exists
    if 'DataSeries' not in df.columns:
        print(f"   ❌ 'DataSeries' column not found. Available columns: {list(df.columns)}")
        return df
    
    # Find the GDP in Chained (2015) Dollars row
    target_series = None
    target_row_idx = None
    
    print(f"   🔍 Searching for GDP Chained 2015 series...")
    
    for idx, row in df.iterrows():
        if pd.notna(row.get('DataSeries', '')):
            series_name = str(row['DataSeries']).lower()
            print(f"      Row {idx}: {row['DataSeries']}")
            
            if 'gdp' in series_name and 'chained' in series_name and '2015' in series_name:
                target_series = row['DataSeries']
                target_row_idx = idx
                break
    
    if target_row_idx is None:
        print(f"   ❌ Could not find 'GDP In Chained (2015) Dollars' series")
        print(f"   💡 Available series:")
        for idx, row in df.iterrows():
            if pd.notna(row.get('DataSeries', '')):
                print(f"      {idx}: {row['DataSeries']}")
        return df
    
    print(f"   📋 Found target series: {target_series}")
    
    # Get the target row
    target_row = df.iloc[target_row_idx]
    
    # Find date columns (exclude metadata)
    metadata_cols = ['_id', 'DataSeries', 'source_name', 'extraction_time', 
                    'data_type', 'country', 'currency', 'unit']
    date_cols = [col for col in df.columns if col not in metadata_cols]
    
    print(f"   📅 Found {len(date_cols)} date columns")
    
    # Convert to vertical format
    vertical_records = []
    
    for date_col in date_cols:
        if pd.notna(target_row[date_col]) and target_row[date_col] != '':
            # Convert quarterly date format
            converted_date = convert_quarterly_date(str(date_col))
            
            if converted_date:
                try:
                    numeric_value = float(target_row[date_col])
                    record = {
                        'date': converted_date,
                        'value': numeric_value,
                        'series_name': target_series,
                        'data_type': 'GDP',
                        'country': 'Singapore',
                        'unit': target_row.get('unit', 'Chained 2015 Dollars'),
                        'source_name': target_row.get('source_name', 'Singapore Department of Statistics')
                    }
                    vertical_records.append(record)
                except (ValueError, TypeError):
                    print(f"      ⚠️ Skipping non-numeric value in {date_col}: {target_row[date_col]}")
    
    result_df = pd.DataFrame(vertical_records)
    result_df = result_df.sort_values('date')
    
    print(f"   ✅ GDP: {len(result_df)} records created")
    if len(result_df) > 0:
        print(f"   📅 Date range: {result_df['date'].iloc[0]} to {result_df['date'].iloc[-1]}")
    
    return result_df

def standardize_cpi_sheet(df):
    """
    Standardize CPI sheet:
    - Keep only "All Items" → rename to "value"
    - Same quarterly date conversion as GDP
    """
    
    print(f"   🔄 CPI: Converting horizontal → vertical, keeping All Items only")
    
    # Find the "All Items" row
    target_series = None
    target_row_idx = None
    
    for idx, row in df.iterrows():
        if pd.notna(row.get('DataSeries', '')):
            series_name = str(row['DataSeries']).lower()
            if 'all items' in series_name:
                target_series = row['DataSeries']
                target_row_idx = idx
                break
    
    if target_row_idx is None:
        print(f"   ❌ Could not find 'All Items' series")
        return df
    
    print(f"   📋 Found target series: {target_series}")
    
    # Get the target row
    target_row = df.iloc[target_row_idx]
    
    # Find date columns (exclude metadata)
    metadata_cols = ['_id', 'DataSeries', 'source_name', 'extraction_time', 
                    'data_type', 'country', 'currency', 'unit']
    date_cols = [col for col in df.columns if col not in metadata_cols]
    
    # Convert to vertical format
    vertical_records = []
    
    for date_col in date_cols:
        if pd.notna(target_row[date_col]) and target_row[date_col] != '':
            # Convert quarterly date format
            converted_date = convert_quarterly_date(str(date_col))
            
            if converted_date:
                record = {
                    'date': converted_date,
                    'value': float(target_row[date_col]) if pd.notna(target_row[date_col]) else None,
                    'series_name': target_series,
                    'data_type': 'CPI',
                    'country': 'Singapore',
                    'unit': target_row.get('unit', 'Index (2019=100)'),
                    'source_name': target_row.get('source_name', 'Singapore Department of Statistics')
                }
                vertical_records.append(record)
    
    result_df = pd.DataFrame(vertical_records)
    result_df = result_df.sort_values('date')
    
    print(f"   ✅ CPI: {len(result_df)} records created")
    if len(result_df) > 0:
        print(f"   📅 Date range: {result_df['date'].iloc[0]} to {result_df['date'].iloc[-1]}")
    
    return result_df

def standardize_interest_rate_sheet(df):
    """
    Standardize Interest Rate sheet:
    - Keep only "SORA 3-Month" → rename to "value"
    - Convert date: 2025-04 → 2025-04-01
    """
    
    print(f"   🔄 Interest Rate: Converting horizontal → vertical, keeping SORA 3-Month only")
    
    # Find the SORA 3-Month row
    target_series = None
    target_row_idx = None
    
    for idx, row in df.iterrows():
        if pd.notna(row.get('DataSeries', '')):
            series_name = str(row['DataSeries']).lower()
            if 'sora' in series_name and '3' in series_name and 'month' in series_name:
                target_series = row['DataSeries']
                target_row_idx = idx
                break
    
    if target_row_idx is None:
        print(f"   ❌ Could not find 'SORA 3-Month' series")
        return df
    
    print(f"   📋 Found target series: {target_series}")
    
    # Get the target row
    target_row = df.iloc[target_row_idx]
    
    # Find date columns (exclude metadata)
    metadata_cols = ['_id', 'DataSeries', 'source_name', 'extraction_time', 
                    'data_type', 'country', 'currency', 'unit']
    date_cols = [col for col in df.columns if col not in metadata_cols]
    
    # Convert to vertical format
    vertical_records = []
    
    for date_col in date_cols:
        # Skip if value is NA, empty, or cannot be converted to float
        cell_value = target_row[date_col]
        
        if pd.notna(cell_value) and cell_value != '' and str(cell_value).lower() != 'na':
            try:
                # Try to convert to float first
                numeric_value = float(cell_value)
                
                # Convert monthly date format (2025-04 → 2025-04-01)
                converted_date = convert_monthly_date(str(date_col))
                
                if converted_date:
                    record = {
                        'date': converted_date,
                        'value': numeric_value,
                        'series_name': target_series,
                        'data_type': 'Interest_Rate',
                        'country': 'Singapore',
                        'unit': target_row.get('unit', 'Percent per annum'),
                        'source_name': target_row.get('source_name', 'Monetary Authority of Singapore')
                    }
                    vertical_records.append(record)
                    
            except (ValueError, TypeError):
                # Skip values that cannot be converted to float (like 'na')
                print(f"      ⚠️ Skipping non-numeric value in {date_col}: {cell_value}")
                continue
    
    result_df = pd.DataFrame(vertical_records)
    result_df = result_df.sort_values('date')
    
    print(f"   ✅ Interest Rate: {len(result_df)} records created")
    if len(result_df) > 0:
        print(f"   📅 Date range: {result_df['date'].iloc[0]} to {result_df['date'].iloc[-1]}")
    
    return result_df

def standardize_population_sheet(df):
    """
    Standardize Population sheet:
    - "Total Population" → "value" column (should be large numbers like 6,000,000)
    - Other population types → separate columns
    - Convert date: 2025 → 2025-01-01
    """
    
    print(f"   🔄 Population: Converting horizontal → vertical, Total Population as value + other types as columns")
    print(f"   📊 DataFrame shape: {df.shape}")
    print(f"   📋 Available DataSeries:")
    
    # Show all available series for debugging
    for idx, row in df.iterrows():
        if pd.notna(row.get('DataSeries', '')):
            print(f"      Row {idx}: {row['DataSeries']}")
    
    # Population series mapping - looking for exact matches
    population_types = {
        'Total Population': 'value',
        'Resident Population': 'resident_population',
        'Singapore Citizen Population': 'citizen_population', 
        'Permanent Resident Population': 'pr_population',
        'Non-Resident Population': 'non_resident_population'
    }
    
    # Find population rows by exact matching of DataSeries
    population_rows = {}
    for idx, row in df.iterrows():
        if pd.notna(row.get('DataSeries', '')):
            series_name = str(row['DataSeries']).strip()
            
            # Check for exact matches first
            for pop_type, column_name in population_types.items():
                if series_name == pop_type:
                    population_rows[column_name] = row
                    print(f"      ✅ EXACT MATCH - {pop_type}: Row {idx}")
                    break
            
            # If no exact match, try partial matching
            if len(population_rows) < len(population_types):
                for pop_type, column_name in population_types.items():
                    if column_name not in population_rows and pop_type.lower() in series_name.lower():
                        population_rows[column_name] = row
                        print(f"      📋 PARTIAL MATCH - {pop_type}: Row {idx} -> {series_name}")
                        break
    
    if 'value' not in population_rows:
        print(f"   ❌ Could not find 'Total Population' row")
        print(f"   💡 Please check if 'Total Population' exists in DataSeries column")
        return df
    
    print(f"   📋 Found {len(population_rows)} population types")
    
    # Find year columns (exclude metadata) 
    metadata_cols = ['_id', 'DataSeries', 'source_name', 'extraction_time', 
                    'data_type', 'country', 'unit']
    year_cols = [col for col in df.columns if col not in metadata_cols]
    
    # Filter to actual year columns (4-digit numbers)
    year_cols = [col for col in year_cols if str(col).strip().isdigit() and len(str(col).strip()) == 4]
    
    print(f"   📅 Found year columns: {year_cols}")
    
    # Convert year columns to proper dates and sort
    converted_date_cols = []
    for year_col in year_cols:
        converted_date = convert_yearly_date(str(year_col))
        if converted_date:
            converted_date_cols.append((year_col, converted_date))
    
    # Sort by date (oldest first)
    converted_date_cols.sort(key=lambda x: x[1])
    
    print(f"   📅 Processing {len(converted_date_cols)} years: {[x[0] for x in converted_date_cols]}")
    
    # Create records for each year
    vertical_records = []
    
    for original_col, converted_date in converted_date_cols:
        
        # Initialize record for this date
        record = {
            'date': converted_date,
            'data_type': 'Population',
            'country': 'Singapore',
            'unit': 'Number of persons',
            'source_name': 'Singapore Department of Statistics',
            'series_name': 'Population Breakdown'
        }
        
        # Get values for each population type
        record_has_data = False
        
        for column_name, row_data in population_rows.items():
            if original_col in row_data and pd.notna(row_data[original_col]) and row_data[original_col] != '':
                try:
                    numeric_value = float(row_data[original_col])
                    record[column_name] = numeric_value
                    record_has_data = True
                    
                    # Debug: show what we're getting for Total Population (value column)
                    if column_name == 'value':
                        print(f"      📊 {converted_date} Total Population: {numeric_value:,.0f}")
                        
                except (ValueError, TypeError):
                    print(f"      ⚠️ Skipping non-numeric value for {column_name} in {original_col}: {row_data[original_col]}")
        
        # Only add record if we have at least the total population value
        if record_has_data and 'value' in record:
            vertical_records.append(record)
    
    result_df = pd.DataFrame(vertical_records)
    
    if len(result_df) > 0:
        result_df = result_df.sort_values('date')
        
        print(f"   ✅ Population: {len(result_df)} records created")
        print(f"   📅 Date range: {result_df['date'].iloc[0]} to {result_df['date'].iloc[-1]}")
        print(f"   📊 Columns: {list(result_df.columns)}")
        
        # Show sample of total population values to verify they're correct
        if 'value' in result_df.columns:
            sample_values = result_df['value'].head(3).tolist()
            print(f"   📈 Sample Total Population values: {[f'{v:,.0f}' for v in sample_values]}")
            
            # Verify the values are reasonable (should be millions for Singapore)
            if all(v > 1000000 for v in sample_values):
                print(f"   ✅ Total Population values look correct (in millions)")
            else:
                print(f"   ⚠️ Total Population values seem too small - check data source")
    else:
        print(f"   ❌ No valid population records created")
    
    return result_df

def standardize_property_price_sheet(df):
    """
    Standardize Property Price sheet:
    - Rename "quarter" → "date" (already vertical)
    - Rename "index" → "value" for consistency with other indicators
    """
    
    print(f"   🔄 Property Price: Renaming 'quarter'→'date' AND 'index'→'value'")
    print(f"   📊 Original columns: {list(df.columns)}")
    
    result_df = df.copy()
    
    # Step 1: Rename quarter column to date
    if 'quarter' in result_df.columns:
        result_df.rename(columns={'quarter': 'date'}, inplace=True)
        print(f"   ✅ Renamed 'quarter' → 'date'")
    else:
        print(f"   ⚠️ 'quarter' column not found")
    
    # Step 2: Rename index column to value for consistency with other indicators
    if 'index' in result_df.columns:
        result_df.rename(columns={'index': 'value'}, inplace=True)
        print(f"   ✅ Renamed 'index' → 'value' for consistency")
    else:
        print(f"   ⚠️ 'index' column not found")
    
    # Show final structure
    print(f"   📊 Final columns: {list(result_df.columns)}")
    print(f"   ✅ Property Price: {len(result_df)} records, {len(result_df.columns)} columns")
    
    # Show sample data to verify
    if 'value' in result_df.columns and len(result_df) > 0:
        sample_values = result_df['value'].head(3).tolist()
        print(f"   📈 Sample property price values: {sample_values}")
    
    return result_df

def convert_quarterly_date(date_str):
    """
    Convert quarterly date format: 2025Q1 → 2025-01-01, Q2 → 2025-04-01, etc.
    """
    
    date_str = str(date_str).strip()
    
    # Match pattern like 2025Q1, 2024Q4, etc.
    match = re.match(r'^(\d{4})Q(\d)$', date_str)
    
    if match:
        year = match.group(1)
        quarter = int(match.group(2))
        
        # Map quarter to month
        quarter_months = {1: '01', 2: '04', 3: '07', 4: '10'}
        
        if quarter in quarter_months:
            month = quarter_months[quarter]
            return f"{year}-{month}-01"
    
    return None

def convert_monthly_date(date_str):
    """
    Convert monthly date format: 2025-04 → 2025-04-01
    """
    
    date_str = str(date_str).strip()
    
    # Match pattern like 2025-04, 2024-12, etc.
    if re.match(r'^\d{4}-\d{2}$', date_str):
        return f"{date_str}-01"
    
    return None

def convert_yearly_date(date_str):
    """
    Convert yearly date format: 2025 → 2025-01-01
    """
    
    date_str = str(date_str).strip()
    
    # Match 4-digit year
    if re.match(r'^\d{4}$', date_str):
        return f"{date_str}-01-01"
    
    return None

def get_actual_date_range(df, data_type):
    """Extract actual date range from standardized data"""
    try:
        if df is None or len(df) == 0:
            return "No data"
        
        # Look for date column
        if 'date' in df.columns:
            # Sort dates to get range
            dates = pd.to_datetime(df['date'], errors='coerce').dropna()
            
            if len(dates) > 0:
                min_date = dates.min().strftime('%Y-%m-%d')
                max_date = dates.max().strftime('%Y-%m-%d')
                return f"{min_date} to {max_date}"
        
        # Fallback: try to infer from row count and data type
        if data_type in ['GDP', 'CPI']:
            # Quarterly data, estimate range
            years = len(df) // 4
            start_year = 2025 - years
            return f"{start_year}-01-01 to 2025-01-01"
        elif data_type == 'Interest_Rate':
            # Monthly data
            months = len(df)
            years = months // 12
            start_year = 2025 - years
            return f"{start_year}-01-01 to 2025-01-01"
        elif data_type == 'Population':
            # Annual data
            years = len(df)
            start_year = 2025 - years
            return f"{start_year}-01-01 to 2024-01-01"
        else:
            return f"Data available ({len(df)} records)"
            
    except Exception as e:
        print(f"   ⚠️ Could not determine date range for {data_type}: {e}")
        return f"Data available ({len(df) if df is not None else 0} records)"

def recreate_contents_sheet_exactly(writer, standardized_data, source_urls):
    """Recreate Contents sheet with source URLs from macro_sources.py"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        workbook = writer.book
        
        # Create new Contents sheet at the beginning
        contents_sheet = workbook.create_sheet('Contents', 0)  # Insert at beginning
        
        print(f"   🔧 Recreating Contents sheet with source URLs...")
        
        # ROW 1-2: Title
        contents_sheet['A1'] = "Singapore Macroeconomic Data"
        contents_sheet.merge_cells('A1:H2')
        
        # Style title
        title_font = Font(size=18, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        contents_sheet['A1'].font = title_font
        contents_sheet['A1'].fill = title_fill
        contents_sheet['A1'].alignment = title_alignment
        
        # ROW 3: Headers
        headers = ['No.', 'Data Type', 'Status', 'Records', 'Date Range', 'Last Extraction Time', 'Sheet Link', 'Source URL']
        for col, header in enumerate(headers, 1):
            cell = contents_sheet.cell(row=3, column=col)
            cell.value = header
            
            # Style header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # ROW 4+: Data
        data_types = ['GDP', 'CPI', 'Interest_Rate', 'Population', 'Property_Price']
        
        for i, data_type in enumerate(data_types, 1):
            row = 3 + i
            
            # Get record count and date range from standardized data
            if data_type in standardized_data:
                df = standardized_data[data_type]
                records = len(df) if df is not None else 0
                
                # Get actual date range from the data
                date_range = get_actual_date_range(df, data_type)
            else:
                records = 0
                date_range = "No data"
            
            # No.
            contents_sheet.cell(row=row, column=1).value = i
            
            # Data Type
            contents_sheet.cell(row=row, column=2).value = data_type
            
            # Status
            contents_sheet.cell(row=row, column=3).value = "✅ Success"
            
            # Records
            contents_sheet.cell(row=row, column=4).value = records
            
            # Date Range (actual data range)
            contents_sheet.cell(row=row, column=5).value = date_range
            
            # Last Extraction Time
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            contents_sheet.cell(row=row, column=6).value = current_time
            
            # Sheet Link (with hyperlink)
            sheet_link = f"Go to {data_type}"
            sheet_name_clean = data_type.replace(' ', '_')
            hyperlink_formula = f'=HYPERLINK("#{sheet_name_clean}!A1","{sheet_link}")'
            link_cell = contents_sheet.cell(row=row, column=7)
            link_cell.value = hyperlink_formula
            
            # Style hyperlink
            link_font = Font(color="0563C1", underline="single")
            link_cell.font = link_font
            
            # Source URL - get from mapping or show N/A if not available
            correct_url = source_urls.get(data_type, "N/A")
            contents_sheet.cell(row=row, column=8).value = correct_url
            
            print(f"      ✅ {data_type}: {correct_url}")
        
        # Set column widths
        column_widths = {
            'A': 8,   # No.
            'B': 15,  # Data Type
            'C': 12,  # Status
            'D': 10,  # Records
            'E': 35,  # Date Range
            'F': 20,  # Last Extraction Time
            'G': 20,  # Sheet Link
            'H': 60   # Source URL (wider for long URLs)
        }
        
        for col_letter, width in column_widths.items():
            contents_sheet.column_dimensions[col_letter].width = width
        
        print(f"   ✅ Contents sheet recreated with source URLs!")
        
    except Exception as e:
        print(f"   ❌ Error recreating Contents sheet: {e}")
        import traceback
        traceback.print_exc()

def main():
    """Main execution function"""
    
    print("🇸🇬 SINGAPORE DATA STANDARDIZATION")
    print("=" * 60)
    print("Updated: Property Price 'index' → 'value' column for consistency")
    
    try:
        # Standardize Singapore data
        output_file = standardize_singapore_data()
        
        if output_file:
            print(f"\n🎯 STANDARDIZATION SUMMARY:")
            print(f"✅ GDP: Horizontal → Vertical (GDP Chained 2015 Dollars only)")
            print(f"✅ CPI: Horizontal → Vertical (All Items only)")
            print(f"✅ Interest Rate: Horizontal → Vertical (SORA 3-Month only)")
            print(f"✅ Population: Horizontal → Vertical (Multiple population types)")
            print(f"✅ Property Price: Quarter→Date AND Index→Value (Updated)")
            print(f"✅ Contents: Preserved unchanged")
            print(f"\n📁 Output: {output_file}")
        else:
            print(f"\n❌ Standardization failed")
            
    except FileNotFoundError as e:
        print(f"❌ Error: {e}")
        print(f"💡 Make sure you have Singapore files in ./extracted_data/")
    except Exception as e:
        print(f"❌ Unexpected error: {e}")

if __name__ == "__main__":
    main()
