"""
FRED Data Cleaner - Population Unit Conversion Only
Clean script for US, Euro Area, and Japan macroeconomic data from FRED
Converts only population units (thousands/millions → persons) to avoid currency issues
"""

import pandas as pd
import numpy as np
from pathlib import Path
import re
from datetime import datetime
import warnings
import time
from typing import Dict, List, Any, Optional

# Suppress pandas performance warnings
warnings.filterwarnings('ignore', category=pd.errors.PerformanceWarning)

def find_latest_fred_files(data_dir: str = "./extracted_data"):
    """
    Find the latest FRED data files for US, Euro Area, and Japan
    Returns dict with country names as keys and file paths as values
    """
    
    data_path = Path(data_dir)
    if not data_path.exists():
        raise FileNotFoundError(f"Data directory not found: {data_dir}")
    
    # Define file patterns for each country
    patterns = {
        'China': 'macro_data_china_*.xlsx',          # NEW
        'US': 'macro_data_us_*.xlsx',
        'Euro Area': 'macro_data_euro area_*.xlsx', 
        'Japan': 'macro_data_japan_*.xlsx',
        'UK': 'macro_data_uk_*.xlsx',              # NEW
        'India': 'macro_data_india_*.xlsx',        # NEW
        'Indonesia': 'macro_data_indonesia_*.xlsx' # NEW
    }
    
    latest_files = {}
    
    for country, pattern in patterns.items():
        files = list(data_path.glob(pattern))
        if files:
            # Sort by modification time (most recent first)
            latest_file = max(files, key=lambda x: x.stat().st_mtime)
            latest_files[country] = latest_file
            print(f"🔍 {country}: Found {len(files)} files, latest is {latest_file.name}")
        else:
            print(f"⚠️ {country}: No files found with pattern {pattern}")
    
    return latest_files

def analyze_population_frequency(df, country):
    """Analyze the frequency of population data"""
    
    if df.empty or 'date' not in df.columns:
        return "No data", 0
    
    # Convert dates and sort
    df['date'] = pd.to_datetime(df['date'])
    df_sorted = df.sort_values('date')
    
    # Calculate date differences
    date_diffs = df_sorted['date'].diff().dropna()
    
    if len(date_diffs) == 0:
        return "Single record", 1
    
    # Get median difference
    median_diff = date_diffs.median()
    
    if median_diff <= pd.Timedelta(days=32):
        return "Monthly", len(df)
    elif median_diff <= pd.Timedelta(days=100):
        return "Quarterly", len(df)
    else:
        return "Annual", len(df)

def standardize_population_to_annual(df, country):
    """Convert population data to annual format (January 1st values)"""
    
    if df.empty or 'date' not in df.columns:
        return df
    
    print(f"   📅 Converting {country} population to annual format...")
    
    # Convert dates and sort
    df['date'] = pd.to_datetime(df['date'])
    df_sorted = df.sort_values('date')
    
    # Extract year from date
    df_sorted['year'] = df_sorted['date'].dt.year
    
    # For each year, try to get January 1st value, or closest available
    annual_records = []
    
    for year in df_sorted['year'].unique():
        year_data = df_sorted[df_sorted['year'] == year]
        
        # Try to find January 1st data
        jan_1_data = year_data[year_data['date'].dt.strftime('%m-%d') == '01-01']
        
        if not jan_1_data.empty:
            # Perfect - we have January 1st data
            record = jan_1_data.iloc[0].copy()
        else:
            # Take the first available record for that year
            record = year_data.iloc[0].copy()
        
        # Standardize date to January 1st of that year
        record['date'] = f"{year}-01-01"
        annual_records.append(record)
    
    # Create annual DataFrame
    annual_df = pd.DataFrame(annual_records)
    annual_df = annual_df.drop('year', axis=1, errors='ignore')  # Remove helper column
    
    print(f"   ✅ {country}: {len(df_sorted)} records → {len(annual_df)} annual records")
    print(f"   📅 Annual range: {annual_df['date'].min()} to {annual_df['date'].max()}")
    
    return annual_df

def convert_population_units(value, unit, data_type):
    """
    Convert population values with thousands, millions, billions units to base units
    Only applies to Population data type to avoid currency comparison issues
    Returns: (converted_value, updated_unit)
    """
    
    if pd.isna(value) or pd.isna(unit) or pd.isna(data_type):
        return value, unit
    
    # Only convert if it's population data
    if str(data_type).lower() != 'population':
        return value, unit
    
    unit_str = str(unit).lower()
    
    # Check for thousands
    if 'thousand' in unit_str:
        converted_value = value * 1000
        # Update unit description to remove "thousands"
        new_unit = re.sub(r'\bthousands?\b', '', unit_str, flags=re.IGNORECASE).strip()
        new_unit = re.sub(r'\s+', ' ', new_unit)  # Clean up extra spaces
        if not new_unit or new_unit.lower() in ['of', '']:
            new_unit = "Persons"
        return converted_value, new_unit.title()
    
    # Check for millions
    elif 'million' in unit_str:
        converted_value = value * 1000000
        new_unit = re.sub(r'\bmillions?\b', '', unit_str, flags=re.IGNORECASE).strip()
        new_unit = re.sub(r'\s+', ' ', new_unit)
        if not new_unit or new_unit.lower() in ['of', '']:
            new_unit = "Persons"
        return converted_value, new_unit.title()
    
    # Check for billions
    elif 'billion' in unit_str:
        converted_value = value * 1000000000
        new_unit = re.sub(r'\bbillions?\b', '', unit_str, flags=re.IGNORECASE).strip()
        new_unit = re.sub(r'\s+', ' ', new_unit)
        if not new_unit or new_unit.lower() in ['of', '']:
            new_unit = "Persons"
        return converted_value, new_unit.title()
    
    # No conversion needed
    return value, unit

def copy_source_urls_from_original(input_file_path: str) -> Dict[str, str]:
    """Copy source URLs from original file's Contents sheet"""
    source_urls = {}
    
    try:
        # Read the original Contents sheet - skiprows=2 to skip merged title rows
        contents_df = pd.read_excel(input_file_path, sheet_name='Contents', skiprows=2)
        
        print(f"   📋 Reading URLs from original Contents sheet...")
        print(f"   Available columns: {list(contents_df.columns)}")
        
        # Find the data type and URL columns flexibly
        data_type_col = None
        url_col = None
        
        for col in contents_df.columns:
            col_str = str(col).lower().strip()
            if 'data type' in col_str or 'data_type' in col_str:
                data_type_col = col
                print(f"   Found data type column: {col}")
            elif 'source url' in col_str or 'url' in col_str:
                url_col = col
                print(f"   Found URL column: {col}")
        
        if data_type_col and url_col:
            for idx, row in contents_df.iterrows():
                data_type = row[data_type_col]
                url = row[url_col]
                
                if pd.notna(data_type) and pd.notna(url) and str(url) not in ['N/A', 'nan', '']:
                    source_urls[str(data_type)] = str(url)
                    print(f"   📎 {data_type} -> {url}")
        
        print(f"   ✅ Copied {len(source_urls)} URLs from original file")
        
    except Exception as e:
        print(f"   ⚠️ Could not read URLs from original file: {e}")
    
    return source_urls

def clean_fred_data_2000_plus(input_file_path: str, country_name: str) -> str:
    """
    Clean FRED data: Filter to 2000+, remove unnecessary columns, convert population units only
    Process data first, then update Contents sheet with correct date ranges
    """
    
    input_path = Path(input_file_path)
    output_file_path = input_path.parent / f"cleaned_{input_path.name}"
    
    print(f"\n🧹 CLEANING {country_name.upper()} DATA (2000+ & POPULATION CONVERSION)")
    print("=" * 65)
    print(f"Input:  {input_path.name}")
    print(f"Output: {output_file_path.name}")
    
    original_source_urls = copy_source_urls_from_original(input_file_path)
    
    # Read all sheets
    excel_data = pd.read_excel(input_file_path, sheet_name=None)
    cleaned_data = {}
    
    # STEP 1: Process all data sheets first (except Contents)
    print(f"\n📊 STEP 1: Processing data sheets...")
    for sheet_name, df in excel_data.items():
        if sheet_name != 'Contents':
            print(f"\n🔄 Processing sheet: {sheet_name}")
            cleaned_df = clean_fred_data_sheet(df, sheet_name, country_name)
            if cleaned_df is not None and len(cleaned_df) > 0:
                cleaned_data[sheet_name] = cleaned_df
                print(f"   ✅ {sheet_name}: {len(cleaned_df)} records after processing")
            else:
                print(f"   ❌ {sheet_name}: No data after filtering - sheet will be empty")
                # Create empty dataframe with proper columns to maintain sheet structure
                cleaned_data[sheet_name] = pd.DataFrame(columns=['source_name', 'series_id', 'country', 'data_type', 'date', 'value', 'unit', 'currency', 'extraction_time'])
    
    # STEP 2: Write cleaned data to Excel first (without Contents sheet)
    print(f"\n💾 STEP 2: Writing cleaned data to Excel...")
    
    # Handle existing file
    if output_file_path.exists():
        print(f"   📝 File exists: {output_file_path.name}")
        try:
            output_file_path.unlink()
            print(f"   🗑️ Deleted existing file")
        except PermissionError:
            print(f"   ⚠️ File is open - using timestamp suffix")
            timestamp = datetime.now().strftime('%H%M%S')
            new_name = output_file_path.stem + f"_{timestamp}" + output_file_path.suffix
            output_file_path = output_file_path.parent / new_name
    
    # Write Excel file with data sheets only
    try:
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            # Write data sheets
            for sheet_name, df in cleaned_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                apply_basic_formatting(writer, sheet_name, df)
        
        print(f"   ✅ Data sheets written successfully")
        
        # STEP 3: Create perfect Contents sheet by reading the actual written data
        print(f"\n📋 STEP 3: Creating Contents sheet from actual written data...")
        apply_exact_contents_formatting(output_file_path, country_name, original_source_urls)
        
        print(f"   ✅ File written successfully: {output_file_path.name}")
        
    except Exception as e:
        print(f"   ❌ Write error: {e}")
        raise
    
    print(f"\n🎉 {country_name} cleaning completed!")
    print(f"📁 Output saved: {output_file_path}")
    return str(output_file_path)

def clean_fred_data_sheet(df: pd.DataFrame, sheet_name: str, country_name: str) -> pd.DataFrame:
    """
    Clean individual FRED data sheet:
    1. Standardize population frequency to annual (if needed)
    2. Filter dates to 2000+
    3. Apply population unit conversions only (thousands, millions, billions)
    4. Remove unnecessary columns
    5. Keep essential columns in proper order
    6. Sort by date ascending (oldest first)
    """
    
    if df is None or len(df) == 0:
        return df
    
    print(f"   📅 Original records: {len(df)}")
    
    # STEP 1: Standardize population frequency to annual (if needed)
    if sheet_name == 'Population':
        frequency, record_count = analyze_population_frequency(df, country_name)
        print(f"   📊 Population frequency: {frequency} ({record_count} records)")
        
        if frequency in ["Monthly", "Quarterly"]:
            df = standardize_population_to_annual(df, country_name)
        else:
            print(f"   ✅ Population already annual format")
            # Still standardize dates to January 1st format for consistency
            df = df.copy()
            df['date'] = pd.to_datetime(df['date'])
            df['date'] = df['date'].dt.year.astype(str) + '-01-01'
    
    # STEP 2: Filter dates to 2000+
    df_filtered = df[df['date'] >= '2000-01-01'].copy()
    print(f"   📅 After 2000+ filter: {len(df_filtered)}")
    
    if len(df_filtered) == 0:
        return None
    
    # STEP 3: Apply population unit conversions only
    if 'unit' in df_filtered.columns and 'value' in df_filtered.columns and 'data_type' in df_filtered.columns:
        print(f"   🔄 Checking for population unit conversions...")
        
        conversions_applied = 0
        unit_changes = {}
        
        for idx, row in df_filtered.iterrows():
            original_value = row['value']
            original_unit = row['unit']
            data_type = row['data_type']
            
            converted_value, new_unit = convert_population_units(original_value, original_unit, data_type)
            
            if converted_value != original_value:
                df_filtered.at[idx, 'value'] = converted_value
                df_filtered.at[idx, 'unit'] = new_unit
                conversions_applied += 1
                
                # Track unit changes for reporting
                conversion_key = f"{original_unit} → {new_unit}"
                if conversion_key not in unit_changes:
                    unit_changes[conversion_key] = 0
                unit_changes[conversion_key] += 1
        
        if conversions_applied > 0:
            print(f"   👥 Applied {conversions_applied} population unit conversions:")
            for change, count in unit_changes.items():
                print(f"      • {count} records: {change}")
        else:
            print(f"   ✅ No population unit conversions needed")
    
    # STEP 4: Remove unwanted columns
    columns_to_remove = ['records_removed', 'cleaning_applied']
    for col in columns_to_remove:
        if col in df_filtered.columns:
            df_filtered = df_filtered.drop(columns=[col])
            print(f"   🗑️ Removed column: {col}")
    
    # STEP 5: Ensure proper column order
    desired_order = [
        'source_name', 'series_id', 'country', 'data_type', 
        'date', 'value', 'unit', 'currency', 'extraction_time'
    ]
    
    # Keep columns in desired order (if they exist)
    final_columns = []
    for col in desired_order:
        if col in df_filtered.columns:
            final_columns.append(col)
    
    # Add any remaining columns
    for col in df_filtered.columns:
        if col not in final_columns:
            final_columns.append(col)
    
    df_clean = df_filtered[final_columns].copy()
    
    # STEP 6: Sort by date ascending (oldest first, newest last)
    df_clean = df_clean.sort_values('date', ascending=True)
    
    # Show final statistics
    if len(df_clean) > 0:
        min_date = df_clean['date'].min()
        max_date = df_clean['date'].max()
        print(f"   📅 Final date range: {min_date} to {max_date}")
        
        # Show sample values for verification (only for population)
        if 'value' in df_clean.columns and sheet_name == 'Population':
            sample_values = df_clean['value'].head(3).tolist()
            latest_values = df_clean['value'].tail(3).tolist()
            print(f"   👥 Sample early population: {[f'{v:,.0f}' for v in sample_values]}")
            print(f"   👥 Sample latest population: {[f'{v:,.0f}' for v in latest_values]}")
    
    return df_clean

def apply_basic_formatting(writer, sheet_name: str, df: pd.DataFrame):
    """Apply basic column width formatting to data sheets"""
    try:
        from openpyxl.utils import get_column_letter
        
        workbook = writer.book
        worksheet = workbook[sheet_name]
        
        # Auto-adjust column widths
        for col_num, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_num)
            
            # Calculate reasonable width
            max_length = len(str(column))
            for value in df[column].astype(str):
                max_length = max(max_length, len(str(value)))
            
            # Set width with reasonable limits
            if column == 'source_name':
                width = min(max(max_length + 2, 25), 50)
            elif column == 'unit':
                width = min(max(max_length + 2, 30), 60)
            elif column == 'value':
                width = min(max(max_length + 2, 15), 30)  # Wider for large numbers
            else:
                width = min(max(max_length + 2, 10), 40)
            
            worksheet.column_dimensions[column_letter].width = width
        
    except Exception as e:
        print(f"   ⚠️ Could not apply formatting to {sheet_name}: {e}")

def apply_exact_contents_formatting(file_path: Path, country_name: str, original_urls: Dict[str, str] = None):
    """Apply exact Contents sheet formatting - DYNAMIC headers and sheets"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        workbook = load_workbook(file_path)
        
        # Get list of all data sheets (excluding Contents) - DYNAMIC, not hardcoded
        data_sheets = [sheet for sheet in workbook.sheetnames if sheet != 'Contents']
        
        # Read headers from original file's Contents sheet - TRULY DYNAMIC
        headers = []
        original_file_path = str(file_path).replace('cleaned_', '')  # Try to find original file
        
        try:
            # Try to read headers from the original file
            if Path(original_file_path).exists():
                print(f"   📋 Reading headers from original file: {Path(original_file_path).name}")
                original_contents = pd.read_excel(original_file_path, sheet_name='Contents', skiprows=2, nrows=1)
                headers = list(original_contents.columns)
                print(f"   ✅ Found headers: {headers}")
            else:
                print(f"   ⚠️ Original file not found: {original_file_path}")
                raise FileNotFoundError("Original file not found")
        except Exception as e:
            print(f"   ⚠️ Could not read headers from original file: {e}")
            print(f"   📋 Using fallback standard headers")
            headers = ['No.', 'Data Type', 'Status', 'Records', 'Date Range', 'Last Extraction Time', 'Sheet Link', 'Source URL']
        
        # Delete the old Contents sheet and create new one
        if 'Contents' in workbook.sheetnames:
            del workbook['Contents']
        contents_sheet = workbook.create_sheet('Contents', 0)  # Insert at beginning
        
        print(f"   🔧 Recreating Contents sheet from scratch...")
        print(f"   📊 Found {len(data_sheets)} data sheets: {data_sheets}")
        
        # ROW 1-2: Title (merged) - dynamic based on number of columns
        num_cols = len(headers)
        title_range = f'A1:{get_column_letter(num_cols)}2'
        contents_sheet['A1'] = f"{country_name} Macroeconomic Data (Annual Population + Unit Converted)"
        contents_sheet.merge_cells(title_range)
        
        # Style title exactly like original
        title_font = Font(size=18, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        contents_sheet['A1'].font = title_font
        contents_sheet['A1'].fill = title_fill
        contents_sheet['A1'].alignment = title_alignment
        
        # ROW 3: Headers - DYNAMIC based on actual headers
        for col, header in enumerate(headers, 1):
            cell = contents_sheet.cell(row=3, column=col)
            cell.value = header
            
            # Style header exactly like original
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # ROW 4+: Create data rows for each ACTUAL sheet (dynamic)
        for i, sheet_name in enumerate(data_sheets, 1):
            row = 3 + i
            
            # Get data from the actual sheet to calculate metrics
            try:
                sheet_data = workbook[sheet_name]
                record_count = sheet_data.max_row - 1 if sheet_data.max_row > 1 else 0
                
                # Try to get date range from the sheet data
                date_range = "N/A"
                if record_count > 0:
                    # Find the date column dynamically
                    date_column = None
                    for col in range(1, sheet_data.max_column + 1):
                        header_cell = sheet_data.cell(row=1, column=col)
                        if header_cell.value and 'date' in str(header_cell.value).lower():
                            date_column = col
                            break

                    if date_column:
                        first_date = None
                        last_date = None
                        for row_num in range(2, min(sheet_data.max_row + 1, 12)):
                            date_cell = sheet_data.cell(row=row_num, column=date_column)
                            if date_cell.value:
                                if first_date is None:
                                    first_date = str(date_cell.value)
                                last_date = str(date_cell.value)
                    
                    # Get last date from bottom of sheet
                    if sheet_data.max_row > 1 and date_column:
                        last_date_cell = sheet_data.cell(row=sheet_data.max_row, column=date_column)
                        if last_date_cell.value:
                            last_date = str(last_date_cell.value)
                    
                    if first_date and last_date:
                        date_range = f"{first_date} to {last_date}"
                
            except Exception as e:
                print(f"   ⚠️ Could not read data from {sheet_name}: {e}")
                record_count = 0
                date_range = "N/A"
            
            # Populate cells based on header positions - DYNAMIC
            for col, header in enumerate(headers, 1):
                cell = contents_sheet.cell(row=row, column=col)
                
                if header == 'No.':
                    cell.value = i
                elif header == 'Data Type':
                    cell.value = sheet_name
                elif header == 'Status':
                    cell.value = "✅ Success" if record_count > 0 else "❌ No Data"
                elif header == 'Records':
                    cell.value = record_count
                elif header == 'Date Range':
                    cell.value = date_range
                elif header == 'Last Extraction Time':
                    cell.value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                elif header == 'Sheet Link':
                    if record_count > 0:
                        sheet_name_clean = sheet_name.replace(' ', '_')
                        hyperlink_formula = f'=HYPERLINK("#{sheet_name_clean}!A1","Go to {sheet_name}")'
                        cell.value = hyperlink_formula
                        
                        # Style hyperlink exactly like original
                        link_font = Font(color="0563C1", underline="single")
                        cell.font = link_font
                    else:
                        cell.value = "No data available"
                elif header == 'Source URL':
                    # Source URL - get from copied URLs using actual sheet name
                    if original_urls:
                        correct_url = original_urls.get(sheet_name, "N/A")
                        print(f"      ✅ Using copied URL for {sheet_name}: {correct_url}")
                    else:
                        correct_url = "N/A"
                        print(f"      ⚠️ No copied URL found for {sheet_name}")
                    cell.value = correct_url
        
        # Set column widths - DYNAMIC based on header content
        # Default widths for common headers
        default_widths = {
            'No.': 8,
            'Data Type': 15,
            'Status': 12,
            'Records': 10,
            'Date Range': 35,
            'Last Extraction Time': 20,
            'Sheet Link': 20,
            'Source URL': 50
        }
        
        for col, header in enumerate(headers, 1):
            col_letter = get_column_letter(col)
            width = default_widths.get(header, 15)  # Default to 15 if header not in mapping
            contents_sheet.column_dimensions[col_letter].width = width
        
        # Save workbook
        workbook.save(file_path)
        print(f"   ✅ Contents sheet recreated with {len(data_sheets)} sheets and {len(headers)} columns!")
        
    except Exception as e:
        print(f"   ❌ Error recreating Contents: {e}")
        import traceback
        traceback.print_exc()

def analyze_cleaning_results(cleaned_files: dict):
    """Analyze cleaning results for all countries with population conversion info"""
    
    print(f"\n📊 CLEANING & POPULATION CONVERSION RESULTS")
    print("=" * 60)
    
    for country, file_path in cleaned_files.items():
        print(f"\n🌍 {country.upper()}:")
        
        try:
            excel_data = pd.read_excel(file_path, sheet_name=None)
            
            for sheet_name, df in excel_data.items():
                if sheet_name == 'Contents':
                    continue
                    
                if df is not None and len(df) > 0:
                    min_date = df['date'].min()
                    max_date = df['date'].max()
                    
                    # Show unit information
                    unit_info = ""
                    if 'unit' in df.columns:
                        unique_units = df['unit'].unique()
                        if len(unique_units) == 1:
                            unit_info = f" - Unit: {unique_units[0]}"
                        else:
                            unit_info = f" - Units: {', '.join(unique_units[:2])}"
                    
                    # Show value range for verification (highlight population)
                    value_info = ""
                    if 'value' in df.columns:
                        min_val = df['value'].min()
                        max_val = df['value'].max()
                        if sheet_name == 'Population':
                            value_info = f" - Population Range: {min_val:,.0f} to {max_val:,.0f} persons"
                        else:
                            value_info = f" - Range: {min_val:,.2f} to {max_val:,.2f}"
                    
                    print(f"   📋 {sheet_name}: {len(df)} records ({min_date} to {max_date}){unit_info}{value_info}")
                
        except Exception as e:
            print(f"   ❌ Error analyzing {country}: {e}")

def main():
    """
    Main function to clean all FRED data files with population standardization and unit conversion
    """
    
    print("🚀 FRED DATA CLEANER - POPULATION STANDARDIZATION + CONVERSION")
    print("=" * 70)
    print("📅 Converting population to annual format (January 1st values)")
    print("👥 Converting only population units (thousands/millions → persons)")
    print("💰 Preserving original GDP/economic units for proper currency handling")
    print("🇺🇸🇪🇺🇯🇵 Processing: US, Euro Area, Japan (FRED data only)")
    print("=" * 70)
    
    try:
        # Find latest files
        print(f"\n📂 Finding latest FRED files...")
        latest_files = find_latest_fred_files()
        
        if not latest_files:
            print("❌ No FRED data files found!")
            return
        
        cleaned_files = {}
        
        # Process each country
        for country, file_path in latest_files.items():
            try:
                cleaned_file = clean_fred_data_2000_plus(str(file_path), country)
                cleaned_files[country] = cleaned_file
            except Exception as e:
                print(f"❌ Error processing {country}: {e}")
        
        # Analyze results
        if cleaned_files:
            analyze_cleaning_results(cleaned_files)
            
            print(f"\n🎯 SUMMARY:")
            print(f"✅ Standardized population to annual format (January 1st)")
            print(f"✅ Filtered data to 2000+ only")
            print(f"✅ Applied population unit conversions only (thousands → persons)")
            print(f"✅ Preserved GDP/economic units for proper currency handling")
            print(f"✅ Removed unnecessary columns (records_removed, cleaning_applied)")
            print(f"✅ Updated Contents sheet with processing notes")
            print(f"✅ Processed {len(cleaned_files)} countries")
            
            print(f"\n📁 Cleaned files:")
            for country, file_path in cleaned_files.items():
                print(f"   • {country}: {Path(file_path).name}")
            
            print(f"\n💡 BENEFITS:")
            print(f"   • All countries now have consistent annual population data")
            print(f"   • US: Monthly → Annual (Jan 1st values)")
            print(f"   • Euro Area & Japan: Already annual (date standardized)")
            print(f"   • Cross-country population comparison now accurate")
            print(f"   • Ready for visualization with proper time alignment")
        
        return cleaned_files
        
    except Exception as e:
        print(f"❌ Main process error: {e}")
        return None

if __name__ == "__main__":
    main()
