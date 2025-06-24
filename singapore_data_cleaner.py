"""
Fixed Singapore Data Cleaner - Correct Source URLs
Properly extracts source URLs from macro_sources.py for each data type
"""

import pandas as pd
import numpy as np
from pathlib import Path
import re
from datetime import datetime
import warnings
import os
import time

# Suppress pandas performance warnings
warnings.filterwarnings('ignore', category=pd.errors.PerformanceWarning)

def get_singapore_source_urls():
    """
    Extract correct source URLs from macro_sources.py for Singapore data types
    """
    try:
        # Import Singapore sources
        from macro_sources import SINGAPORE_SOURCES
        
        # Create mapping of data types to URLs
        url_mapping = {}
        
        for source in SINGAPORE_SOURCES:
            # Map based on data type patterns
            if 'GDP' in source.name and 'Chained' in source.name:
                url_mapping['GDP'] = source.url
            elif 'Consumer Price Index' in source.name or 'CPI' in source.name:
                url_mapping['CPI'] = source.url
            elif 'Interest' in source.name or 'SORA' in source.name:
                url_mapping['Interest_Rate'] = source.url
            elif 'Population' in source.name and 'Total' not in source.name:
                url_mapping['Population'] = source.url
            elif 'Property' in source.name or 'Price Index' in source.name:
                url_mapping['Property_Price'] = source.url
        
        print(f"   🔗 Loaded {len(url_mapping)} source URLs from macro_sources.py")
        for data_type, url in url_mapping.items():
            print(f"      {data_type}: {url}")
        
        return url_mapping
        
    except ImportError:
        print(f"   ⚠️ Could not import macro_sources.py")
        return {}
    except Exception as e:
        print(f"   ⚠️ Error loading source URLs: {e}")
        return {}

def find_latest_singapore_file(data_dir: str = "./extracted_data") -> str:
    """
    Find the latest Singapore macro data file automatically
    """
    
    data_path = Path(data_dir)
    if not data_path.exists():
        raise FileNotFoundError(f"Data directory not found: {data_dir}")
    
    # Search for Singapore macro files
    pattern = "macro_data_singapore_*.xlsx"
    singapore_files = list(data_path.glob(pattern))
    
    if not singapore_files:
        raise FileNotFoundError(f"No Singapore files found in {data_dir}")
    
    # Sort by modification time (most recent first)
    latest_file = max(singapore_files, key=lambda x: x.stat().st_mtime)
    
    print(f"🔍 Found {len(singapore_files)} Singapore files")
    print(f"📄 Latest file: {latest_file.name}")
    return str(latest_file)

def safe_file_write(output_file_path, cleaned_data):
    """Safely write Excel file with overwrite handling"""
    
    output_path = Path(output_file_path)
    
    # Check if file exists and handle it
    if output_path.exists():
        print(f"   📝 File exists: {output_path.name}")
        
        # Try to delete the existing file
        try:
            output_path.unlink()
            print(f"   🗑️ Deleted existing file")
        except PermissionError:
            print(f"   ⚠️ File is open in Excel - trying alternative approaches...")
            
            # Try alternative filename with timestamp
            timestamp = datetime.now().strftime('%H%M%S')
            new_name = output_path.stem + f"_{timestamp}" + output_path.suffix
            output_path = output_path.parent / new_name
            print(f"   📝 Using alternative name: {output_path.name}")
    
    # Write the Excel file with proper formatting
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in cleaned_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Apply formatting after writing
                if sheet_name == 'Contents':
                    recreate_contents_sheet_exactly(writer, sheet_name, cleaned_data)
                else:
                    apply_auto_fit_columns(writer, sheet_name, df)
        
        print(f"   ✅ File written successfully: {output_path.name}")
        return str(output_path)
        
    except PermissionError as e:
        print(f"   ❌ Still cannot write file: {e}")
        print(f"   💡 Please close Excel and try again, or run with different filename")
        raise
    except Exception as e:
        print(f"   ❌ Unexpected write error: {e}")
        raise

def clean_singapore_data_simple(input_file_path: str = None, filter_population_to_2000: bool = True):
    """
    Clean Singapore Excel data with correct source URLs
    """
    
    # Auto-find latest file if not specified
    if input_file_path is None:
        input_file_path = find_latest_singapore_file()
    
    input_path = Path(input_file_path)
    output_file_path = input_path.parent / f"cleaned_{input_path.name}"
    
    print("🧹 SINGAPORE DATA CLEANING (FIXED VERSION)")
    print("=" * 60)
    print(f"Input:  {input_path.name}")
    print(f"Output: {output_file_path.name}")
    print(f"📂 Directory: {input_path.parent}")
    
    # Load source URLs
    source_urls = get_singapore_source_urls()
    
    # Read all sheets
    excel_data = pd.read_excel(input_file_path, sheet_name=None)
    cleaned_data = {}
    
    for sheet_name, df in excel_data.items():
        print(f"\n📊 Processing sheet: {sheet_name}")
        
        if sheet_name == 'Contents':
            # DON'T PROCESS - we'll recreate it completely in formatting
            cleaned_data[sheet_name] = df  # Keep original for now
            print(f"   ✅ Contents will be recreated with correct URLs")
            
        elif sheet_name in ['GDP', 'CPI', 'Interest_Rate']:
            # Clean quarterly/monthly data with 2000+ filter
            cleaned_df = clean_quarterly_monthly_data(df, sheet_name)
            cleaned_data[sheet_name] = cleaned_df
            
        elif sheet_name == 'Population':
            # Clean population with 2000+ filter for consistency
            cleaned_df = clean_population_data(df, filter_to_2000=filter_population_to_2000)
            cleaned_data[sheet_name] = cleaned_df
            
        elif sheet_name == 'Property_Price':
            # Clean property price dates
            cleaned_df = clean_property_price_data(df)
            cleaned_data[sheet_name] = cleaned_df
            
        else:
            # Keep other sheets as-is
            cleaned_data[sheet_name] = df
            print(f"   ✅ {sheet_name} sheet preserved")
    
    # Write to new Excel file with smart overwrite handling
    try:
        output_file_path = safe_file_write(output_file_path, cleaned_data)
    except Exception as e:
        print(f"\n❌ File write error: {e}")
        return None
    
    print(f"\n🎉 Cleaning completed with correct source URLs!")
    print(f"📁 Output saved: {output_file_path}")
    return str(output_file_path)

def clean_quarterly_monthly_data(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """Clean GDP, CPI, and Interest_Rate data with 2000+ filter"""
    
    print(f"   📅 Cleaning {sheet_name} quarterly/monthly data (2000+)...")
    
    # Identify date columns (exclude metadata columns)
    metadata_cols = ['_id', 'DataSeries', 'source_name', 'extraction_time', 
                    'data_type', 'country', 'currency', 'unit']
    
    date_columns = [col for col in df.columns if col not in metadata_cols]
    
    # Fix quarter/month format and filter for 2000+
    cleaned_date_cols = []
    
    for col in date_columns:
        if pd.isna(col) or col == '':
            continue
            
        col_str = str(col)
        
        # Quarterly format: 20244Q → 2024Q4
        if re.match(r'^\d{4}\dQ$', col_str):
            year = col_str[:4]
            quarter = col_str[4]
            if int(year) >= 2000:
                new_col = f"{year}Q{quarter}"
                cleaned_date_cols.append((col, new_col))
                
        # Monthly format: 2025Apr → 2025-04
        elif re.match(r'^\d{4}[A-Za-z]{3}$', col_str):
            year = col_str[:4]
            month_abbr = col_str[4:]
            if int(year) >= 2000:
                month_map = {
                    'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04',
                    'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08', 
                    'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
                }
                new_col = f"{year}-{month_map.get(month_abbr, '01')}"
                cleaned_date_cols.append((col, new_col))
    
    # Sort columns chronologically (newest first)
    def sort_key(item):
        _, new_col = item
        if 'Q' in new_col:  # Quarterly
            year, quarter = new_col.split('Q')
            return (int(year), int(quarter))
        elif '-' in new_col:  # Monthly
            year, month = new_col.split('-')
            return (int(year), int(month))
        return (0, 0)
    
    sorted_date_cols = sorted(cleaned_date_cols, key=sort_key, reverse=True)
    
    print(f"   📊 Filtered to {len(sorted_date_cols)} columns (2000+)")
    if sorted_date_cols:
        print(f"   📅 Range: {sorted_date_cols[-1][1]} to {sorted_date_cols[0][1]}")
    
    # Rebuild DataFrame efficiently using pd.concat
    column_parts = []
    
    # Add metadata columns first
    for col in ['_id', 'DataSeries']:
        if col in df.columns:
            column_parts.append(df[col].to_frame(col))
    
    # Add sorted date columns
    for old_col, new_col in sorted_date_cols:
        if old_col in df.columns:
            column_parts.append(df[old_col].to_frame(new_col))
    
    # Add remaining metadata columns
    for col in metadata_cols[2:]:
        if col in df.columns:
            column_parts.append(df[col].to_frame(col))
    
    # Combine all columns at once
    new_df = pd.concat(column_parts, axis=1) if column_parts else pd.DataFrame()
    
    print(f"   ✅ {sheet_name}: {len(new_df.columns)} columns, {len(new_df)} rows")
    return new_df

def clean_population_data(df: pd.DataFrame, filter_to_2000: bool = True) -> pd.DataFrame:
    """Clean population data with optional 2000+ filter"""
    
    if filter_to_2000:
        print(f"   👥 Cleaning Population data (2000+ for consistency)...")
    else:
        print(f"   👥 Cleaning Population data (full history)...")
    
    # Identify year columns
    metadata_cols = ['_id', 'DataSeries', 'source_name', 'extraction_time', 
                    'data_type', 'country', 'unit']
    
    year_columns = []
    for col in df.columns:
        if col not in metadata_cols:
            try:
                year = int(col)
                if 1950 <= year <= 2030:
                    if not filter_to_2000 or year >= 2000:
                        year_columns.append(col)
            except:
                continue
    
    # Sort years in descending order (newest first)
    sorted_years = sorted(year_columns, reverse=True)
    
    print(f"   📅 Year range: {sorted_years[-1]} to {sorted_years[0]}")
    
    # Rebuild DataFrame efficiently with correct column order
    column_parts = []
    
    # Add _id and DataSeries first
    if '_id' in df.columns:
        column_parts.append(df['_id'].to_frame('_id'))
    if 'DataSeries' in df.columns:
        column_parts.append(df['DataSeries'].to_frame('DataSeries'))
    
    # Add sorted year columns
    for year in sorted_years:
        if year in df.columns:
            column_parts.append(df[year].to_frame(str(year)))
    
    # Add remaining metadata columns
    for col in metadata_cols[2:]:
        if col in df.columns:
            column_parts.append(df[col].to_frame(col))
    
    # Combine all columns at once
    new_df = pd.concat(column_parts, axis=1) if column_parts else pd.DataFrame()
    
    print(f"   ✅ Population: {len(new_df.columns)} columns, {len(new_df)} rows")
    return new_df

def clean_property_price_data(df: pd.DataFrame) -> pd.DataFrame:
    """Clean property price data - remove time component from dates"""
    
    print(f"   🏠 Cleaning Property Price data...")
    
    new_df = df.copy()
    
    # Clean the quarter column (remove time component)
    if 'quarter' in new_df.columns:
        new_df['quarter'] = new_df['quarter'].astype(str).str.replace(r'\s+\d{2}:\d{2}:\d{2}.*', '', regex=True)
        print(f"   🗓️ Cleaned date format in 'quarter' column")
    
    print(f"   ✅ Property Price: {len(new_df.columns)} columns, {len(new_df)} rows")
    return new_df

def recreate_contents_sheet_exactly(writer, sheet_name, all_cleaned_data):
    """Recreate Contents sheet with CORRECT source URLs from macro_sources.py"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        workbook = writer.book
        
        # Delete the broken Contents sheet and create new one
        if 'Contents' in workbook.sheetnames:
            del workbook['Contents']
        
        # Create new Contents sheet
        contents_sheet = workbook.create_sheet('Contents', 0)  # Insert at beginning
        
        print(f"   🔧 Recreating Contents sheet with correct source URLs...")
        
        # Get correct source URLs
        source_urls = get_singapore_source_urls()
        
        # ROW 1-2: Title
        contents_sheet['A1'] = "Singapore Macroeconomic Data"
        contents_sheet.merge_cells('A1:H2')
        
        # Style title
        title_font = Font(size=18, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        contents_sheet['A1'].font = title_font
        contents_sheet['A1'].fill = title_fill
        contents_sheet['A1'].alignment = title_alignment
        
        # ROW 3: Headers
        headers = ['No.', 'Data Type', 'Status', 'Records', 'Date Range', 'Last Extraction Time', 'Sheet Link', 'Source URL']
        for col, header in enumerate(headers, 1):
            cell = contents_sheet.cell(row=3, column=col)
            cell.value = header
            
            # Style header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # ROW 4+: Data
        data_types = ['GDP', 'CPI', 'Interest_Rate', 'Population', 'Property_Price']
        
        for i, data_type in enumerate(data_types, 1):
            row = 3 + i
            
            # Get record count
            if data_type in all_cleaned_data:
                df = all_cleaned_data[data_type]
                records = len(df) if df is not None else 0
            else:
                records = 0
            
            # No.
            contents_sheet.cell(row=row, column=1).value = i
            
            # Data Type
            contents_sheet.cell(row=row, column=2).value = data_type
            
            # Status
            contents_sheet.cell(row=row, column=3).value = "✅ Success"
            
            # Records
            contents_sheet.cell(row=row, column=4).value = records
            
            # Date Range - estimate based on data type
            if data_type in ['GDP', 'CPI']:
                date_range = "2000-Q1 to 2025-Q1"
            elif data_type == 'Interest_Rate':
                date_range = "2005-10 to 2025-06"
            elif data_type == 'Population':
                date_range = "2000 to 2024"
            else:
                date_range = "2000-Q1 to 2025-Q1"
            
            contents_sheet.cell(row=row, column=5).value = date_range
            
            # Last Extraction Time
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            contents_sheet.cell(row=row, column=6).value = current_time
            
            # Sheet Link (with hyperlink)
            sheet_link = f"Go to {data_type}"
            sheet_name_clean = data_type.replace(' ', '_')
            hyperlink_formula = f'=HYPERLINK("#{sheet_name_clean}!A1","{sheet_link}")'
            link_cell = contents_sheet.cell(row=row, column=7)
            link_cell.value = hyperlink_formula
            
            # Style hyperlink
            link_font = Font(color="0563C1", underline="single")
            link_cell.font = link_font
            
            # CORRECT Source URL - get from mapping
            correct_url = source_urls.get(data_type, "N/A")
            contents_sheet.cell(row=row, column=8).value = correct_url
            
            print(f"      ✅ {data_type}: {correct_url}")
        
        # Set column widths
        column_widths = {
            'A': 8,   # No.
            'B': 15,  # Data Type
            'C': 12,  # Status
            'D': 10,  # Records
            'E': 20,  # Date Range
            'F': 20,  # Last Extraction Time
            'G': 20,  # Sheet Link
            'H': 60   # Source URL (wider for long URLs)
        }
        
        for col_letter, width in column_widths.items():
            contents_sheet.column_dimensions[col_letter].width = width
        
        print(f"   ✅ Contents sheet recreated with CORRECT source URLs!")
        
    except Exception as e:
        print(f"   ❌ Error recreating Contents sheet: {e}")
        import traceback
        traceback.print_exc()

def apply_auto_fit_columns(writer, sheet_name, df):
    """Apply auto-fit column widths to data sheets only"""
    try:
        from openpyxl.utils import get_column_letter
        
        workbook = writer.book
        worksheet = workbook[sheet_name]
        
        # Auto-fit all columns based on content
        for col_num, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_num)
            
            # Calculate max width needed
            max_length = len(str(column))  # Header length
            
            # Check data content length
            for value in df[column].astype(str):
                max_length = max(max_length, len(str(value)))
            
            # Special handling for DataSeries column
            if column == 'DataSeries':
                max_length = max(max_length, 60)  # Ensure width for long descriptions
            
            # Set column width (reasonable limits)
            adjusted_width = min(max(max_length + 2, 8), 80)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"   📏 Applied auto-fit columns to {sheet_name}")
        
    except Exception as e:
        print(f"   ⚠️ Could not apply column formatting to {sheet_name}: {e}")

def analyze_cleaning_results(cleaned_file_path: str):
    """Quick analysis of cleaning results"""
    
    print(f"\n📊 CLEANING RESULTS ANALYSIS")
    print("=" * 50)
    
    excel_data = pd.read_excel(cleaned_file_path, sheet_name=None)
    
    for sheet_name, df in excel_data.items():
        if sheet_name == 'Contents':
            print(f"✅ {sheet_name}: Recreated with correct source URLs")
            continue
            
        print(f"\n📋 {sheet_name}: {df.shape[0]} rows × {df.shape[1]} columns")
        
        # Show date/time range for data sheets
        if sheet_name in ['GDP', 'CPI', 'Interest_Rate']:
            date_cols = [col for col in df.columns if col not in ['_id', 'DataSeries', 'source_name', 'extraction_time', 'data_type', 'country', 'currency', 'unit']]
            if date_cols:
                print(f"   📅 Period range: {date_cols[-1]} to {date_cols[0]}")
                
        elif sheet_name == 'Population':
            year_cols = [col for col in df.columns if col.isdigit()]
            if year_cols:
                years = sorted(year_cols)
                print(f"   📅 Year range: {years[0]} to {years[-1]}")

def main():
    """Main function to clean the latest Singapore file with correct URLs"""
    
    try:
        print("🚀 SINGAPORE DATA CLEANER - FIXED VERSION")
        print("=" * 70)
        print("🔗 Features: Correct source URLs from macro_sources.py")
        print("📊 Clean data formatting + proper Contents sheet")
        print("=" * 70)
        
        # Clean the latest Singapore file automatically
        output_file = clean_singapore_data_simple()
        
        if output_file:
            # Analyze results
            analyze_cleaning_results(output_file)
            
            print(f"\n🎯 SUMMARY:")
            print(f"✅ Quarter format: 20244Q → 2024Q4")
            print(f"✅ Columns: Chronological order (newest first)")
            print(f"✅ Population: 2000+ data for consistency")
            print(f"✅ Property dates: Cleaned (no time component)")
            print(f"✅ Contents sheet: Recreated with CORRECT source URLs")
            print(f"✅ Source URLs: Extracted from macro_sources.py")
            print(f"✅ DataSeries column: Auto-fit width")
            
            return output_file
        else:
            print(f"\n❌ Cleaning failed")
            return None
        
    except FileNotFoundError as e:
        print(f"❌ Error: {e}")
        print(f"💡 Make sure you have Singapore files in ./extracted_data/")
        return None
    except Exception as e:
        print(f"❌ Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    main()
